from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Optional

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from st_aggrid import AgGrid, DataReturnMode, GridOptionsBuilder, JsCode

from controllers.presupuesto_ventas_controller import (
    cargar_excel_directo_presupuesto_ventas_ctrl,
    eliminar_carga_completa_presupuesto_ventas_ctrl,
    eliminar_registro_presupuesto_ventas_ctrl,
    guardar_presupuesto_ventas_batch_ctrl,
    insertar_presupuesto_ventas_linea_estatus_ctrl,
    obtener_cargas_presupuesto_ventas_ctrl,
    obtener_catalogo_clientes_pv_ctrl,
    obtener_catalogo_productos_pv_ctrl,
    obtener_existencias_productos_pv_ctrl,
    obtener_ordenes_compra_pendientes_pv_ctrl,
    obtener_presupuesto_ventas_ctrl,
    obtener_presupuesto_ventas_lineas_ctrl,
    obtener_presupuesto_ventas_lineas_pendientes_ctrl,
    obtener_ultimo_precio_venta_ctrl,
    obtener_ventas_reales_sae_pv_ctrl,
    registrar_carga_presupuesto_ventas_ctrl,
    upsert_presupuesto_ventas_linea_ctrl,
)
from views.modulo_presupuesto_ventas.presupuesto_compras_view import (
    _enviar_notificacion_vendedor_compras,
    _panel_carga as _panel_carga_compras,
    _panel_crear_manual as _panel_crear_manual_compras,
    _panel_gestionar_cargas as _panel_gestionar_cargas_compras,
    _panel_pivot as _panel_pivot_compras,
    _selector_carga as _selector_carga_compras,
)
from controllers.presupuesto_admin_controller import (
    obtener_presupuesto_ventas_compras_ctrl,
    obtener_roles_usuario_id_ctrl,
    obtener_usuario_por_id_ctrl,
    obtener_usuarios_presupuesto_ctrl,
)
from controllers.presupuesto_compras_controller import (
    insertar_presupuesto_compras_linea_estatus_ctrl,
    obtener_presupuesto_compras_ctrl,
    obtener_presupuesto_compras_lineas_pendientes_ctrl,
    upsert_presupuesto_compras_linea_ctrl,
)
from controllers.solicitudes_controller import get_correos_usuarios_por_rol_ctrl
from utils.envio_correo import enviar_correo


# ── constantes ────────────────────────────────────────────────────────────────

_MESES = {
    1: "ene", 2: "feb", 3: "mar", 4: "abr",
    5: "may", 6: "jun", 7: "jul", 8: "ago",
    9: "sep", 10: "oct", 11: "nov", 12: "dic",
}

_TABS_PIVOT = [
    ("KG México",        "KG",  "MEXICO"),
]

_COLS_ID = ["company", "cliente_excel", "codigo_origen", "producto_excel"]

# resalta en verde los valores positivos y en rojo los negativos; 0 sin color
_CELL_STYLE_VALORES = JsCode("""
function(params) {
    if (params.value > 0) {
        return {backgroundColor: '#d4edda', color: '#155724'};
    }
    if (params.value < 0) {
        return {backgroundColor: '#f8d7da', color: '#721c24'};
    }
    return null;
}
""")


def _value_formatter_js(decimales: int) -> JsCode:
    return JsCode(f"""
    function(params) {{
        if (params.value === null || params.value === undefined || params.value === '') return '';
        return Number(params.value).toFixed({decimales});
    }}
    """)


# ── helpers ───────────────────────────────────────────────────────────────────

def _get_usuario_id() -> int:
    usuario = st.session_state.get("usuario") or {}
    return int(usuario.get("id") or usuario.get("id_usuario") or 0)


def _norm_roles_list(values) -> list[str]:
    if not values:
        return []
    if isinstance(values, str):
        values = [values]
    return [str(v).strip().lower() for v in values if str(v or "").strip()]


def _tiene_rol(roles: list[str], *objetivos: str) -> bool:
    roles_set = set(_norm_roles_list(roles))
    objetivos_set = set(_norm_roles_list(objetivos))
    return bool(roles_set.intersection(objetivos_set))


def _puede_ver_todos_presupuesto() -> bool:
    usuario = st.session_state.get("usuario") or {}
    return _tiene_rol(usuario.get("roles"), "forecastadmin", "superadmin")


# ── autorización por línea ──────────────────────────────────────────────────
# mismo esquema que usa solicitudes de gastos (tab_solicitud_gastos_view.py:
# _tipo_autorizacion_solicitud / _rol_autorizador_solicitud): el rol de quien
# somete decide quién autoriza. "Gerente de Ventas" no requiere autorización;
# "Jefe de Ventas"/"Supervisor de Ventas" los autoriza el Gerente; cualquier
# otro perfil (p.ej. "Ventas") lo autoriza el Jefe de Ventas.

_ESTATUS_LINEA_BADGE = {
    "captura": "🔵 captura",
    "enviada": "🟡 enviada",
    "autorizada": "🟢 autorizada",
    "rechazada": "🔴 rechazada",
}

# colores de fondo para la columna "Autorización" en tablas de solo lectura
# (st.dataframe con pandas Styler) — mismos tonos que usa AgGrid en la
# tabla de captura para valores positivos/negativos
_COLOR_FONDO_AUTORIZACION = {
    "captura": "#cfe2ff",
    "enviada": "#fff3cd",
    "autorizada": "#d4edda",
    "rechazada": "#f8d7da",
}


def _color_fondo_autorizacion(valor: str) -> str:
    texto = str(valor or "")
    for estatus, color in _COLOR_FONDO_AUTORIZACION.items():
        if estatus in texto:
            return f"background-color: {color}; color: #111"
    return ""


# mismos tonos que usa AgGrid (_CELL_STYLE_VALORES) en la tabla de captura
# para resaltar valores positivos/negativos en las columnas de mes
def _color_valor_mes(valor) -> str:
    try:
        v = float(valor)
    except (TypeError, ValueError):
        return ""
    if v > 0:
        return "background-color: #d4edda; color: #155724"
    if v < 0:
        return "background-color: #f8d7da; color: #721c24"
    return ""


def _formatear_numeros(estilo, df_mostrado: pd.DataFrame, meses_cols: list[str]):
    """Aplica formato de miles/decimales a "Precio" y "Total USD Año" (con $
    y 2 decimales), "Total Kilos Año" (miles y 2 decimales) y a los meses
    (miles y 3 decimales, igual que en la tabla de presupuesto)."""
    for col in ("Precio", "Total USD Año"):
        if col in df_mostrado.columns:
            estilo = estilo.format("${:,.2f}", subset=[col])
    if "Total Kilos Año" in df_mostrado.columns:
        estilo = estilo.format("{:,.2f}", subset=["Total Kilos Año"])
    if meses_cols:
        estilo = estilo.format("{:,.3f}", subset=meses_cols)
    return estilo


def _tipo_autorizacion_linea() -> str:
    usuario = st.session_state.get("usuario") or {}
    roles = usuario.get("roles")
    if _tiene_rol(roles, "gerente de ventas", "gerente ventas"):
        return "sin_autorizacion"
    if _tiene_rol(roles, "jefe de ventas", "supervisor de ventas"):
        return "autoriza_gerente_ventas"
    return "autoriza_jefe_ventas"


def _rol_autorizador_linea() -> str:
    tipo = _tipo_autorizacion_linea()
    if tipo == "autoriza_gerente_ventas":
        return "Gerente de Ventas"
    if tipo == "autoriza_jefe_ventas":
        return "Jefe de Ventas"
    return ""


def _cambiar_estatus_linea_ventas(
    *,
    id_carga: int,
    company: Optional[str],
    cliente_excel: Optional[str],
    codigo_origen: Optional[str],
    producto_excel: str,
    estatus_nuevo: str,
    usuario_id: int,
    usuario_nombre: Optional[str],
    usuario_email: Optional[str],
    comentario: Optional[str] = None,
) -> int:
    linea_id, estatus_anterior = upsert_presupuesto_ventas_linea_ctrl(
        id_carga=id_carga,
        company=company,
        cliente_excel=cliente_excel,
        codigo_origen=codigo_origen,
        producto_excel=producto_excel,
        estatus=estatus_nuevo,
        usuario_id=usuario_id,
    )
    insertar_presupuesto_ventas_linea_estatus_ctrl(
        linea_id=linea_id,
        estatus_anterior=estatus_anterior,
        estatus_nuevo=estatus_nuevo,
        usuario_id=usuario_id,
        usuario_nombre=usuario_nombre,
        usuario_email=usuario_email,
        comentario=comentario,
    )
    return linea_id


def _enviar_notificacion_autorizador_ventas(
    *, lineas: list[dict], id_carga: int, anio: int, nombre_rol: str, token, remitente: str,
) -> tuple[bool, str]:
    destinatarios = sorted({
        str(e).strip() for e in (get_correos_usuarios_por_rol_ctrl(nombre_rol) or []) if str(e or "").strip()
    })
    if not destinatarios:
        return False, f"no hay correos configurados para el rol \"{nombre_rol}\""

    filas_html = "".join(
        f"<tr><td>{l.get('producto_excel', '')}</td><td>{l.get('company', '') or ''}</td>"
        f"<td>{l.get('cliente_excel', '') or ''}</td></tr>"
        for l in lineas
    )
    asunto = f"Presupuesto de ventas {anio} — solicitud de autorización (carga {id_carga})"
    cuerpo_html = f"""
    <div style="font-family: Arial, sans-serif; font-size: 14px; color: #111;">
        <p>Se solicitó autorización para {len(lineas)} línea(s) del presupuesto de ventas
        {anio} (carga {id_carga}).</p>
        <table border="1" cellpadding="6" cellspacing="0" style="border-collapse: collapse;">
            <tr style="background:#1F4E78; color:#fff;">
                <th>Producto</th><th>Company</th><th>Cliente</th>
            </tr>
            {filas_html}
        </table>
        <p>Entra a la app, módulo <b>Presupuesto de Ventas → ✅ autorizaciones</b>,
        para autorizar o rechazar.</p>
    </div>
    """
    return enviar_correo(
        destinatario=destinatarios, asunto=asunto, cuerpo_html=cuerpo_html,
        token=token, remitente=remitente,
    )


def _enviar_notificacion_vendedor_ventas(
    *, destinatario: str, aprobado: bool, id_carga: int, anio: int,
    producto_excel: str, motivo: Optional[str], token, remitente: str,
) -> tuple[bool, str]:
    if not destinatario:
        return False, "sin correo del vendedor"
    asunto = (
        f"Presupuesto de ventas {anio} — línea {'autorizada' if aprobado else 'rechazada'}"
    )
    motivo_html = f"<p><b>Motivo:</b> {motivo}</p>" if (not aprobado and motivo) else ""
    cuerpo_html = f"""
    <div style="font-family: Arial, sans-serif; font-size: 14px; color: #111;">
        <p>La línea <b>{producto_excel}</b> del presupuesto de ventas {anio}
        (carga {id_carga}) fue
        <b style="color:{'#16a34a' if aprobado else '#dc2626'}">
            {'AUTORIZADA' if aprobado else 'RECHAZADA'}
        </b>.</p>
        {motivo_html}
    </div>
    """
    return enviar_correo(
        destinatario=destinatario, asunto=asunto, cuerpo_html=cuerpo_html,
        token=token, remitente=remitente,
    )


def _obtener_hojas(archivo) -> list[str]:
    try:
        archivo.seek(0)
        return pd.ExcelFile(archivo).sheet_names or []
    except Exception:
        return []


def _catalogo_sae() -> tuple[set, dict, dict, dict, dict, dict, dict, dict, list]:
    """Returns (sae_set, code_to_label, label_to_code, code_to_desc, code_to_precio,
    code_to_linea, code_to_origen, code_to_unidad, options_list).

    code_to_precio (precio público en SAE), code_to_linea (cve_linea, "cve — desc"),
    code_to_origen (inve_clib01.camplib10) y code_to_unidad (inve01.uni_med) se
    usan para autocompletar esos campos al agregar un registro nuevo — el
    usuario no los captura a mano.
    """
    df = obtener_catalogo_productos_pv_ctrl()
    if df is None or df.empty:
        return set(), {}, {"": None}, {}, {}, {}, {}, {}, [""]

    records = df.to_dict("records")
    sae_set: set = set()
    code_to_label: dict = {}
    label_to_code: dict = {"": None}
    code_to_desc: dict = {}
    code_to_precio: dict = {}
    code_to_linea: dict = {}
    code_to_origen: dict = {}
    code_to_unidad: dict = {}
    items: list = []

    for r in records:
        code = str(r.get("cve_prod") or "").strip()
        desc = str(r.get("descr") or "").strip()
        if not code:
            continue
        label = f"{code}  {desc}" if desc else code
        sae_set.add(code)
        code_to_label[code] = label
        code_to_desc[code] = desc
        label_to_code[label] = code
        code_to_precio[code] = float(r.get("precio") or 0.0)
        cve_linea = str(r.get("cve_linea") or "").strip()
        desc_linea = str(r.get("linea") or "").strip()
        code_to_linea[code] = (cve_linea, f"{cve_linea} — {desc_linea}" if desc_linea else cve_linea)
        code_to_origen[code] = str(r.get("codigo_origen") or "").strip()
        code_to_unidad[code] = str(r.get("unidad") or "").strip().lower()
        items.append(((desc or code).lower(), label))

    # opciones ordenadas alfabéticamente por nombre de producto
    items.sort(key=lambda t: t[0])
    options = [""] + [lbl for _, lbl in items]
    return (
        sae_set, code_to_label, label_to_code, code_to_desc, code_to_precio,
        code_to_linea, code_to_origen, code_to_unidad, options,
    )


def _catalogo_clientes_sae() -> tuple[set, list]:
    """Catálogo de clientes SAE (mismo que consume Solicitudes vía
    buscar_clientes_sae_ctrl). Returns (clientes_set, opciones) donde cada
    opción es "clave - nombre - estado", igual que en Solicitudes — usado
    para exigir cliente de SAE cuando el estatus es "Budgeted"."""
    df = obtener_catalogo_clientes_pv_ctrl()
    if df is None or df.empty:
        return set(), [""]

    labels = []
    for r in df.to_dict("records"):
        clave = str(r.get("clave") or "").strip()
        nombre = str(r.get("nombre") or "").strip()
        estado = str(r.get("estado") or "").strip()
        partes = [p for p in (clave, nombre, estado) if p]
        label = " - ".join(partes)
        if label:
            labels.append(label)

    clientes_set = set(labels)
    opciones = [""] + sorted(labels, key=lambda s: s.lower())
    return clientes_set, opciones


def _construir_pivot(
    df: pd.DataFrame,
    sae_set: set,
    code_to_label: dict,
) -> tuple[pd.DataFrame, dict, dict]:
    cols_id = [c for c in _COLS_ID if c in df.columns]

    if df.empty:
        pivot_vacio = pd.DataFrame(
            columns=cols_id + ["_status", "_cve_prod_label", "estatus_excel", "precio", "precio_venta"]
        )
        return pivot_vacio, {}, {}

    # pivot_table descarta filas con NaN en el índice; rellenamos con ""
    df = df.copy()
    for c in cols_id:
        df[c] = df[c].fillna("")

    # mapping (row_key, mes) → id_presupuesto  (solo meses con registro real)
    mapping: dict = {}
    for _, row in df.iterrows():
        key = tuple(str(row.get(c) or "") for c in cols_id)
        mapping[(key, int(row["mes"]))] = int(row["id_presupuesto"])

    # row_meta: datos constantes por fila para insertar nuevos meses
    meta_cols = ["id_carga", "seccion", "region", "anio",
                 "cve_prod", "estatus_excel", "precio", "precio_venta"] + cols_id
    meta_cols = [c for c in meta_cols if c in df.columns]
    row_meta: dict = {}
    for _, row in df.iterrows():
        key = tuple(str(row.get(c) or "") for c in cols_id)
        if key not in row_meta:
            row_meta[key] = {c: row.get(c) for c in meta_cols}

    meta_map = df.groupby(cols_id, dropna=False)[
        [c for c in ["precio", "precio_venta", "cve_prod", "estatus_excel"] if c in df.columns]
    ].first().reset_index()

    pivot = df.pivot_table(
        index=cols_id,
        columns="mes",
        values="valor",
        aggfunc="sum",
        fill_value=0.0,
    ).reset_index()
    pivot = pivot.rename(columns=_MESES)
    pivot = pivot.merge(meta_map, on=cols_id, how="left")

    meses_presentes = [_MESES[m] for m in range(1, 13) if _MESES[m] in pivot.columns]

    # indicador SAE por fila
    def _status(cve):
        return "🟢" if str(cve or "").strip() in sae_set else "🟠"

    def _label(cve):
        code = str(cve or "").strip()
        return code_to_label.get(code, code)  # fallback: raw code

    pivot["_status"] = pivot["cve_prod"].apply(_status) if "cve_prod" in pivot.columns else "🟠"
    pivot["_cve_prod_label"] = pivot["cve_prod"].apply(_label) if "cve_prod" in pivot.columns else ""

    col_order = cols_id + ["_status", "_cve_prod_label", "estatus_excel", "precio", "precio_venta"] + meses_presentes
    pivot = pivot[[c for c in col_order if c in pivot.columns]]
    return pivot, mapping, row_meta


def _pivotear_meses(
    df: pd.DataFrame,
    cols_grupo: list[str],
    col_valor: str = "valor",
) -> pd.DataFrame:
    """Pivotea un dataframe 'largo' (una fila por mes) a 'ancho', con una
    columna ene..dic por grupo — mismo formato que la tabla de presupuesto.
    Se usa en las tablas de solo lectura ("ver todos", "autorizaciones")."""
    cols_grupo = [c for c in cols_grupo if c in df.columns] if df is not None else []

    if df is None or df.empty or not cols_grupo:
        return pd.DataFrame(columns=cols_grupo + list(_MESES.values()))

    df = df.copy()
    for c in cols_grupo:
        df[c] = df[c].fillna("")
    df["mes"] = pd.to_numeric(df["mes"], errors="coerce").fillna(0).astype(int)
    df[col_valor] = pd.to_numeric(df[col_valor], errors="coerce").fillna(0.0)

    pivot = df.pivot_table(
        index=cols_grupo,
        columns="mes",
        values=col_valor,
        aggfunc="sum",
        fill_value=0.0,
    ).reset_index()
    pivot = pivot.rename(columns=_MESES)

    for m in _MESES.values():
        if m not in pivot.columns:
            pivot[m] = 0.0

    return pivot[[c for c in (cols_grupo + list(_MESES.values())) if c in pivot.columns]]


def _agregar_totales_anio(pivote: pd.DataFrame) -> pd.DataFrame:
    """Calcula total_kg_anio/total_usd_anio a partir de los meses ya
    pivoteados (columna "valor") + "seccion" + "precio" (y "precio_venta"
    si viene en el pivote) — NO se usan las columnas cantidad_kg/importe
    guardadas en BD porque en registros antiguos pueden venir en NULL, lo
    que hace que la suma salga en 0 sin avisar. Mismo criterio que
    _guardar_pivot: en sección KG, "valor" son los kilos y el importe es
    valor×precio (precio venta si el usuario ya lo capturó, si no el de
    SAE); en cualquier otra sección "valor" ya es el monto en USD."""
    if pivote is None or pivote.empty:
        pivote = pivote.copy() if pivote is not None else pd.DataFrame()
        pivote["total_kg_anio"] = pd.Series(dtype=float)
        pivote["total_usd_anio"] = pd.Series(dtype=float)
        return pivote

    pivote = pivote.copy()
    meses_cols = [m for m in _MESES.values() if m in pivote.columns]
    suma_anio = pivote[meses_cols].sum(axis=1) if meses_cols else pd.Series(0.0, index=pivote.index)
    # comparación insensible a mayúsculas/espacios — hay datos (sobre todo
    # cargas antiguas por Excel) donde "seccion" no llega como "KG" exacto
    es_kg = (
        pivote["seccion"].astype(str).str.strip().str.upper() == "KG"
        if "seccion" in pivote.columns else pd.Series(False, index=pivote.index)
    )
    precio_col = pd.to_numeric(pivote["precio"], errors="coerce").fillna(0.0) if "precio" in pivote.columns else 0.0
    if "precio_venta" in pivote.columns:
        precio_venta_col = pd.to_numeric(pivote["precio_venta"], errors="coerce").fillna(0.0)
        precio_col = precio_venta_col.where(precio_venta_col > 0, precio_col)

    pivote["total_kg_anio"] = suma_anio.where(es_kg, 0.0)
    pivote["total_usd_anio"] = (suma_anio * precio_col).where(es_kg, suma_anio)
    return pivote


_ENCABEZADOS_EXPORT = {
    "company": "Company",
    "cliente_excel": "Cliente",
    "codigo_origen": "Código origen",
    "producto_excel": "Producto",
    "_cve_prod_label": "Cve prod / SAE",
    "_status": "En catálogo SAE",
    "estatus_excel": "Estatus",
    "precio": "Precio USD/unidad",
    "precio_venta": "Precio Venta",
}


def _pivot_a_excel_bytes(hojas: list[tuple[str, str, pd.DataFrame]]) -> bytes:
    """hojas: lista de (nombre_hoja, seccion, dataframe ya armado por _construir_pivot)."""
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for nombre_hoja, seccion, df in hojas:
        ws = wb.create_sheet(title=nombre_hoja[:31])

        if df is None or df.empty:
            ws.append(["sin datos para esta sección"])
            continue

        df_export = df.copy()
        df_export["_status"] = df_export["_status"].map({"🟢": "Sí", "🟠": "No"}).fillna("No")

        cols_id = [c for c in _COLS_ID if c in df_export.columns]
        meses = [m for m in _MESES.values() if m in df_export.columns]
        col_order = cols_id + ["_cve_prod_label", "_status", "estatus_excel", "precio", "precio_venta"] + meses
        df_export = df_export[[c for c in col_order if c in df_export.columns]]

        columnas = list(df_export.columns)
        ws.append([_ENCABEZADOS_EXPORT.get(c, c.upper()) for c in columnas])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        ws.freeze_panes = "A2"

        fmt_mes = "#,##0.000"
        meses_idx = {m: columnas.index(m) + 1 for m in meses}
        precio_idx = columnas.index("precio") + 1 if "precio" in columnas else None
        precio_venta_idx = columnas.index("precio_venta") + 1 if "precio_venta" in columnas else None

        for _, fila in df_export.iterrows():
            # openpyxl no acepta NaN/pd.NA como valor de celda ("Cannot
            # convert <NA> to Excel") — se limpian a None (celda en blanco)
            ws.append([None if pd.isna(v) else v for v in fila])

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for idx in meses_idx.values():
                row[idx - 1].number_format = fmt_mes
            if precio_idx:
                row[precio_idx - 1].number_format = "#,##0.0000"
            if precio_venta_idx:
                row[precio_venta_idx - 1].number_format = "#,##0.000"

        for idx, col_name in enumerate(columnas, start=1):
            largo = len(str(_ENCABEZADOS_EXPORT.get(col_name, col_name)))
            valores = df_export[col_name].astype(str).tolist()[:200]
            if valores:
                largo = max(largo, max(len(v) for v in valores))
            ws.column_dimensions[get_column_letter(idx)].width = min(largo + 2, 40)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


# ── plantilla de carga ("layout") ───────────────────────────────────────────
#
# Formato "standard" del parser (utils/presupuesto_ventas_excel_parser.py):
#   - fila ancla de sección: debe contener TURNOVER+VOLUME+KG (→ KG) o
#     TURNOVER+USD (→ USD), en cualquier celda de la fila.
#   - fila ancla de región: debe contener MEXICO o CAM+CARIBE.
#   - fila de encabezado: columnas FIJAS por posición — col0 estatus,
#     col1 company, col2 cliente, col3 código origen, col4 cve_prod,
#     col5 producto, col6 precio, col7 precio venta, col8+ meses
#     (detectados por nombre, ES/EN).
#   - los datos terminan en la primera fila con "producto" vacío.
#   - estatus válidos: BUDGETED, BUDGETEED, NOT IN BGT (o vacío).
# Se evita a propósito cualquier texto "CLAVE SAE" / "CODIGO UNIVERSAL" para
# no disparar por error uno de los parsers de vendedor (brewing/baking/juice).

_LAYOUT_HEADERS = [
    "ESTATUS", "COMPANY", "CLIENTE", "CODIGO ORIGEN", "CVE PROD",
    "PRODUCTO", "PRECIO USD/KG", "PRECIO VENTA",
    "ENE", "FEB", "MAR", "ABR", "MAY", "JUN",
    "JUL", "AGO", "SEP", "OCT", "NOV", "DIC",
]

_LAYOUT_BLOQUES = [
    ("2026 BGT Calculated TURNOVER in VOLUME KG", "MEXICO", [
        ["Budgeted", "NZMX", "Cliente Ejemplo 1", "CON-0001", "B0001", "PRODUCTO EJEMPLO 1", 10.5, 12.0,
         100, 100, 100, 120, 120, 120, 110, 110, 110, 100, 100, 100],
        ["Budgeted", "NZMX", "Cliente Ejemplo 2", "CON-0002", "B0002", "PRODUCTO EJEMPLO 2", 8.25, 9.5,
         50, 50, 60, 60, 70, 70, 70, 60, 60, 50, 50, 50],
    ]),
]

_LAYOUT_INSTRUCCIONES = [
    "PLANTILLA DE CARGA — PRESUPUESTO DE VENTAS",
    "",
    "Reglas para que la carga se procese correctamente:",
    "1. No borres ni renombres las filas que dicen \"TURNOVER in VOLUME KG\" ni la que dice "
    "\"MEXICO\" — son las que el sistema usa para saber en qué sección/región va el bloque de datos "
    "(por ahora la app solo captura KG México).",
    "2. No borres la fila de encabezados (ESTATUS, COMPANY, CLIENTE, ...) del bloque.",
    "3. Las columnas están en un orden fijo: ESTATUS, COMPANY, CLIENTE, CODIGO ORIGEN, CVE PROD, "
    "PRODUCTO, PRECIO USD/KG, PRECIO VENTA y luego los 12 meses — no insertes ni borres columnas en medio.",
    "4. PRECIO VENTA es opcional (déjalo en 0 o vacío si no aplica) — es el precio de venta que "
    "captura el usuario, distinto de PRECIO USD/KG (que es el precio de referencia de SAE).",
    "5. ESTATUS solo acepta: Budgeted, Budgeteed o Not in BGT (o dejarlo vacío).",
    "6. CVE PROD es la clave del producto en SAE (código B-xxxx). Si coincide con el catálogo, "
    "en la app aparecerá marcado en verde 🟢; si no, en naranja 🟠 (puedes corregirlo después "
    "desde la app con el buscador de producto SAE).",
    "7. Las cifras de los meses son KILOGRAMOS.",
    "8. Una fila en blanco marca el final del bloque de datos — no dejes filas en blanco "
    "en medio del bloque.",
    "9. Borra las filas de ejemplo (PRODUCTO EJEMPLO 1/2) antes de cargar tu información real; "
    "se incluyen solo para mostrar el formato esperado.",
]


def _generar_layout_presupuesto_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "layout_carga"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    seccion_font = Font(bold=True)

    for anchor_seccion, anchor_region, filas_ejemplo in _LAYOUT_BLOQUES:
        ws.append([anchor_seccion])
        for cell in ws[ws.max_row]:
            cell.font = seccion_font
        ws.append([anchor_region])
        for cell in ws[ws.max_row]:
            cell.font = seccion_font

        ws.append(_LAYOUT_HEADERS)
        for cell in ws[ws.max_row]:
            cell.fill = header_fill
            cell.font = header_font

        for fila in filas_ejemplo:
            ws.append(fila)

        ws.append([])  # fila en blanco: cierra el bloque

    for idx, header in enumerate(_LAYOUT_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(len(header) + 2, 12)
    ws.column_dimensions["F"].width = 24
    ws.freeze_panes = "A1"

    ws_instr = wb.create_sheet("instrucciones")
    for linea in _LAYOUT_INSTRUCCIONES:
        ws_instr.append([linea])
    ws_instr["A1"].font = Font(bold=True, size=13)
    ws_instr.column_dimensions["A"].width = 110
    for row in ws_instr.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _guardar_pivot(
    orig: pd.DataFrame,
    edited: pd.DataFrame,
    mapping: dict,
    row_meta: dict,
    seccion: str,
    region: Optional[str],
    cols_id: list[str],
    usuario_id: int,
    label_to_code: dict,
    id_carga: int,
    anio: int,
    code_to_linea: dict,
) -> tuple[int, int]:
    """Calcula los cambios fila por fila y los aplica en un solo lote (una conexión)."""
    inserts: list[dict] = []
    updates: list[dict] = []
    cve_prod_updates: list[dict] = []
    identidad_updates: list[dict] = []
    errores = 0

    for i in range(len(orig)):
        es_nueva = bool(orig.iloc[i].get("_nueva"))

        if es_nueva:
            producto = str(edited.iloc[i].get("producto_excel") or "").strip()
            sin_llenar = not producto and all(
                not str(edited.iloc[i].get(c) or "").strip() for c in cols_id
            )
            if sin_llenar:
                # fila agregada pero nunca llenada: se ignora sin marcar error
                continue
            if not producto:
                errores += 1
                continue

            row_key = tuple(str(edited.iloc[i].get(c) or "").strip() for c in cols_id)
            meta = {
                "id_carga": id_carga,
                "seccion": seccion,
                "region": region,
                "anio": anio,
                **{c: edited.iloc[i].get(c) for c in cols_id},
            }
        else:
            row_key = tuple(str(orig.iloc[i].get(c) or "") for c in cols_id)
            meta = dict(row_meta.get(row_key, {}))
            meta_orig = meta  # identidad tal como está hoy en BD, antes de cualquier rename

            # cambios en la identidad de la fila (company/cliente/código/producto):
            # mientras la línea no esté congelada se puede editar cualquier campo,
            # incluida su identidad — se actualiza en BD por id_carga + identidad
            # anterior, igual que cve_prod_updates
            row_key_edit = tuple(str(edited.iloc[i].get(c) or "") for c in cols_id)
            producto_edit_id = str(edited.iloc[i].get("producto_excel") or "").strip()
            if row_key_edit != row_key and producto_edit_id:
                identidad_nueva = {
                    "company": str(edited.iloc[i].get("company") or "").strip() or None,
                    "cliente_excel": str(edited.iloc[i].get("cliente_excel") or "").strip() or None,
                    "codigo_origen": str(edited.iloc[i].get("codigo_origen") or "").strip() or None,
                    "producto_excel": producto_edit_id,
                }
                identidad_updates.append({
                    "id_carga": int(meta_orig.get("id_carga") or 0),
                    "producto_excel_orig": str(meta_orig.get("producto_excel") or ""),
                    "cliente_excel_orig": meta_orig.get("cliente_excel") or None,
                    "codigo_origen_orig": meta_orig.get("codigo_origen") or None,
                    "company_orig": meta_orig.get("company") or None,
                    **identidad_nueva,
                })
                # los inserts de meses nuevos para esta fila usan ya la
                # identidad renombrada, para no quedar inconsistentes con el
                # update anterior
                meta = {**meta, **identidad_nueva}

        # ── cve_prod / cve_linea (nuevo valor u origen — la línea siempre se
        # deriva del producto SAE elegido, nunca la captura el usuario) ───────
        cve_edit_lbl = str(edited.iloc[i].get("_cve_prod_label") or "")
        cve_edit_cod = label_to_code.get(cve_edit_lbl, cve_edit_lbl.strip() or None)
        cve_linea_edit = code_to_linea.get(cve_edit_cod, ("", ""))[0] if cve_edit_cod else None

        # ── estatus (editable) ────────────────────────────────────────────────
        estatus_orig = "" if es_nueva else str(orig.iloc[i].get("estatus_excel") or "")
        estatus_edit = str(edited.iloc[i].get("estatus_excel") or "").strip()
        estatus_cambio = es_nueva or estatus_orig != estatus_edit

        if es_nueva:
            meta["cve_prod"] = cve_edit_cod
            meta["cve_linea"] = cve_linea_edit
            meta["estatus_excel"] = estatus_edit or None
        else:
            cve_orig_lbl = str(orig.iloc[i].get("_cve_prod_label") or "")
            if cve_orig_lbl != cve_edit_lbl:
                cve_prod_updates.append({
                    "id_carga": int(meta_orig.get("id_carga") or 0),
                    "producto_excel": str(meta_orig.get("producto_excel") or ""),
                    "cliente_excel": meta_orig.get("cliente_excel") or None,
                    "codigo_origen": meta_orig.get("codigo_origen") or None,
                    "company": meta_orig.get("company") or None,
                    "cve_linea": cve_linea_edit,
                    "cve_prod": cve_edit_cod,
                })

        # ── cambios en precio / precio venta / valores mensuales ────────────
        precio_orig = 0.0 if es_nueva else float(orig.iloc[i].get("precio") or 0)
        precio_edit = float(edited.iloc[i].get("precio") or 0)
        precio_cambio = es_nueva or abs(precio_edit - precio_orig) > 1e-6

        precio_venta_orig = 0.0 if es_nueva else float(orig.iloc[i].get("precio_venta") or 0)
        precio_venta_edit = float(edited.iloc[i].get("precio_venta") or 0)
        precio_venta_cambio = es_nueva or abs(precio_venta_edit - precio_venta_orig) > 1e-6

        for mes_num in range(1, 13):
            mes_name = _MESES[mes_num]
            if mes_name not in orig.columns:
                continue

            val_orig = 0.0 if es_nueva else float(orig.iloc[i].get(mes_name) or 0)
            val_edit = float(edited.iloc[i].get(mes_name) or 0)
            val_cambio = abs(val_edit - val_orig) > 1e-4

            if es_nueva and abs(val_edit) < 1e-9:
                # fila nueva: no crea registros para meses que quedaron en cero
                continue

            id_pv = None if es_nueva else mapping.get((row_key, mes_num))

            # un cambio de precio_venta actualiza los meses que ya existen en BD,
            # pero no crea registros nuevos en cero para meses sin capturar aún
            if not es_nueva and not val_cambio and not precio_cambio and not estatus_cambio \
                    and not (precio_venta_cambio and id_pv):
                continue

            precio_final = precio_edit if precio_cambio else precio_orig
            precio_venta_final = precio_venta_edit if precio_venta_cambio else precio_venta_orig
            valor_final = val_edit if val_cambio else val_orig

            if seccion == "KG":
                cantidad_kg = valor_final
                importe = round(valor_final * precio_final, 2)
            else:
                cantidad_kg = 0.0
                importe = valor_final

            if id_pv:
                updates.append({
                    "id_presupuesto": id_pv,
                    "valor": valor_final,
                    "precio": precio_final,
                    "precio_venta": precio_venta_final,
                    "cantidad_kg": cantidad_kg,
                    "importe": importe,
                    "estatus_excel": estatus_edit if estatus_cambio else None,
                })
            else:
                inserts.append({
                    "id_carga": int(meta.get("id_carga") or 0),
                    "seccion": seccion,
                    "region": meta.get("region") or None,
                    "anio": int(meta.get("anio") or 0),
                    "mes": mes_num,
                    "company": meta.get("company") or None,
                    "cliente_excel": meta.get("cliente_excel") or None,
                    "codigo_origen": meta.get("codigo_origen") or None,
                    "producto_excel": str(meta.get("producto_excel") or ""),
                    "cve_prod": meta.get("cve_prod") or None,
                    "cve_linea": meta.get("cve_linea") or None,
                    "estatus_excel": meta.get("estatus_excel") or None,
                    "precio": precio_final,
                    "precio_venta": precio_venta_final,
                    "valor": valor_final,
                    "cantidad_kg": cantidad_kg,
                    "importe": importe,
                    "usuario_id": usuario_id,
                })

    if not inserts and not updates and not cve_prod_updates and not identidad_updates:
        return 0, errores

    try:
        resultado = guardar_presupuesto_ventas_batch_ctrl(
            inserts=inserts,
            updates=updates,
            cve_prod_updates=cve_prod_updates,
            identidad_updates=identidad_updates,
        )
        cambios = (
            resultado.get("insertados", 0)
            + resultado.get("actualizados", 0)
            + resultado.get("cve_prod_actualizados", 0)
            + resultado.get("identidad_actualizados", 0)
        )
    except Exception:
        cambios = 0
        errores += len(inserts) + len(updates) + len(cve_prod_updates) + len(identidad_updates)

    return cambios, errores


# ── panel: selector de carga ──────────────────────────────────────────────────

def _selector_carga() -> Optional[int]:
    df = obtener_cargas_presupuesto_ventas_ctrl(limit=50, usuario_id=_get_usuario_id())

    if df is None or df.empty:
        st.info("aún no hay presupuestos cargados")
        return None

    opciones = {
        f"{r['id_carga']} | {r['nombre_archivo']} | {r['anio']} | {r.get('version', '')} | {r.get('comentarios', '') or ''}": int(r["id_carga"])
        for r in df.to_dict(orient="records")
    }
    labels = list(opciones.keys())
    default = st.session_state.get("pv_id_carga")
    idx = next((i for i, l in enumerate(labels) if opciones[l] == default), 0)

    label = st.selectbox("presupuesto", options=labels, index=idx, key="pv_select_carga")
    id_carga = opciones[label]
    st.session_state["pv_id_carga"] = id_carga
    return id_carga


# ── panel: crear presupuesto manual (sin depender de un Excel) ────────────────

def _panel_crear_manual() -> None:
    with st.expander("➕ crear presupuesto manual (sin Excel)"):
        col1, col2, col3 = st.columns([1, 1, 2])
        with col1:
            anio = st.number_input(
                "año", min_value=2020, max_value=2100, value=2026, step=1, key="pv_manual_anio"
            )
        with col2:
            version = st.text_input("versión", value="manual", key="pv_manual_version")
        with col3:
            comentarios = st.text_input("comentarios", value="", key="pv_manual_comentarios")

        if st.button("crear presupuesto manual", key="pv_btn_manual"):
            usuario_id = _get_usuario_id()
            if usuario_id <= 0:
                st.error("no se encontró el usuario en sesión")
                return
            try:
                id_carga = registrar_carga_presupuesto_ventas_ctrl(
                    nombre_archivo="Presupuesto manual",
                    hoja_origen="manual",
                    anio=int(anio),
                    version=version or None,
                    comentarios=comentarios or None,
                    usuario_id=usuario_id,
                )
                st.session_state["pv_id_carga"] = int(id_carga)
                st.success(f"presupuesto manual creado — id={id_carga}")
                st.rerun()
            except Exception as e:
                st.error(f"error al crear el presupuesto manual: {e}")


# ── panel: carga de Excel ─────────────────────────────────────────────────────

def _generar_excel_productos_sae_bytes(df: pd.DataFrame) -> bytes:
    """Excel con clave, nombre y existencia al día de los productos activos en
    SAE — usado como referencia rápida al capturar el presupuesto de ventas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "productos_sae"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    encabezados = ["Clave", "Nombre", "Existencia"]
    ws.append(encabezados)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A2"

    if df is not None and not df.empty:
        for _, fila in df.iterrows():
            ws.append([
                fila.get("cve_art"),
                fila.get("descr"),
                float(fila.get("existencia") or 0),
            ])

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row[2].number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 14

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _panel_carga() -> None:
    col_layout, col_productos = st.columns(2)
    with col_layout:
        st.download_button(
            "⬇️ descargar layout de carga",
            data=_generar_layout_presupuesto_bytes(),
            file_name="layout_presupuesto_ventas.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="pv_btn_layout",
            use_container_width=True,
            help="plantilla en blanco con el formato que espera la carga (incluye hoja de instrucciones)",
        )
    with col_productos:
        df_productos_sae = obtener_existencias_productos_pv_ctrl()
        st.download_button(
            "⬇️ descargar productos SAE (clave, nombre, existencia)",
            data=_generar_excel_productos_sae_bytes(df_productos_sae),
            file_name="productos_sae_existencia.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="pv_btn_productos_sae",
            use_container_width=True,
            help="catálogo de productos activos en SAE con su existencia al día (clave, nombre, existencia)",
        )

    archivo = st.file_uploader(
        "sube el archivo de presupuesto ventas",
        type=["xlsx", "xls"],
        key="pv_archivo",
    )
    if archivo is None:
        return

    hojas = _obtener_hojas(archivo)
    if not hojas:
        st.error("no fue posible leer las hojas del archivo")
        return

    col1, col2, col3 = st.columns([2, 1, 2])
    with col1:
        hoja = st.selectbox("hoja", options=hojas, key="pv_hoja")
    with col2:
        anio = st.number_input("año", min_value=2020, max_value=2100, value=2026, step=1, key="pv_anio")
    with col3:
        version = st.text_input("versión", value="v1", key="pv_version")

    comentarios = st.text_input("comentarios", value="", key="pv_comentarios")

    if st.button("cargar presupuesto", type="primary", use_container_width=True, key="pv_btn_cargar"):
        usuario_id = _get_usuario_id()
        if usuario_id <= 0:
            st.error("no se encontró el usuario en sesión")
            return
        try:
            archivo.seek(0)
            res = cargar_excel_directo_presupuesto_ventas_ctrl(
                archivo=archivo,
                nombre_archivo=archivo.name,
                hoja=hoja,
                anio=int(anio),
                usuario_id=usuario_id,
                version=version or None,
                comentarios=comentarios or None,
                reemplazar=True,
            )
            st.session_state["pv_id_carga"] = int(res["id_carga"])
            st.success(
                f"cargado — id={res['id_carga']} | "
                f"tablas={res['tablas_detectadas']} | "
                f"registros={res['total_registros']}"
            )
            st.rerun()
        except Exception as e:
            st.error(f"error al cargar: {e}")


# ── formulario: agregar registro manual (en vez de fila vacía en la tabla) ────

def _form_agregar_registro(
    *,
    seccion: str,
    region: Optional[str],
    id_carga: int,
    work_key: str,
    cols_id: list[str],
    meses_presentes: list[str],
    sae_opciones: list[str],
    label_to_code: dict,
    code_to_label: dict,
    code_to_desc: dict,
    code_to_precio: dict,
    code_to_linea: dict,
    code_to_origen: dict,
    code_to_unidad: dict,
    sae_set: set,
    clientes_set: set,
    clientes_opciones: list[str],
) -> None:
    prefix = f"pv_man_{seccion}_{region}_{id_carga}"
    with st.expander("➕ agregar registro"):
        st.caption(
            "campos obligatorios: **nombre producto** y **al menos un mes con valor distinto de cero** "
            "— company, cliente y estatus son opcionales; precio, línea y código origen se "
            "obtienen de SAE al elegir el producto y no se pueden editar a mano"
        )
        col1, col2 = st.columns(2)
        with col1:
            company = st.text_input("company (opcional)", key=f"{prefix}_company")
        with col2:
            cli_sel_key = f"{prefix}_cliente_sel"
            cliente_key = f"{prefix}_cliente"

            def _autofill_cliente() -> None:
                label = st.session_state.get(cli_sel_key) or ""
                if label:
                    st.session_state[cliente_key] = label

            st.selectbox(
                "cliente SAE (obligatorio si estatus = Budgeted)",
                clientes_opciones, key=cli_sel_key, on_change=_autofill_cliente,
            )
            cliente = st.text_input(
                "cliente (libre si estatus no es Budgeted)", key=cliente_key,
            )

        sel_key = f"{prefix}_prod_sel"
        nom_key = f"{prefix}_prod_nom"

        def _autofill_producto() -> None:
            code = label_to_code.get(st.session_state.get(sel_key) or "", "")
            st.session_state[nom_key] = code_to_desc.get(code, "") if code else ""

        cve_label = st.selectbox(
            "producto SAE (opcional — autocompleta nombre, precio, línea y código origen)",
            sae_opciones, key=sel_key, on_change=_autofill_producto,
        )
        producto_excel = st.text_input("nombre producto *obligatorio*", key=nom_key)

        # precio, línea y código origen son solo lectura y dependen únicamente
        # del producto seleccionado: se recalculan en cada render (en vez de
        # guardarse en session_state vía on_change) para que nunca queden
        # desincronizados de lo que el selectbox muestra, sin importar qué
        # otro widget del formulario haya disparado el rerun (p. ej. el
        # cliente SAE)
        _code_sel = label_to_code.get(cve_label or "", "")
        precio = float(code_to_precio.get(_code_sel, 0.0)) if _code_sel else 0.0
        linea_desc = code_to_linea.get(_code_sel, ("", ""))[1] if _code_sel else ""
        codigo_origen = code_to_origen.get(_code_sel, "") if _code_sel else ""
        unidad_sel = (code_to_unidad.get(_code_sel, "") if _code_sel else "") or "unidad"

        # si el precio público (tabla de precios x art) es 0, se usa como
        # respaldo el último precio de venta en facturas — cruzado con el
        # cliente si ya se conoce, o el último precio a cualquier cliente si no
        _cve_clie_sel = (
            cliente.strip().split(" - ", 1)[0].strip()
            if cliente.strip() in clientes_set
            else None
        )
        if _code_sel and precio == 0:
            precio = obtener_ultimo_precio_venta_ctrl(cve_art=_code_sel, cve_clie=_cve_clie_sel)

        col4, col5, col6, col7, col8 = st.columns(5)
        with col4:
            estatus_excel = st.selectbox(
                "estatus (opcional)", ["", "Budgeted", "Not in BGT", "Prospecto"], key=f"{prefix}_estatus",
            )
        with col5:
            st.number_input(
                f"precio USD/{unidad_sel} (de SAE, no editable)",
                min_value=0.0, format="%.4f", value=precio, disabled=True,
                key=f"{prefix}_precio_{_code_sel or 'none'}_{_cve_clie_sel or 'none'}",
            )
        with col6:
            st.text_input(
                "línea (de SAE, no editable)", value=linea_desc, disabled=True,
                key=f"{prefix}_linea_{_code_sel or 'none'}",
            )
        with col7:
            st.text_input(
                "código origen (de SAE, no editable)", value=codigo_origen, disabled=True,
                key=f"{prefix}_origen_{_code_sel or 'none'}",
            )
        with col8:
            precio_venta = st.number_input(
                "precio venta (opcional, lo capturas tú)",
                min_value=0.0, format="%.3f", value=0.0,
                key=f"{prefix}_precio_venta",
            )

        # se arma en bloques de hasta 6 columnas por fila para mantener el
        # orden ene→dic; los 12 meses están disponibles sin restricción
        st.caption("*obligatorio* captura al menos un mes con valor distinto de cero")
        fmt = "%.3f"
        meses_form = meses_presentes
        valores: dict[str, float] = {}
        CHUNK = 6
        for inicio in range(0, len(meses_form), CHUNK):
            fila_meses = meses_form[inicio:inicio + CHUNK]
            cols = st.columns(len(fila_meses))
            for col, m in zip(cols, fila_meses):
                with col:
                    valores[m] = st.number_input(
                        m.upper(), value=0.0, format=fmt, key=f"{prefix}_val_{m}",
                    )

        if st.button("agregar", key=f"{prefix}_btn"):
            if not producto_excel.strip():
                st.error("ingresa el nombre del producto")
                return
            if not any(abs(v) > 1e-9 for v in valores.values()):
                st.error("captura al menos un mes con valor distinto de cero")
                return
            if estatus_excel == "Budgeted" and cliente.strip() not in clientes_set:
                st.error(
                    "con estatus \"Budgeted\" el cliente debe ser uno del catálogo SAE — "
                    "selecciónalo en \"cliente SAE\" (Not in BGT / Prospecto no lo requieren)"
                )
                return
            code = label_to_code.get(cve_label, cve_label.strip() or None)
            fila = {c: "" for c in cols_id}
            fila.update({
                "company": company.strip(),
                "cliente_excel": cliente.strip(),
                "codigo_origen": codigo_origen.strip(),
                "producto_excel": producto_excel.strip(),
                "_status": "🟢" if code in sae_set else "🟠",
                "_cve_prod_label": code_to_label.get(code, "") if code else "",
                "estatus_excel": estatus_excel,
                "precio": float(precio),
                "precio_venta": float(precio_venta),
                "_nueva": True,
                "_estatus_linea": "captura",
                "_estatus_linea_badge": _ESTATUS_LINEA_BADGE["captura"],
            })
            for m in meses_presentes:
                fila[m] = valores.get(m, 0.0)

            # se agrega directo al estado de trabajo, sin tocar ver_key: no
            # hay nada nuevo que recargar de BD, solo mostrar la fila
            st.session_state[work_key] = pd.concat(
                [st.session_state[work_key], pd.DataFrame([fila])], ignore_index=True
            )
            st.success("registro agregado — revísalo en la tabla y guarda los cambios")
            st.rerun()


# ── panel: stock y órdenes de compra por llegar (SAE, solo lectura) ───────────

def _panel_stock_ordenes_sae() -> None:
    with st.container(border=True):
        st.markdown("**📦 stock y órdenes de compra por llegar (SAE)**")

        st.caption("stock de productos")
        df_stock = obtener_existencias_productos_pv_ctrl()
        if df_stock is None or df_stock.empty:
            st.info("sin datos de existencias en SAE")
        else:
            filtro = st.text_input(
                "buscar producto (clave o nombre)", key="pv_stock_filtro",
            )
            df_mostrar = df_stock
            if filtro.strip():
                f = filtro.strip().lower()
                df_mostrar = df_stock[
                    df_stock["cve_art"].astype(str).str.lower().str.contains(f, na=False)
                    | df_stock["descr"].astype(str).str.lower().str.contains(f, na=False)
                ]
            st.caption(f"{len(df_mostrar):,} producto(s)")
            st.dataframe(
                df_mostrar.rename(columns={
                    "cve_art": "clave", "descr": "producto", "linea": "línea",
                    "uni_med": "unidad", "existencia": "existencia",
                    "costo_prom": "costo prom.", "ult_costo": "último costo",
                })[["clave", "producto", "línea", "unidad", "existencia", "costo prom.", "último costo"]],
                use_container_width=True, height=300, hide_index=True,
            )

        st.caption("órdenes de compra por llegar")
        df_oc = obtener_ordenes_compra_pendientes_pv_ctrl()
        if df_oc is None or df_oc.empty:
            st.info("sin órdenes de compra pendientes de recibir en SAE")
        else:
            st.caption(f"{len(df_oc):,} línea(s) de orden de compra")
            st.dataframe(
                df_oc.rename(columns={
                    "cve_doc": "orden", "fecha_doc": "fecha", "fecha_rec": "fecha requerida",
                    "proveedor": "proveedor", "cve_art": "clave", "producto": "producto",
                    "linea": "línea", "cantidad": "cantidad", "unidad": "unidad", "precio": "precio",
                })[[
                    "orden", "fecha", "fecha requerida", "proveedor",
                    "clave", "producto", "línea", "cantidad", "unidad", "precio",
                ]],
                use_container_width=True, height=300, hide_index=True,
            )


# ── panel: comparación vs ventas reales SAE (solo lectura) ────────────────────

def _panel_comparacion_sae_ventas(
    work_df: pd.DataFrame,
    seccion: str,
    anio_actual: int,
    productos_sel: list,
    clientes_sel: list,
    label_to_code: dict,
    clientes_set: set,
) -> None:
    """Compara el presupuesto capturado (respetando los mismos filtros de
    producto/cliente de arriba) contra las ventas reales de SAE del año de la
    carga — cruzando por cliente + producto SAE (mismo criterio de
    cve_art/cve_clie que usa Construcción de forecast). Panel de solo
    lectura, no toca la tabla editable."""
    with st.container(border=True):
        st.markdown("**📈 comparación vs ventas reales (SAE)**")
        st.caption(
            "cruza por cliente + producto SAE — solo compara líneas con cliente/producto "
            "reconocidos en el catálogo SAE; respeta los filtros de arriba"
        )

        df_pv = work_df.copy()
        if productos_sel:
            df_pv = df_pv[df_pv["producto_excel"].astype(str).str.strip().isin(productos_sel)]
        if clientes_sel:
            df_pv = df_pv[df_pv["cliente_excel"].astype(str).str.strip().isin(clientes_sel)]

        if df_pv.empty:
            st.info("sin registros de presupuesto para comparar con los filtros elegidos")
            return

        meses_cols = [c for c in _MESES.values() if c in df_pv.columns]
        suma_meses = (
            df_pv[meses_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0).sum(axis=1)
            if meses_cols else pd.Series(0.0, index=df_pv.index)
        )
        precio_num = pd.to_numeric(df_pv.get("precio"), errors="coerce").fillna(0.0)
        precio_venta_num = pd.to_numeric(df_pv.get("precio_venta"), errors="coerce").fillna(0.0)
        precio_efectivo = precio_venta_num.where(precio_venta_num > 0, precio_num)
        if seccion == "KG":
            df_pv["_presupuesto_kg"] = suma_meses
            df_pv["_presupuesto_usd"] = suma_meses * precio_efectivo
        else:
            df_pv["_presupuesto_kg"] = 0.0
            df_pv["_presupuesto_usd"] = suma_meses

        df_pv["_cve_art"] = df_pv.get("_cve_prod_label", "").astype(str).map(
            lambda lbl: label_to_code.get(lbl, "") or ""
        )

        def _cve_clie_de(v) -> str:
            v = str(v or "").strip()
            return v.split(" - ", 1)[0].strip() if v in clientes_set and " - " in v else ""

        df_pv["_cve_clie"] = df_pv.get("cliente_excel", "").apply(_cve_clie_de)

        grp_pv = df_pv.groupby(["_cve_art", "_cve_clie"], as_index=False).agg(
            producto=("producto_excel", "first"),
            cliente=("cliente_excel", "first"),
            presupuesto_kg=("_presupuesto_kg", "sum"),
            presupuesto_usd=("_presupuesto_usd", "sum"),
        )

        # el real de SAE se acota siempre a los códigos (producto/cliente) que
        # sí aparecen en el presupuesto ya filtrado — evita traer el año
        # completo de ventas SAE (miles de líneas ajenas al presupuesto)
        codigos_art = set(grp_pv["_cve_art"]) - {""}
        codigos_clie = set(grp_pv["_cve_clie"]) - {""}

        df_real = obtener_ventas_reales_sae_pv_ctrl(anio_actual)
        if df_real is not None and not df_real.empty and (codigos_art or codigos_clie):
            df_real = df_real.copy()
            df_real["cve_art"] = df_real["cve_art"].astype(str).str.strip()
            df_real["cve_clie"] = df_real["cve_clie"].astype(str).str.strip()
            if codigos_art:
                df_real = df_real[df_real["cve_art"].isin(codigos_art)]
            if codigos_clie:
                df_real = df_real[df_real["cve_clie"].isin(codigos_clie)]
            grp_real = df_real.groupby(["cve_art", "cve_clie"], as_index=False).agg(
                real_kg=("cantidad", "sum"),
                real_usd=("importe", "sum"),
            )
        else:
            grp_real = pd.DataFrame(columns=["cve_art", "cve_clie", "real_kg", "real_usd"])

        comp = pd.merge(
            grp_pv, grp_real,
            left_on=["_cve_art", "_cve_clie"], right_on=["cve_art", "cve_clie"],
            how="outer",
        )
        comp["producto"] = comp["producto"].fillna(comp["_cve_art"])
        comp["cliente"] = comp["cliente"].fillna(comp["_cve_clie"])
        for c in ("presupuesto_kg", "presupuesto_usd", "real_kg", "real_usd"):
            comp[c] = pd.to_numeric(comp.get(c), errors="coerce").fillna(0.0)

        comp["diferencia_usd"] = comp["real_usd"] - comp["presupuesto_usd"]
        den_usd = comp["presupuesto_usd"].replace(0, float("nan"))
        comp["cumplimiento_%"] = (comp["real_usd"] / den_usd * 100).round(1)

        comp = comp[[
            "cliente", "producto", "presupuesto_kg", "presupuesto_usd",
            "real_kg", "real_usd", "diferencia_usd", "cumplimiento_%",
        ]].sort_values("presupuesto_usd", ascending=False)

        if comp.empty:
            st.info("sin ventas reales de SAE que crucen con este presupuesto")
            return

        st.caption(f"{len(comp):,} combinación(es) cliente/producto")
        st.dataframe(
            comp,
            use_container_width=True,
            hide_index=True,
            column_config={
                "cliente": st.column_config.TextColumn("cliente"),
                "producto": st.column_config.TextColumn("producto"),
                "presupuesto_kg": st.column_config.NumberColumn("presupuesto kg", format="%.0f"),
                "presupuesto_usd": st.column_config.NumberColumn("presupuesto USD", format="%.2f"),
                "real_kg": st.column_config.NumberColumn("real SAE kg", format="%.0f"),
                "real_usd": st.column_config.NumberColumn("real SAE USD", format="%.2f"),
                "diferencia_usd": st.column_config.NumberColumn("diferencia USD (real − presupuesto)", format="%.2f"),
                "cumplimiento_%": st.column_config.NumberColumn("cumplimiento %", format="%.1f"),
            },
            height=min(56 + len(comp) * 35, 500),
        )


_MES_NUM_DE = {v: k for k, v in _MESES.items()}


def _tabla_filtrada_presupuesto(
    work_df: pd.DataFrame,
    seccion: str,
    meses_presentes: list,
    productos_sel: list,
    clientes_sel: list,
    anio_actual: int,
    label_to_code: dict,
    clientes_set: set,
) -> None:
    """Tabla de solo lectura con las filas del presupuesto que coinciden con
    los filtros de producto/cliente elegidos arriba — la tabla editable de
    arriba no se toca, esto es nada más para revisar el resultado del
    filtro. Incluye, junto a cada mes capturado, la venta real de Aspel SAE
    de ese mismo mes ("Real <mes>"), cruzando por cliente + producto SAE —
    mismo criterio que usa Construcción de forecast."""
    if not productos_sel and not clientes_sel:
        st.caption("elige un producto o cliente arriba para ver los resultados aquí")
        return

    df = work_df.copy()
    if productos_sel:
        df = df[df["producto_excel"].astype(str).str.strip().isin(productos_sel)]
    if clientes_sel:
        df = df[df["cliente_excel"].astype(str).str.strip().isin(clientes_sel)]

    st.markdown("**resultado del filtro**")
    if df.empty:
        st.info("sin registros que coincidan con los filtros elegidos")
        return

    meses_cols = [c for c in meses_presentes if c in df.columns]
    suma_meses = (
        df[meses_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0).sum(axis=1)
        if meses_cols else pd.Series(0.0, index=df.index)
    )
    precio_num = pd.to_numeric(df.get("precio"), errors="coerce").fillna(0.0)
    precio_venta_num = pd.to_numeric(df.get("precio_venta"), errors="coerce").fillna(0.0)
    precio_efectivo = precio_venta_num.where(precio_venta_num > 0, precio_num)
    if seccion == "KG":
        df["total_kg_anio"] = suma_meses
        df["total_usd_anio"] = suma_meses * precio_efectivo
    else:
        df["total_kg_anio"] = 0.0
        df["total_usd_anio"] = suma_meses

    # venta real de Aspel SAE mes a mes — cruza por cliente + producto SAE
    df["_cve_art"] = df.get("_cve_prod_label", "").astype(str).map(
        lambda lbl: label_to_code.get(lbl, "") or ""
    )

    def _cve_clie_de(v) -> str:
        v = str(v or "").strip()
        return v.split(" - ", 1)[0].strip() if v in clientes_set and " - " in v else ""

    df["_cve_clie"] = df.get("cliente_excel", "").apply(_cve_clie_de)

    col_real = "cantidad" if seccion == "KG" else "importe"
    real_por_mes: dict[int, dict] = {}
    df_real = obtener_ventas_reales_sae_pv_ctrl(anio_actual)
    if df_real is not None and not df_real.empty:
        dfr = df_real.copy()
        dfr["cve_art"] = dfr["cve_art"].astype(str).str.strip()
        dfr["cve_clie"] = dfr["cve_clie"].astype(str).str.strip()
        dfr["mes"] = pd.to_numeric(dfr["mes"], errors="coerce").astype(int)
        for mes, sub in dfr.groupby("mes"):
            grp = sub.groupby(["cve_art", "cve_clie"], as_index=False)[col_real].sum()
            real_por_mes[int(mes)] = dict(zip(zip(grp["cve_art"], grp["cve_clie"]), grp[col_real]))

    meses_cols_out: list[str] = []
    encabezados = {
        "cliente_excel": "cliente", "producto_excel": "producto",
        "estatus_excel": "status", "precio_venta": "precio venta",
        "total_kg_anio": "Total Kg", "total_usd_anio": "Total USD",
    }
    for mn in meses_cols:
        mapa_mes = real_por_mes.get(_MES_NUM_DE.get(mn), {})
        real_col = f"_real_{mn}"
        df[real_col] = [
            mapa_mes.get((art, clie), 0.0)
            for art, clie in zip(df["_cve_art"], df["_cve_clie"])
        ]
        meses_cols_out.append(real_col)
        meses_cols_out.append(mn)
        encabezados[real_col] = f"Real {mn.capitalize()}"
        encabezados[mn] = mn.upper()

    cols_mostrar = [c for c in (
        "company", "cliente_excel", "producto_excel", "estatus_excel",
        "precio", "precio_venta", "total_kg_anio", "total_usd_anio",
    ) if c in df.columns] + meses_cols_out

    st.caption(
        f"{len(df):,} registro(s) coinciden con el filtro — "
        "\"Real <mes>\" es la venta real de Aspel SAE de ese mes (cruzada por cliente + producto)"
    )
    st.dataframe(
        df[cols_mostrar].rename(columns=encabezados),
        use_container_width=True, hide_index=True,
        height=min(56 + len(df) * 35, 500),
    )


# ── panel: tabla pivot ────────────────────────────────────────────────────────

def _panel_pivot(id_carga: int) -> None:
    df_all = obtener_presupuesto_ventas_ctrl(id_carga=id_carga)
    if df_all is None:
        df_all = pd.DataFrame()

    anio_default = None
    if df_all.empty:
        # presupuesto sin registros aún (p. ej. creado de forma manual, sin Excel):
        # se arma una estructura vacía con las columnas base para permitir captura manual
        df_all = pd.DataFrame(columns=_COLS_ID + [
            "mes", "anio", "seccion", "region", "valor", "importe",
            "cantidad_kg", "precio", "precio_venta", "cve_prod", "estatus_excel",
            "id_carga", "id_presupuesto",
        ])
        carga_meta = obtener_cargas_presupuesto_ventas_ctrl(id_carga=id_carga, limit=1)
        if carga_meta is not None and not carga_meta.empty and "anio" in carga_meta.columns:
            anio_default = int(carga_meta.iloc[0]["anio"])

    for col in ("valor", "importe", "cantidad_kg", "precio", "precio_venta"):
        if col in df_all.columns:
            df_all[col] = pd.to_numeric(df_all[col], errors="coerce").fillna(0.0)
        else:
            df_all[col] = 0.0

    if df_all["valor"].eq(0).all() and "importe" in df_all.columns:
        df_all["valor"] = df_all["importe"]

    df_all["mes"] = pd.to_numeric(df_all["mes"], errors="coerce").fillna(0).astype(int)

    # catálogo SAE (cacheado 1 hora)
    sae_set, code_to_label, label_to_code, code_to_desc, code_to_precio, code_to_linea, code_to_origen, code_to_unidad, sae_opciones = _catalogo_sae()
    clientes_set, clientes_opciones = _catalogo_clientes_sae()

    if "anio" in df_all.columns and not df_all.empty:
        anio_actual = int(df_all["anio"].iloc[0])
    else:
        anio_actual = int(anio_default or pd.Timestamp.today().year)

    # se muestran siempre los 12 meses (aunque no tengan datos aún) para poder
    # capturar cualquier mes al agregar o completar un registro
    meses_todos = list(_MESES.values())

    col_exp1, col_exp2 = st.columns([1, 3])
    with col_exp1:
        export_key = f"pv_export_bytes_{id_carga}"
        if st.button("📊 generar Excel", key=f"pv_export_gen_{id_carga}", use_container_width=True):
            hojas = []
            for label, seccion, region in _TABS_PIVOT:
                mask = pd.Series(True, index=df_all.index)
                if "seccion" in df_all.columns:
                    mask &= df_all["seccion"].astype(str) == seccion
                if region and "region" in df_all.columns:
                    mask &= df_all["region"].astype(str) == region
                df_sec_exp = df_all[mask].copy()
                pivot_exp, _, _ = _construir_pivot(df_sec_exp, sae_set, code_to_label)
                hojas.append((label, seccion, pivot_exp))
            st.session_state[export_key] = _pivot_a_excel_bytes(hojas)

        if st.session_state.get(export_key):
            st.download_button(
                "⬇️ descargar Excel",
                data=st.session_state[export_key],
                file_name=f"presupuesto_ventas_{id_carga}_{anio_actual}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"pv_export_dl_{id_carga}",
            )

    sub_tabs = st.tabs([t[0] for t in _TABS_PIVOT])

    for tab_ui, (label, seccion, region) in zip(sub_tabs, _TABS_PIVOT):
        with tab_ui:
            # `orig_key` guarda la última copia confiable de BD (pivot + mapping +
            # row_meta), y solo se reconstruye cuando `ver_key` cambia (justo
            # después de agregar/guardar/eliminar). `work_key` es el estado que
            # se edita en vivo: se pasa como `data=` a AgGrid en cada render y se
            # vuelve a guardar con lo que el grid reporte, para que un rerun
            # disparado por la edición de OTRA celda nunca le reenvíe al grid un
            # estado viejo y revierta lo que el usuario acaba de escribir.
            ver_key = f"pv_ver_{seccion}_{region}_{id_carga}"
            orig_key = f"pv_orig_{seccion}_{region}_{id_carga}"
            work_key = f"pv_work_{seccion}_{region}_{id_carga}"
            st.session_state.setdefault(ver_key, 0)

            necesita_reload = (
                orig_key not in st.session_state
                or st.session_state[orig_key]["ver"] != st.session_state[ver_key]
            )

            if necesita_reload:
                mask = pd.Series(True, index=df_all.index)
                if "seccion" in df_all.columns:
                    mask &= df_all["seccion"].astype(str) == seccion
                if region and "region" in df_all.columns:
                    mask &= df_all["region"].astype(str) == region
                df_sec = df_all[mask].copy()

                cols_id_reload = [c for c in _COLS_ID if c in df_all.columns]
                pivot_db, mapping_db, row_meta_db = _construir_pivot(df_sec, sae_set, code_to_label)

                for m in meses_todos:
                    if m not in pivot_db.columns:
                        pivot_db[m] = 0.0

                col_order = (
                    cols_id_reload + ["_status", "_cve_prod_label", "estatus_excel", "precio", "precio_venta"] + meses_todos
                )
                pivot_db = pivot_db[[c for c in col_order if c in pivot_db.columns]]
                pivot_db["_nueva"] = False

                # estatus de autorización por línea (no por mes): si no hay
                # fila en presupuesto_ventas_lineas para esa combinación
                # company/cliente/código/producto, es "captura" implícito
                lineas_df = obtener_presupuesto_ventas_lineas_ctrl(id_carga)
                estatus_por_linea: dict[tuple, str] = {}
                if lineas_df is not None and not lineas_df.empty:
                    for r in lineas_df.to_dict("records"):
                        clave = tuple(
                            str(r.get(c) or "") for c in
                            ("company", "cliente_excel", "codigo_origen", "producto_excel")
                        )
                        estatus_por_linea[clave] = str(r.get("estatus") or "captura")

                def _estatus_linea_de(row):
                    clave = tuple(str(row.get(c) or "") for c in cols_id_reload)
                    return estatus_por_linea.get(clave, "captura")

                pivot_db["_estatus_linea"] = (
                    pivot_db.apply(_estatus_linea_de, axis=1) if not pivot_db.empty else "captura"
                )
                pivot_db["_estatus_linea_badge"] = pivot_db["_estatus_linea"].map(_ESTATUS_LINEA_BADGE).fillna("🔵 captura")

                # la columna de autorización va al inicio de la tabla
                pivot_db = pivot_db[
                    ["_estatus_linea_badge"] + [c for c in pivot_db.columns if c != "_estatus_linea_badge"]
                ]

                st.session_state[orig_key] = {
                    "ver": st.session_state[ver_key],
                    "pivot": pivot_db,
                    "mapping": mapping_db,
                    "row_meta": row_meta_db,
                    "cols_id": cols_id_reload,
                }
                st.session_state[work_key] = pivot_db.copy()

            orig_data = st.session_state[orig_key]
            pivot = orig_data["pivot"]
            mapping = orig_data["mapping"]
            row_meta = orig_data["row_meta"]
            cols_id = orig_data["cols_id"]
            meses_presentes = meses_todos

            decimales = 3

            _form_agregar_registro(
                seccion=seccion,
                region=region,
                id_carga=id_carga,
                work_key=work_key,
                cols_id=cols_id,
                meses_presentes=meses_presentes,
                sae_opciones=sae_opciones,
                label_to_code=label_to_code,
                code_to_label=code_to_label,
                code_to_desc=code_to_desc,
                code_to_precio=code_to_precio,
                code_to_linea=code_to_linea,
                code_to_origen=code_to_origen,
                code_to_unidad=code_to_unidad,
                sae_set=sae_set,
                clientes_set=clientes_set,
                clientes_opciones=clientes_opciones,
            )

            work_df = st.session_state[work_key]
            if work_df.empty:
                st.info(f"sin datos para {label} — usa \"➕ agregar registro\" para capturar uno")
                continue

            # la columna de autorización siempre va primero (estática/pinned
            # a la izquierda) — se fuerza aquí en cada render porque, tras un
            # ciclo de edición, `work_key` se reemplaza por lo que devuelve
            # el grid y el orden de columnas ya no está garantizado
            if "_estatus_linea_badge" in work_df.columns:
                work_df = work_df[
                    ["_estatus_linea_badge"] + [c for c in work_df.columns if c != "_estatus_linea_badge"]
                ]

            # filtros de producto/cliente (widgets abajo de los botones, ver
            # más adelante) — se leen aquí porque el grid se arma antes en el
            # layout; como el valor ya quedó en session_state de la corrida
            # anterior, no hace falta que el widget se dibuje primero
            filtro_prod_key = f"pv_filtro_prod_{seccion}_{region}_{id_carga}"
            filtro_cli_key = f"pv_filtro_cli_{seccion}_{region}_{id_carga}"
            productos_filtro_sel = st.session_state.get(filtro_prod_key, [])
            clientes_filtro_sel = st.session_state.get(filtro_cli_key, [])

            # única restricción de edición: una línea deja de ser editable en
            # cuanto se solicita autorización ("enviada") y mientras esté
            # "autorizada" — "captura" y "rechazada" permiten editar
            # cualquier campo sin restricción
            _NOT_FROZEN = (
                "!(params.data && (params.data._estatus_linea === 'enviada'"
                " || params.data._estatus_linea === 'autorizada'))"
            )
            editable_no_congelada = JsCode(f"function(params) {{ return {_NOT_FROZEN}; }}")

            gb = GridOptionsBuilder.from_dataframe(work_df)
            gb.configure_default_column(editable=False, resizable=True, width=100)
            gb.configure_selection("multiple", use_checkbox=True, header_checkbox=True)

            # columnas sin ningún dato capturado (todo vacío/NaN en la carga
            # actual) se ocultan — p.ej. "company" o "código origen" cuando
            # esa carga no los trae
            def _columna_vacia(nombre_col: str) -> bool:
                if nombre_col not in work_df.columns:
                    return True
                serie = work_df[nombre_col].astype(str).str.strip()
                return bool(serie.isin(["", "nan", "None"]).all())

            _COLS_OCULTAS_SI_VACIAS = {"company", "codigo_origen"}

            # cliente, producto y código origen quedan fijos una vez
            # capturado el registro (cliente y producto se definen al
            # agregarlo, código origen viene de SAE) — solo "company"
            # permanece editable en la tabla
            _CAMPOS_ID_BLOQUEADOS = {"producto_excel", "cliente_excel", "codigo_origen"}
            for c in cols_id:
                gb.configure_column(
                    c, headerName=c.replace("_excel", "").replace("_", " "),
                    editable=False if c in _CAMPOS_ID_BLOQUEADOS else editable_no_congelada,
                    width=130,
                    pinned="left",
                    hide=c in _COLS_OCULTAS_SI_VACIAS and _columna_vacia(c),
                )

            gb.configure_column("_nueva", hide=True)
            gb.configure_column("_estatus_linea", hide=True)
            gb.configure_column(
                "_estatus_linea_badge", headerName="autorización", editable=False, width=120,
                pinned="left",
                headerTooltip="🔵 captura  |  🟡 enviada (esperando autorización)  |  🟢 autorizada  |  🔴 rechazada",
            )
            gb.configure_column(
                "_status", headerName="SAE", editable=False, width=70,
                pinned="left",
                headerTooltip="🟢 producto en catálogo SAE  |  🟠 no encontrado en SAE",
            )
            gb.configure_column(
                "_cve_prod_label",
                headerName="cve prod",
                # producto SAE: fijo una vez capturado, igual que precio y línea —
                # se elige al agregar el registro, no se reasigna después
                editable=False,
                width=200,
                pinned="left",
            )
            gb.configure_column(
                "estatus_excel", headerName="status", editable=editable_no_congelada, width=110,
                pinned="left",
                cellEditor="agSelectCellEditor",
                cellEditorParams={"values": ["", "Budgeted", "Not in BGT", "Prospecto"]},
            )
            gb.configure_column(
                "precio",
                headerName="precio USD/unidad",
                # precio: viene de SAE al agregar el registro, no editable en la tabla
                editable=False,
                width=110,
                pinned="left",
                type=["numericColumn"],
                valueFormatter=_value_formatter_js(4),
            )
            gb.configure_column(
                "precio_venta",
                headerName="Precio Venta",
                # precio_venta: lo captura el usuario a mano en la tabla
                editable=editable_no_congelada,
                width=120,
                pinned="left",
                type=["numericColumn"],
                cellEditor="agNumberCellEditor",
                valueFormatter=_value_formatter_js(3),
            )
            for m in meses_presentes:
                gb.configure_column(
                    m,
                    headerName=m.upper(),
                    editable=editable_no_congelada,
                    width=90,
                    type=["numericColumn"],
                    cellEditor="agNumberCellEditor",
                    cellStyle=_CELL_STYLE_VALORES,
                    valueFormatter=_value_formatter_js(decimales),
                )

            # totales al final — solo lectura, calculados en vivo a partir de
            # los meses ya capturados en el estado de trabajo (no se guardan
            # como columnas propias del work_key, solo se muestran)
            work_df_mostrar = work_df.copy()
            suma_meses = (
                work_df_mostrar[meses_presentes].apply(pd.to_numeric, errors="coerce").fillna(0.0).sum(axis=1)
                if meses_presentes else pd.Series(0.0, index=work_df_mostrar.index)
            )
            precio_num = pd.to_numeric(work_df_mostrar.get("precio"), errors="coerce").fillna(0.0)
            precio_venta_num = pd.to_numeric(work_df_mostrar.get("precio_venta"), errors="coerce").fillna(0.0)
            # precio venta manda sobre el precio SAE para el total en USD
            # cuando el usuario ya lo capturó (>0); si no, se usa el de SAE
            precio_efectivo = precio_venta_num.where(precio_venta_num > 0, precio_num)
            if seccion == "KG":
                work_df_mostrar["total_kg_anio"] = suma_meses
                work_df_mostrar["total_usd_anio"] = suma_meses * precio_efectivo
            else:
                work_df_mostrar["total_kg_anio"] = 0.0
                work_df_mostrar["total_usd_anio"] = suma_meses

            gb.configure_column(
                "total_kg_anio", headerName="Total Kg", editable=False, width=120,
                type=["numericColumn"], valueFormatter=_value_formatter_js(4),
            )
            gb.configure_column(
                "total_usd_anio", headerName="Total USD", editable=False, width=120,
                type=["numericColumn"], valueFormatter=_value_formatter_js(2),
            )

            st.caption(
                "🟢 en SAE  |  🟠 no en SAE  |  🟩 valor positivo  |  🟥 valor negativo "
                " — cliente, producto, cve prod, código origen y precio quedan fijos desde que se agrega"
                " el registro  |  selecciona filas con el checkbox para eliminarlas o para solicitar su"
                " autorización  |  🔵 captura 🟡 enviada 🟢 autorizada 🔴 rechazada — mientras esté en"
                " captura o rechazada puedes editar company, estatus y los meses; una vez enviada o"
                " autorizada ya no se puede editar nada"
            )

            grid_response = AgGrid(
                work_df_mostrar,
                gridOptions=gb.build(),
                update_on=[("cellValueChanged", 600), "selectionChanged"],
                data_return_mode=DataReturnMode.AS_INPUT,
                fit_columns_on_grid_load=False,
                allow_unsafe_jscode=True,
                height=min(56 + len(work_df) * 35, 680),
                key=f"pv_pivot_{seccion}_{region}_{id_carga}_{st.session_state[ver_key]}",
            )
            edited = pd.DataFrame(grid_response.get("data", []))
            edited = edited.drop(columns=["total_kg_anio", "total_usd_anio"], errors="ignore")

            # se guarda de inmediato lo que el grid reportó como estado de
            # trabajo, para que el próximo rerun (disparado por la edición de
            # OTRA celda) siga desde aquí en vez de reconstruir desde BD y
            # revertir lo que el usuario acaba de escribir
            if not edited.empty:
                st.session_state[work_key] = edited

            # al elegir cve_prod en una fila nueva, se llena "producto" con el
            # nombre del producto del catálogo SAE
            if not edited.empty and "_cve_prod_label" in edited.columns and "producto_excel" in edited.columns:
                hubo_cambio = False
                work_actual = st.session_state[work_key].copy()
                for i in range(len(work_actual)):
                    if not bool(work_actual.iloc[i].get("_nueva")):
                        continue
                    lbl = str(work_actual.iloc[i].get("_cve_prod_label") or "").strip()
                    if not lbl:
                        continue
                    code = label_to_code.get(lbl, lbl)
                    nombre = code_to_desc.get(code, "")
                    if nombre and str(work_actual.iloc[i].get("producto_excel") or "") != nombre:
                        work_actual.at[i, "producto_excel"] = nombre
                        hubo_cambio = True
                    if "_status" in work_actual.columns and code in sae_set:
                        work_actual.at[i, "_status"] = "🟢"

                if hubo_cambio:
                    # se mutó el estado de trabajo directamente, sin tocar
                    # ver_key: no hace falta recargar de BD, solo redibujar
                    st.session_state[work_key] = work_actual
                    edited = work_actual
                    st.rerun()

            seleccionadas = grid_response.get("selected_rows")
            if seleccionadas is None:
                seleccionadas = []
            elif isinstance(seleccionadas, pd.DataFrame):
                seleccionadas = seleccionadas.to_dict("records")

            col_save, col_del, col_aut = st.columns(3)
            with col_save:
                if st.button(
                    "💾 guardar cambios",
                    type="primary",
                    use_container_width=True,
                    key=f"pv_save_{seccion}_{region}_{id_carga}",
                ):
                    # `_guardar_pivot` compara `orig`/`edited` fila por fila
                    # posicionalmente; `pivot` (línea base de BD) no incluye las
                    # filas nuevas que solo viven en `work_key`, así que se
                    # rellena con placeholders "_nueva" para igualar el largo
                    # antes de diffear.
                    work_actual_save = st.session_state[work_key]
                    n_extra = len(work_actual_save) - len(pivot)
                    if n_extra > 0:
                        relleno = pd.DataFrame(
                            [{**{c: "" for c in cols_id}, "_nueva": True}] * n_extra
                        )
                        orig_para_diff = pd.concat([pivot, relleno], ignore_index=True)
                    else:
                        orig_para_diff = pivot

                    cambios, errores = _guardar_pivot(
                        orig_para_diff, work_actual_save, mapping, row_meta,
                        seccion, region, cols_id, _get_usuario_id(), label_to_code,
                        id_carga, anio_actual, code_to_linea,
                    )
                    if cambios:
                        st.success(f"guardados {cambios} registros")
                        # se recarga desde BD: descarta filas nuevas vacías y
                        # refleja lo recién guardado como la nueva línea base
                        st.session_state[ver_key] += 1
                        st.rerun()
                    if errores:
                        st.error(f"{errores} filas con error")
                    if not cambios and not errores:
                        st.info("sin cambios detectados")
            with col_del:
                if st.button(
                    "🗑️ eliminar seleccionados",
                    use_container_width=True,
                    disabled=not seleccionadas,
                    key=f"pv_del_{seccion}_{region}_{id_carga}",
                ):
                    # una línea enviada o autorizada ya no se puede eliminar
                    congeladas = [
                        f for f in seleccionadas
                        if str(f.get("_estatus_linea") or "captura") in ("enviada", "autorizada")
                    ]
                    seleccionadas = [
                        f for f in seleccionadas
                        if str(f.get("_estatus_linea") or "captura") not in ("enviada", "autorizada")
                    ]
                    if congeladas:
                        st.warning(
                            f"{len(congeladas)} línea(s) enviada(s)/autorizada(s) no se eliminaron "
                            "(ya no se pueden editar ni eliminar)."
                        )

                    registros_borrados = 0
                    filas_bd = 0
                    claves_borradas = {
                        tuple(str(f.get(c) or "").strip() for c in cols_id)
                        for f in seleccionadas
                    }
                    nuevas_borradas = [
                        f for f in seleccionadas if f.get("_nueva")
                    ]
                    for fila in seleccionadas:
                        if fila.get("_nueva"):
                            continue
                        try:
                            n = eliminar_registro_presupuesto_ventas_ctrl(
                                id_carga=id_carga,
                                seccion=seccion,
                                region=region,
                                producto_excel=str(fila.get("producto_excel") or ""),
                                cliente_excel=fila.get("cliente_excel") or None,
                                codigo_origen=fila.get("codigo_origen") or None,
                                company=fila.get("company") or None,
                            )
                            filas_bd += n
                            if n:
                                registros_borrados += 1
                        except Exception:
                            pass

                    # se quitan de inmediato del estado de trabajo (filas nuevas
                    # y/o ya guardadas); si hubo borrado real en BD, además se
                    # fuerza una recarga desde BD para que quede consistente
                    work_actual = st.session_state[work_key]
                    mask_fuera = work_actual.apply(
                        lambda r: tuple(str(r.get(c) or "").strip() for c in cols_id) in claves_borradas,
                        axis=1,
                    )
                    st.session_state[work_key] = work_actual[~mask_fuera].reset_index(drop=True)

                    if registros_borrados:
                        st.session_state[ver_key] += 1

                    if registros_borrados or nuevas_borradas:
                        st.success(
                            f"{registros_borrados} registro(s) eliminados ({filas_bd} filas de detalle)"
                            + (f", {len(nuevas_borradas)} fila(s) nueva(s) descartadas" if nuevas_borradas else "")
                        )
                        st.rerun()
                    else:
                        st.info("no se eliminó nada")
            with col_aut:
                elegibles = [
                    f for f in seleccionadas
                    if str(f.get("_estatus_linea") or "captura") in ("captura", "rechazada")
                    and str(f.get("producto_excel") or "").strip()
                ]
                if st.button(
                    "📤 solicitar autorización",
                    use_container_width=True,
                    disabled=not elegibles,
                    key=f"pv_aut_{seccion}_{region}_{id_carga}",
                ):
                    usuario = st.session_state.get("usuario") or {}
                    usuario_id = int(usuario.get("id") or 0)
                    usuario_nombre = str(usuario.get("nombre") or usuario.get("username") or "").strip()
                    usuario_email = str(usuario.get("email") or "").strip()

                    tipo_aut = _tipo_autorizacion_linea()
                    estatus_destino = "autorizada" if tipo_aut == "sin_autorizacion" else "enviada"

                    for fila in elegibles:
                        _cambiar_estatus_linea_ventas(
                            id_carga=id_carga,
                            company=fila.get("company") or None,
                            cliente_excel=fila.get("cliente_excel") or None,
                            codigo_origen=fila.get("codigo_origen") or None,
                            producto_excel=str(fila.get("producto_excel") or "").strip(),
                            estatus_nuevo=estatus_destino,
                            usuario_id=usuario_id,
                            usuario_nombre=usuario_nombre,
                            usuario_email=usuario_email,
                        )

                    if estatus_destino == "enviada":
                        nombre_rol = _rol_autorizador_linea()
                        token = st.session_state.get("microsoft_token")
                        ok_mail, msg_mail = _enviar_notificacion_autorizador_ventas(
                            lineas=elegibles, id_carga=id_carga, anio=anio_actual,
                            nombre_rol=nombre_rol, token=token, remitente=usuario_email,
                        )
                        st.success(f"{len(elegibles)} línea(s) enviada(s) a autorización ({nombre_rol})")
                        if not ok_mail:
                            st.warning(f"no se pudo enviar el correo de notificación: {msg_mail}")
                    else:
                        st.success(f"{len(elegibles)} línea(s) autorizada(s) automáticamente (tu rol no requiere autorización)")

                    st.session_state[ver_key] += 1
                    st.rerun()

            st.divider()
            st.markdown("**🔍 filtros**")
            productos_opciones_tab = (
                sorted(work_df["producto_excel"].astype(str).str.strip().replace("", pd.NA).dropna().unique())
                if "producto_excel" in work_df.columns else []
            )
            clientes_opciones_tab = (
                sorted(work_df["cliente_excel"].astype(str).str.strip().replace("", pd.NA).dropna().unique())
                if "cliente_excel" in work_df.columns else []
            )
            col_f1, col_f2 = st.columns(2)
            with col_f1:
                st.multiselect(
                    "filtrar por producto", options=productos_opciones_tab, key=filtro_prod_key,
                )
            with col_f2:
                st.multiselect(
                    "filtrar por cliente", options=clientes_opciones_tab, key=filtro_cli_key,
                )

            _tabla_filtrada_presupuesto(
                work_df=work_df,
                seccion=seccion,
                meses_presentes=meses_presentes,
                productos_sel=productos_filtro_sel,
                clientes_sel=clientes_filtro_sel,
                anio_actual=anio_actual,
                label_to_code=label_to_code,
                clientes_set=clientes_set,
            )

            _panel_comparacion_sae_ventas(
                work_df=work_df,
                seccion=seccion,
                anio_actual=anio_actual,
                productos_sel=productos_filtro_sel,
                clientes_sel=clientes_filtro_sel,
                label_to_code=label_to_code,
                clientes_set=clientes_set,
            )

    _panel_stock_ordenes_sae()


# ── panel: gestión de cargas ──────────────────────────────────────────────────

def _panel_gestionar_cargas() -> None:
    df = obtener_cargas_presupuesto_ventas_ctrl(limit=200, usuario_id=_get_usuario_id())

    if df is None or df.empty:
        st.info("no hay cargas registradas")
        return

    cols_mostrar = [c for c in
                    ["id_carga", "nombre_archivo", "anio", "version", "estatus", "comentarios", "created_at"]
                    if c in df.columns]
    st.dataframe(df[cols_mostrar], use_container_width=True, hide_index=True)

    st.divider()
    st.markdown("#### eliminar carga")
    st.caption("esto borra la carga, sus registros de presupuesto y el staging asociado de forma permanente")

    opciones = {
        f"{r['id_carga']} | {r['nombre_archivo']} | {r['anio']} | {r.get('version', '')}": int(r["id_carga"])
        for r in df.to_dict(orient="records")
    }
    label = st.selectbox("selecciona la carga a eliminar", options=list(opciones.keys()), key="pv_gc_select")
    id_sel = opciones[label]

    confirmar = st.checkbox(f"confirmo que quiero eliminar la carga {id_sel}", key="pv_gc_confirmar")

    if st.button("🗑️ eliminar carga", type="primary", disabled=not confirmar, key="pv_gc_btn_eliminar"):
        try:
            eliminar_carga_completa_presupuesto_ventas_ctrl(id_sel)
            st.success(f"carga {id_sel} eliminada correctamente")
            if st.session_state.get("pv_id_carga") == id_sel:
                del st.session_state["pv_id_carga"]
            st.rerun()
        except Exception as e:
            st.error(f"error al eliminar: {e}")


# ── panel: ver todos (forecastAdmin / SuperAdmin) ──────────────────────────────

_ENCABEZADOS_VER_TODOS = {
    "estatus_autorizacion_badge": "Autorización",
    "tipo": "Tipo",
    "usuario_nombre": "Usuario",
    "id_carga": "Carga",
    "anio": "Año",
    "seccion": "Sección",
    "region": "Región",
    "company": "Company",
    "cliente_excel": "Cliente",
    "codigo_origen": "Código origen",
    "producto_excel": "Producto",
    "cve_prod": "Cve prod",
    "precio": "Precio",
    "total_kg_anio": "Total Kilos Año",
    "total_usd_anio": "Total USD Año",
    "nombre_archivo": "Archivo origen",
    "version": "Versión",
}


def _tabla_generica_a_excel_bytes(nombre_hoja: str, df: pd.DataFrame, encabezados: dict[str, str]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja[:31]

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    columnas = list(df.columns)
    ws.append([encabezados.get(c, c) for c in columnas])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws.freeze_panes = "A2"

    for _, fila in df.iterrows():
        ws.append([None if pd.isna(v) else v for v in fila])

    for idx, col_name in enumerate(columnas, start=1):
        largo = len(str(encabezados.get(col_name, col_name)))
        valores = df[col_name].astype(str).tolist()[:200]
        if valores:
            largo = max(largo, max(len(v) for v in valores))
        ws.column_dimensions[get_column_letter(idx)].width = min(largo + 2, 40)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _panel_ver_todos() -> None:
    st.caption(
        "Presupuestos de ventas y compras de todos los usuarios — solo lectura. "
        "Para editar, entra a la carga correspondiente en las pestañas de arriba."
    )

    df_cat = obtener_catalogo_productos_pv_ctrl()
    lineas_disponibles = (
        sorted(df_cat["linea"].dropna().astype(str).str.strip().unique())
        if df_cat is not None and not df_cat.empty and "linea" in df_cat.columns
        else []
    )

    df_usuarios = obtener_usuarios_presupuesto_ctrl()
    usuarios_opciones = {"(todos)": None}
    if df_usuarios is not None and not df_usuarios.empty:
        for r in df_usuarios.to_dict("records"):
            nombre = str(r.get("usuario_nombre") or f"usuario {r.get('usuario_id')}").strip()
            usuarios_opciones[nombre] = int(r["usuario_id"])

    col1, col2, col3, col4, col5, col6 = st.columns(6)
    with col1:
        anio_sel = st.number_input(
            "año", min_value=2020, max_value=2100,
            value=int(date.today().year), step=1, key="pvt_anio",
        )
        filtrar_anio = st.checkbox("filtrar por año", value=True, key="pvt_filtrar_anio")
    with col2:
        usuario_label = st.selectbox("usuario", list(usuarios_opciones.keys()), key="pvt_usuario")
    with col3:
        tipo_sel = st.selectbox("tipo", ["(todos)", "venta", "compra"], key="pvt_tipo")
    with col4:
        linea_sel = st.selectbox("línea de producto", ["(todas)"] + lineas_disponibles, key="pvt_linea")
    with col5:
        df_cat_prod = df_cat
        if df_cat_prod is not None and not df_cat_prod.empty and linea_sel != "(todas)" and "linea" in df_cat_prod.columns:
            df_cat_prod = df_cat_prod[df_cat_prod["linea"].astype(str).str.strip() == linea_sel]
        productos_opciones = {"(todos)": None}
        if df_cat_prod is not None and not df_cat_prod.empty:
            for r in df_cat_prod.to_dict("records"):
                cve = str(r.get("cve_prod") or "").strip()
                desc = str(r.get("descr") or "").strip()
                if cve:
                    productos_opciones[f"{cve} - {desc}" if desc else cve] = cve
        producto_label = st.selectbox("producto", list(productos_opciones.keys()), key="pvt_producto")
    with col6:
        estatus_aut_opciones = ["(todos)"] + list(_ESTATUS_LINEA_BADGE.keys())
        estatus_aut_sel = st.selectbox(
            "autorización", estatus_aut_opciones,
            format_func=lambda k: _ESTATUS_LINEA_BADGE.get(k, k),
            key="pvt_estatus_aut",
        )

    df = obtener_presupuesto_ventas_compras_ctrl(
        anio=int(anio_sel) if filtrar_anio else None,
        usuario_id=usuarios_opciones[usuario_label],
        cve_prod=productos_opciones[producto_label],
        tipo=None if tipo_sel == "(todos)" else tipo_sel,
        estatus_autorizacion=None if estatus_aut_sel == "(todos)" else estatus_aut_sel,
    )

    if df is None or df.empty:
        st.info("sin resultados para los filtros seleccionados")
        return

    if linea_sel != "(todas)" and df_cat is not None and not df_cat.empty and "cve_prod" in df.columns:
        cve_de_linea = set(
            df_cat.loc[df_cat["linea"].astype(str).str.strip() == linea_sel, "cve_prod"]
            .astype(str).str.strip()
        )
        df = df[df["cve_prod"].astype(str).str.strip().isin(cve_de_linea)]

    if df.empty:
        st.info("sin resultados para los filtros seleccionados")
        return

    df = df.copy()
    if "estatus_autorizacion" in df.columns:
        df["estatus_autorizacion_badge"] = (
            df["estatus_autorizacion"].map(_ESTATUS_LINEA_BADGE).fillna(_ESTATUS_LINEA_BADGE["captura"])
        )

    # se arma por año, con los meses como columnas — mismo formato que la
    # tabla de presupuesto; se agrupa también por id_carga para no mezclar
    # cifras de cargas/versiones distintas que compartan la misma identidad
    cols_grupo = [c for c in _ENCABEZADOS_VER_TODOS if c in df.columns and c != "precio"]
    for c in cols_grupo:
        df[c] = df[c].fillna("")

    if "precio" in df.columns:
        df["precio"] = df.groupby(cols_grupo, dropna=False)["precio"].transform("first")
        cols_extra = ["precio"]
    else:
        cols_extra = []

    if "precio_venta" in df.columns:
        # solo se usa para calcular total_usd_anio (precio venta > 0 manda
        # sobre el de SAE) — no se muestra como columna aparte en esta tabla
        df["precio_venta"] = df.groupby(cols_grupo, dropna=False)["precio_venta"].transform("first")
        cols_extra.append("precio_venta")

    pivote = _pivotear_meses(df, cols_grupo + cols_extra, col_valor="valor")
    pivote = _agregar_totales_anio(pivote)

    st.caption(f"{len(pivote):,} línea(s)  |  🔵 captura 🟡 enviada 🟢 autorizada 🔴 rechazada")

    encabezados_pivote = {**_ENCABEZADOS_VER_TODOS, **{m: m.upper() for m in _MESES.values()}}
    col_order = (
        cols_grupo
        + [c for c in ("precio", "total_kg_anio", "total_usd_anio") if c in pivote.columns]
        + list(_MESES.values())
    )
    df_show = pivote[[c for c in col_order if c in pivote.columns]]
    df_renombrado = df_show.rename(columns=encabezados_pivote)

    estilo = df_renombrado.style
    if "Autorización" in df_renombrado.columns:
        estilo = estilo.map(_color_fondo_autorizacion, subset=["Autorización"])
    meses_upper = [m.upper() for m in _MESES.values() if m.upper() in df_renombrado.columns]
    if meses_upper:
        estilo = estilo.map(_color_valor_mes, subset=meses_upper)
    estilo = _formatear_numeros(estilo, df_renombrado, meses_upper)

    st.dataframe(
        estilo,
        use_container_width=True,
        hide_index=True,
        height=min(56 + len(df_show) * 35, 680),
    )

    xlsx_bytes = _tabla_generica_a_excel_bytes("presupuesto ventas y compras", df_show, encabezados_pivote)
    st.download_button(
        "⬇️ descargar Excel",
        data=xlsx_bytes,
        file_name=f"presupuesto_ventas_compras_{anio_sel if filtrar_anio else 'todos'}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="pvt_dl_excel",
    )


# ── panel: autorizaciones (Jefe de Ventas / Gerente de Ventas / SuperAdmin) ────

def _puede_autorizar_lineas() -> bool:
    usuario = st.session_state.get("usuario") or {}
    roles = set(_norm_roles_list(usuario.get("roles")))
    return bool(roles & {
        "jefe de ventas", "supervisor de ventas", "gerente de ventas", "gerente ventas",
        "forecastadmin", "superadmin",
    })


def _resolver_linea_autorizacion(
    fila: dict, estatus_nuevo: str, usuario: dict, token, remitente: str,
    correo_vendedor: str, motivo: Optional[str],
) -> tuple[bool, str]:
    """Aplica el nuevo estatus a una línea y notifica al vendedor.

    Devuelve (ok_mail, msg_mail) — el llamador es responsable de mostrar el
    resumen (útil para autorizar/rechazar en lote sin un st.success por fila).
    """
    usuario_id = int(usuario.get("id") or 0)
    usuario_nombre = str(usuario.get("nombre") or usuario.get("username") or "").strip()
    usuario_email = str(usuario.get("email") or "").strip()
    id_carga = int(fila["id_carga"])
    anio = int(fila.get("anio") or 0)
    producto_excel = str(fila.get("producto_excel") or "")

    if fila["tipo"] == "venta":
        linea_id, estatus_anterior = upsert_presupuesto_ventas_linea_ctrl(
            id_carga=id_carga, company=fila.get("company") or None,
            cliente_excel=fila.get("cliente_excel") or None,
            codigo_origen=fila.get("codigo_origen") or None,
            producto_excel=producto_excel, estatus=estatus_nuevo, usuario_id=usuario_id,
        )
        insertar_presupuesto_ventas_linea_estatus_ctrl(
            linea_id=linea_id, estatus_anterior=estatus_anterior, estatus_nuevo=estatus_nuevo,
            usuario_id=usuario_id, usuario_nombre=usuario_nombre, usuario_email=usuario_email,
            comentario=motivo,
        )
        ok_mail, msg_mail = _enviar_notificacion_vendedor_ventas(
            destinatario=correo_vendedor, aprobado=(estatus_nuevo == "autorizada"),
            id_carga=id_carga, anio=anio, producto_excel=producto_excel,
            motivo=motivo, token=token, remitente=remitente,
        )
    else:
        linea_id, estatus_anterior = upsert_presupuesto_compras_linea_ctrl(
            id_carga=id_carga, company=fila.get("company") or None,
            cliente_excel=fila.get("cliente_excel") or None,
            codigo_origen=fila.get("codigo_origen") or None,
            producto_excel=producto_excel, estatus=estatus_nuevo, usuario_id=usuario_id,
        )
        insertar_presupuesto_compras_linea_estatus_ctrl(
            linea_id=linea_id, estatus_anterior=estatus_anterior, estatus_nuevo=estatus_nuevo,
            usuario_id=usuario_id, usuario_nombre=usuario_nombre, usuario_email=usuario_email,
            comentario=motivo,
        )
        ok_mail, msg_mail = _enviar_notificacion_vendedor_compras(
            destinatario=correo_vendedor, aprobado=(estatus_nuevo == "autorizada"),
            id_carga=id_carga, anio=anio, producto_excel=producto_excel,
            motivo=motivo, token=token, remitente=remitente,
        )

    return ok_mail, msg_mail


def _agregar_meses_pendientes(pendientes: pd.DataFrame) -> pd.DataFrame:
    """Expande cada línea pendiente con sus valores mensuales, precio y total
    de kilos del año — una línea puede generar más de una fila si tiene datos
    en más de un bloque sección/región (p. ej. USD y KG). Mismo formato que
    la tabla de presupuesto (meses como columnas)."""
    cols_id = ["company", "cliente_excel", "codigo_origen", "producto_excel"]
    claves = ["tipo", "id_carga"] + cols_id
    claves_extendidas = claves + ["seccion", "region"]

    pendientes = pendientes.copy()
    for col in cols_id:
        if col in pendientes.columns:
            pendientes[col] = pendientes[col].fillna("")

    detalles = []
    for c in pendientes[["tipo", "id_carga"]].drop_duplicates().to_dict("records"):
        id_carga = int(c["id_carga"])
        det = (
            obtener_presupuesto_ventas_ctrl(id_carga=id_carga) if c["tipo"] == "venta"
            else obtener_presupuesto_compras_ctrl(id_carga=id_carga)
        )
        if det is not None and not det.empty:
            det = det.copy()
            det["tipo"] = c["tipo"]
            detalles.append(det)

    if detalles:
        detalle = pd.concat(detalles, ignore_index=True)
        # las columnas de agrupación deben quedar limpias de NaN antes de
        # calcular nada — si no, "" (usado internamente por _pivotear_meses)
        # no calza con NaN al mezclar por llave y el resultado queda en 0
        for col in claves_extendidas:
            if col in detalle.columns:
                detalle[col] = detalle[col].fillna("")

        cols_extra: list[str] = []
        if "precio" in detalle.columns:
            detalle["precio"] = detalle.groupby(claves_extendidas, dropna=False)["precio"].transform("first")
            cols_extra.append("precio")

        pivote = _pivotear_meses(detalle, claves_extendidas + cols_extra, col_valor="valor")
        pivote = _agregar_totales_anio(pivote)
        resultado = pendientes.merge(pivote, on=claves, how="left")
    else:
        resultado = pendientes.copy()

    for m in _MESES.values():
        if m not in resultado.columns:
            resultado[m] = 0.0
        resultado[m] = resultado[m].fillna(0.0)
    for col in ("seccion", "region"):
        if col not in resultado.columns:
            resultado[col] = ""
        resultado[col] = resultado[col].fillna("")
    for col in ("precio", "total_kg_anio", "total_usd_anio"):
        if col not in resultado.columns:
            resultado[col] = 0.0
        resultado[col] = resultado[col].fillna(0.0)

    return resultado.reset_index(drop=True)


def _panel_autorizaciones() -> None:
    usuario = st.session_state.get("usuario") or {}
    roles_viewer = set(_norm_roles_list(usuario.get("roles")))
    es_superadmin = bool(roles_viewer & {"superadmin"})
    es_gerente = bool(roles_viewer & {"gerente de ventas", "gerente ventas"})
    es_jefe = bool(roles_viewer & {"jefe de ventas", "supervisor de ventas"})

    if not (es_superadmin or es_gerente or es_jefe):
        st.info(
            "esta pestaña es para los roles Jefe de Ventas, Supervisor de Ventas, "
            "Gerente de Ventas o SuperAdmin."
        )
        return

    pendientes_ventas = obtener_presupuesto_ventas_lineas_pendientes_ctrl()
    if pendientes_ventas is not None and not pendientes_ventas.empty:
        pendientes_ventas = pendientes_ventas.copy()
        pendientes_ventas["tipo"] = "venta"

    pendientes_compras = obtener_presupuesto_compras_lineas_pendientes_ctrl()
    if pendientes_compras is not None and not pendientes_compras.empty:
        pendientes_compras = pendientes_compras.copy()
        pendientes_compras["tipo"] = "compra"

    partes = [df for df in (pendientes_ventas, pendientes_compras) if df is not None and not df.empty]
    if not partes:
        st.info("no hay líneas pendientes de autorización.")
        return
    pendientes = pd.concat(partes, ignore_index=True)

    if not es_superadmin:
        # re-deriva qué rol debe autorizar cada línea, según los roles del
        # dueño de la carga (misma cascada que _tipo_autorizacion_linea, pero
        # aplicada a un tercero — igual que _roles_creador_solicitud en
        # solicitudes de gastos)
        roles_por_usuario: dict[int, set[str]] = {}
        for uid in pendientes["carga_usuario_id"].dropna().unique():
            roles_por_usuario[int(uid)] = set(_norm_roles_list(obtener_roles_usuario_id_ctrl(int(uid))))

        def _me_corresponde(row) -> bool:
            owner_roles = roles_por_usuario.get(int(row["carga_usuario_id"]), set())
            if owner_roles & {"jefe de ventas", "supervisor de ventas"}:
                return es_gerente
            return es_jefe or es_gerente

        pendientes = pendientes[pendientes.apply(_me_corresponde, axis=1)]

    if pendientes.empty:
        st.info("no hay líneas pendientes de autorización para tu rol.")
        return

    pendientes = pendientes.sort_values("fecha_actualizacion").reset_index(drop=True)
    pendientes = _agregar_meses_pendientes(pendientes)

    # "usuario" (dueño de la carga) y "presupuesto" (comentarios de la carga)
    # dan contexto a quien autoriza — van al inicio de la tabla
    nombre_por_usuario: dict[int, str] = {}
    for uid in pendientes["carga_usuario_id"].dropna().unique():
        u = obtener_usuario_por_id_ctrl(int(uid)) or {}
        nombre_por_usuario[int(uid)] = str(u.get("nombre") or u.get("email") or "").strip()
    pendientes["usuario"] = pendientes["carga_usuario_id"].apply(
        lambda uid: nombre_por_usuario.get(int(uid), "") if pd.notna(uid) else ""
    )
    pendientes["presupuesto"] = (
        pendientes["comentarios"].fillna("") if "comentarios" in pendientes.columns else ""
    )

    cols_mostrar = ["usuario", "presupuesto"] + [c for c in [
        "tipo", "id_carga", "region", "company", "cliente_excel", "codigo_origen",
        "producto_excel", "anio", "version", "fecha_actualizacion",
        "precio", "total_kg_anio", "total_usd_anio", "carga_usuario_id",
    ] if c in pendientes.columns] + list(_MESES.values())

    st.caption(
        "marca el checkbox de cada fila (o el de la cabecera para seleccionar todas) para elegir una o "
        "varias líneas y autorizarlas o rechazarlas en lote — una línea puede aparecer en más de una fila "
        "si tiene datos en más de un bloque sección/región"
    )

    df_pendientes_mostrar = pendientes[cols_mostrar].reset_index(drop=True)
    meses_presentes_aut = [m for m in _MESES.values() if m in df_pendientes_mostrar.columns]

    gb = GridOptionsBuilder.from_dataframe(df_pendientes_mostrar)
    gb.configure_default_column(editable=False, resizable=True, width=110)
    gb.configure_selection("multiple", use_checkbox=True, header_checkbox=True)
    gb.configure_column("usuario", headerName="Usuario", width=140, pinned="left")
    gb.configure_column("presupuesto", headerName="Presupuesto", width=220, pinned="left")
    gb.configure_column("carga_usuario_id", hide=True)
    for col, header, decimales in (
        ("precio", "Precio", 4),
        ("total_kg_anio", "Total Kilos Año", 4),
        ("total_usd_anio", "Total USD Año", 2),
    ):
        if col in df_pendientes_mostrar.columns:
            gb.configure_column(
                col, headerName=header, type=["numericColumn"], valueFormatter=_value_formatter_js(decimales),
            )
    for m in meses_presentes_aut:
        gb.configure_column(
            m, headerName=m.upper(), type=["numericColumn"],
            cellStyle=_CELL_STYLE_VALORES, valueFormatter=_value_formatter_js(3),
        )

    grid_response = AgGrid(
        df_pendientes_mostrar,
        gridOptions=gb.build(),
        update_on=["selectionChanged"],
        data_return_mode=DataReturnMode.AS_INPUT,
        fit_columns_on_grid_load=False,
        allow_unsafe_jscode=True,
        height=min(56 + len(df_pendientes_mostrar) * 35, 520),
        key="pv_aut_tabla",
    )

    filas_seleccionadas = grid_response.get("selected_rows")
    if filas_seleccionadas is None:
        filas_seleccionadas = []
    elif isinstance(filas_seleccionadas, pd.DataFrame):
        filas_seleccionadas = filas_seleccionadas.to_dict("records")

    if not filas_seleccionadas:
        st.info("selecciona al menos una línea de la tabla para autorizar o rechazar.")
        return

    # una misma línea puede aparecer en varias filas (una por sección/región);
    # se deduplica por identidad para no procesarla ni notificar por correo más de una vez
    claves_unicas = {
        (f.get("tipo"), f.get("id_carga"), f.get("company"), f.get("cliente_excel"),
         f.get("codigo_origen"), f.get("producto_excel"))
        for f in filas_seleccionadas
    }
    n_lineas = len(claves_unicas)

    st.caption(f"{n_lineas} línea(s) seleccionada(s)")

    token = st.session_state.get("microsoft_token")
    remitente = str(usuario.get("email") or "").strip()

    def _procesar_lote(estatus_nuevo: str, motivo: Optional[str]) -> None:
        vistas: set = set()
        unicas: list[dict] = []
        for fila in filas_seleccionadas:
            clave = (
                fila.get("tipo"), fila.get("id_carga"), fila.get("company"),
                fila.get("cliente_excel"), fila.get("codigo_origen"), fila.get("producto_excel"),
            )
            if clave in vistas:
                continue
            vistas.add(clave)
            unicas.append(fila)

        fallos_mail: list[str] = []
        for fila in unicas:
            dueño = obtener_usuario_por_id_ctrl(int(fila.get("carga_usuario_id") or 0)) or {}
            correo_vendedor = str(dueño.get("email") or "").strip()
            ok_mail, msg_mail = _resolver_linea_autorizacion(
                fila, estatus_nuevo, usuario, token, remitente, correo_vendedor, motivo,
            )
            if not ok_mail:
                fallos_mail.append(f"{fila.get('producto_excel', '')} ({msg_mail})")

        st.success(f"{len(unicas)} línea(s) — {estatus_nuevo}")
        if fallos_mail:
            st.warning("no se pudo notificar por correo a: " + "; ".join(fallos_mail))

    col1, col2 = st.columns(2)
    with col1:
        if st.button(
            f"✅ autorizar {n_lineas} línea(s)",
            type="primary", use_container_width=True, key="pv_aut_btn_ok",
        ):
            _procesar_lote("autorizada", None)
            st.session_state["pv_aut_rechazando"] = False
            st.rerun()
    with col2:
        st.session_state.setdefault("pv_aut_rechazando", False)
        if not st.session_state["pv_aut_rechazando"]:
            if st.button(
                f"❌ rechazar {n_lineas} línea(s)",
                use_container_width=True, key="pv_aut_btn_rechazar",
            ):
                st.session_state["pv_aut_rechazando"] = True
                st.rerun()

    if st.session_state.get("pv_aut_rechazando"):
        motivo = st.text_area("motivo del rechazo (se aplica a todas las líneas seleccionadas)", key="pv_aut_motivo")
        if st.button("confirmar rechazo", key="pv_aut_btn_confirmar_rechazo"):
            _procesar_lote("rechazada", motivo.strip() or None)
            st.session_state["pv_aut_rechazando"] = False
            st.rerun()


# ── entry point ───────────────────────────────────────────────────────────────

def mostrar_modulo_presupuesto_ventas() -> None:
    st.subheader("presupuesto de ventas y compras")

    labels = ["📂 cargar Excel", "📊 tabla presupuesto", "🗑️ gestionar cargas"]
    ver_todos = _puede_ver_todos_presupuesto()
    puede_autorizar = _puede_autorizar_lineas()
    if ver_todos:
        labels.append("👁️ ver todos")
    if puede_autorizar:
        labels.append("✅ autorizaciones")

    tabs = st.tabs(labels)
    tab_carga, tab_tabla, tab_cargas = tabs[:3]
    tabs_extra = tabs[3:]

    with tab_carga:
        sub_ventas, sub_compras = st.tabs(["Ventas", "Compras"])
        with sub_ventas:
            with st.container(border=True):
                _panel_carga()
        with sub_compras:
            with st.container(border=True):
                _panel_carga_compras()

    with tab_tabla:
        sub_ventas, sub_compras = st.tabs(["Ventas", "Compras"])
        with sub_ventas:
            _panel_crear_manual()
            id_carga = _selector_carga()
            if id_carga is not None:
                _panel_pivot(id_carga)
        with sub_compras:
            _panel_crear_manual_compras()
            id_carga_compras = _selector_carga_compras()
            if id_carga_compras is not None:
                _panel_pivot_compras(id_carga_compras)

    with tab_cargas:
        sub_ventas, sub_compras = st.tabs(["Ventas", "Compras"])
        with sub_ventas:
            _panel_gestionar_cargas()
        with sub_compras:
            _panel_gestionar_cargas_compras()

    if ver_todos:
        with tabs_extra[0]:
            _panel_ver_todos()
        tabs_extra = tabs_extra[1:]

    if puede_autorizar:
        with tabs_extra[0]:
            _panel_autorizaciones()
