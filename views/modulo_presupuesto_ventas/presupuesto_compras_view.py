from __future__ import annotations

from io import BytesIO
from typing import Optional

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from st_aggrid import AgGrid, DataReturnMode, GridOptionsBuilder, JsCode

from controllers.presupuesto_compras_controller import (
    cargar_excel_directo_presupuesto_compras_ctrl,
    eliminar_carga_completa_presupuesto_compras_ctrl,
    eliminar_registro_presupuesto_compras_ctrl,
    guardar_presupuesto_compras_batch_ctrl,
    insertar_presupuesto_compras_linea_estatus_ctrl,
    obtener_cargas_presupuesto_compras_ctrl,
    obtener_catalogo_clientes_pv_compras_ctrl,
    obtener_catalogo_productos_pv_compras_ctrl,
    obtener_existencias_productos_pv_compras_ctrl,
    obtener_ordenes_compra_pendientes_pv_compras_ctrl,
    obtener_presupuesto_compras_ctrl,
    obtener_presupuesto_compras_lineas_ctrl,
    obtener_presupuesto_compras_lineas_pendientes_ctrl,
    registrar_carga_presupuesto_compras_ctrl,
    upsert_presupuesto_compras_linea_ctrl,
)
from controllers.solicitudes_controller import get_correos_usuarios_por_rol_ctrl
from utils.banxico import obtener_tipo_cambio_fix_banxico
from utils.envio_correo import enviar_correo


# ── constantes ────────────────────────────────────────────────────────────────

_MESES = {
    1: "ene", 2: "feb", 3: "mar", 4: "abr",
    5: "may", 6: "jun", 7: "jul", 8: "ago",
    9: "sep", 10: "oct", 11: "nov", 12: "dic",
}

_TABS_PIVOT = [
    ("KG México",        "KG",  "MEXICO"),
]

_COLS_ID = ["company", "cliente_excel", "codigo_origen", "producto_excel"]

# resalta en verde los valores positivos y en rojo los negativos; 0 sin color
_CELL_STYLE_VALORES = JsCode("""
function(params) {
    if (params.value > 0) {
        return {backgroundColor: '#d4edda', color: '#155724'};
    }
    if (params.value < 0) {
        return {backgroundColor: '#f8d7da', color: '#721c24'};
    }
    return null;
}
""")


def _value_formatter_js(decimales: int) -> JsCode:
    return JsCode(f"""
    function(params) {{
        if (params.value === null || params.value === undefined || params.value === '') return '';
        return Number(params.value).toFixed({decimales});
    }}
    """)


# ── helpers ───────────────────────────────────────────────────────────────────

def _get_usuario_id() -> int:
    usuario = st.session_state.get("usuario") or {}
    return int(usuario.get("id") or usuario.get("id_usuario") or 0)


def _norm_roles_list(values) -> list[str]:
    if not values:
        return []
    if isinstance(values, str):
        values = [values]
    return [str(v).strip().lower() for v in values if str(v or "").strip()]


def _tiene_rol(roles: list[str], *objetivos: str) -> bool:
    roles_set = set(_norm_roles_list(roles))
    objetivos_set = set(_norm_roles_list(objetivos))
    return bool(roles_set.intersection(objetivos_set))


# ── autorización por línea ──────────────────────────────────────────────────
# misma cascada que presupuesto de ventas (no hay roles "Jefe/Gerente de
# Compras" definidos — el pedido describe una sola jerarquía de autorización,
# la de ventas, aplicada por igual a los presupuestos de venta y de compra
# que arma el mismo equipo).

_ESTATUS_LINEA_BADGE = {
    "captura": "🔵 captura",
    "enviada": "🟡 enviada",
    "autorizada": "🟢 autorizada",
    "rechazada": "🔴 rechazada",
}


def _tipo_autorizacion_linea() -> str:
    usuario = st.session_state.get("usuario") or {}
    roles = usuario.get("roles")
    if _tiene_rol(roles, "gerente de ventas", "gerente ventas"):
        return "sin_autorizacion"
    if _tiene_rol(roles, "jefe de ventas", "supervisor de ventas"):
        return "autoriza_gerente_ventas"
    return "autoriza_jefe_ventas"


def _rol_autorizador_linea() -> str:
    tipo = _tipo_autorizacion_linea()
    if tipo == "autoriza_gerente_ventas":
        return "Gerente de Ventas"
    if tipo == "autoriza_jefe_ventas":
        return "Jefe de Ventas"
    return ""


def _cambiar_estatus_linea_compras(
    *,
    id_carga: int,
    company: Optional[str],
    cliente_excel: Optional[str],
    codigo_origen: Optional[str],
    producto_excel: str,
    estatus_nuevo: str,
    usuario_id: int,
    usuario_nombre: Optional[str],
    usuario_email: Optional[str],
    comentario: Optional[str] = None,
) -> int:
    linea_id, estatus_anterior = upsert_presupuesto_compras_linea_ctrl(
        id_carga=id_carga,
        company=company,
        cliente_excel=cliente_excel,
        codigo_origen=codigo_origen,
        producto_excel=producto_excel,
        estatus=estatus_nuevo,
        usuario_id=usuario_id,
    )
    insertar_presupuesto_compras_linea_estatus_ctrl(
        linea_id=linea_id,
        estatus_anterior=estatus_anterior,
        estatus_nuevo=estatus_nuevo,
        usuario_id=usuario_id,
        usuario_nombre=usuario_nombre,
        usuario_email=usuario_email,
        comentario=comentario,
    )
    return linea_id


def _enviar_notificacion_autorizador_compras(
    *, lineas: list[dict], id_carga: int, anio: int, nombre_rol: str, token, remitente: str,
) -> tuple[bool, str]:
    destinatarios = sorted({
        str(e).strip() for e in (get_correos_usuarios_por_rol_ctrl(nombre_rol) or []) if str(e or "").strip()
    })
    if not destinatarios:
        return False, f"no hay correos configurados para el rol \"{nombre_rol}\""

    filas_html = "".join(
        f"<tr><td>{l.get('producto_excel', '')}</td><td>{l.get('company', '') or ''}</td>"
        f"<td>{l.get('cliente_excel', '') or ''}</td></tr>"
        for l in lineas
    )
    asunto = f"Presupuesto de compra {anio} — solicitud de autorización (carga {id_carga})"
    cuerpo_html = f"""
    <div style="font-family: Arial, sans-serif; font-size: 14px; color: #111;">
        <p>Se solicitó autorización para {len(lineas)} línea(s) del presupuesto de compra
        {anio} (carga {id_carga}).</p>
        <table border="1" cellpadding="6" cellspacing="0" style="border-collapse: collapse;">
            <tr style="background:#1F4E78; color:#fff;">
                <th>Producto</th><th>Company</th><th>Cliente</th>
            </tr>
            {filas_html}
        </table>
        <p>Entra a la app, módulo <b>Presupuesto de Ventas → Presupuesto Compra → ✅ autorizaciones</b>,
        para autorizar o rechazar.</p>
    </div>
    """
    return enviar_correo(
        destinatario=destinatarios, asunto=asunto, cuerpo_html=cuerpo_html,
        token=token, remitente=remitente,
    )


def _enviar_notificacion_vendedor_compras(
    *, destinatario: str, aprobado: bool, id_carga: int, anio: int,
    producto_excel: str, motivo: Optional[str], token, remitente: str,
) -> tuple[bool, str]:
    if not destinatario:
        return False, "sin correo del vendedor"
    asunto = (
        f"Presupuesto de compra {anio} — línea {'autorizada' if aprobado else 'rechazada'}"
    )
    motivo_html = f"<p><b>Motivo:</b> {motivo}</p>" if (not aprobado and motivo) else ""
    cuerpo_html = f"""
    <div style="font-family: Arial, sans-serif; font-size: 14px; color: #111;">
        <p>La línea <b>{producto_excel}</b> del presupuesto de compra {anio}
        (carga {id_carga}) fue
        <b style="color:{'#16a34a' if aprobado else '#dc2626'}">
            {'AUTORIZADA' if aprobado else 'RECHAZADA'}
        </b>.</p>
        {motivo_html}
    </div>
    """
    return enviar_correo(
        destinatario=destinatario, asunto=asunto, cuerpo_html=cuerpo_html,
        token=token, remitente=remitente,
    )


def _obtener_hojas(archivo) -> list[str]:
    try:
        archivo.seek(0)
        return pd.ExcelFile(archivo).sheet_names or []
    except Exception:
        return []


def _catalogo_sae() -> tuple[set, dict, dict, dict, dict, dict, dict, dict, list]:
    """Returns (sae_set, code_to_label, label_to_code, code_to_desc, code_to_precio,
    code_to_linea, code_to_origen, code_to_unidad, options_list).

    code_to_precio es en realidad el costo (ult_costo de SAE / tipo de cambio
    FIX oficial de Banxico del día), no el precio público — se usa ese nombre
    de variable por continuidad con el resto del módulo. code_to_linea
    (cve_linea, "cve — desc"), code_to_origen (inve_clib01.camplib10) y
    code_to_unidad (inve01.uni_med) se usan para autocompletar esos campos
    al agregar un registro nuevo — el usuario no los captura a mano.
    """
    df = obtener_catalogo_productos_pv_compras_ctrl()
    if df is None or df.empty:
        return set(), {}, {"": None}, {}, {}, {}, {}, {}, [""]

    df_existencias = obtener_existencias_productos_pv_compras_ctrl()
    code_to_ult_costo: dict = {}
    if df_existencias is not None and not df_existencias.empty:
        for r in df_existencias.to_dict("records"):
            code_art = str(r.get("cve_art") or "").strip()
            if code_art:
                code_to_ult_costo[code_art] = float(r.get("ult_costo") or 0.0)

    try:
        tipo_cambio = obtener_tipo_cambio_fix_banxico()
    except Exception:
        tipo_cambio = 0.0

    records = df.to_dict("records")
    sae_set: set = set()
    code_to_label: dict = {}
    label_to_code: dict = {"": None}
    code_to_desc: dict = {}
    code_to_precio: dict = {}
    code_to_linea: dict = {}
    code_to_origen: dict = {}
    code_to_unidad: dict = {}
    items: list = []

    for r in records:
        code = str(r.get("cve_prod") or "").strip()
        desc = str(r.get("descr") or "").strip()
        if not code:
            continue
        label = f"{code}  {desc}" if desc else code
        sae_set.add(code)
        code_to_label[code] = label
        code_to_desc[code] = desc
        label_to_code[label] = code
        ult_costo = code_to_ult_costo.get(code, 0.0)
        code_to_precio[code] = round(ult_costo / tipo_cambio, 4) if tipo_cambio else 0.0
        cve_linea = str(r.get("cve_linea") or "").strip()
        desc_linea = str(r.get("linea") or "").strip()
        code_to_linea[code] = (cve_linea, f"{cve_linea} — {desc_linea}" if desc_linea else cve_linea)
        code_to_origen[code] = str(r.get("codigo_origen") or "").strip()
        code_to_unidad[code] = str(r.get("unidad") or "").strip().lower()
        items.append(((desc or code).lower(), label))

    # opciones ordenadas alfabéticamente por nombre de producto
    items.sort(key=lambda t: t[0])
    options = [""] + [lbl for _, lbl in items]
    return (
        sae_set, code_to_label, label_to_code, code_to_desc, code_to_precio,
        code_to_linea, code_to_origen, code_to_unidad, options,
    )


def _catalogo_clientes_sae() -> tuple[set, list]:
    """Catálogo de clientes SAE (mismo que consume Solicitudes vía
    buscar_clientes_sae_ctrl). Returns (clientes_set, opciones) donde cada
    opción es "clave - nombre - estado", igual que en Solicitudes — usado
    para exigir cliente de SAE cuando el estatus es "Budgeted"."""
    df = obtener_catalogo_clientes_pv_compras_ctrl()
    if df is None or df.empty:
        return set(), [""]

    labels = []
    for r in df.to_dict("records"):
        clave = str(r.get("clave") or "").strip()
        nombre = str(r.get("nombre") or "").strip()
        estado = str(r.get("estado") or "").strip()
        partes = [p for p in (clave, nombre, estado) if p]
        label = " - ".join(partes)
        if label:
            labels.append(label)

    clientes_set = set(labels)
    opciones = [""] + sorted(labels, key=lambda s: s.lower())
    return clientes_set, opciones


def _construir_pivot(
    df: pd.DataFrame,
    sae_set: set,
    code_to_label: dict,
) -> tuple[pd.DataFrame, dict, dict]:
    cols_id = [c for c in _COLS_ID if c in df.columns]

    if df.empty:
        pivot_vacio = pd.DataFrame(
            columns=cols_id + ["_status", "_cve_prod_label", "estatus_excel", "precio"]
        )
        return pivot_vacio, {}, {}

    # pivot_table descarta filas con NaN en el índice; rellenamos con ""
    df = df.copy()
    for c in cols_id:
        df[c] = df[c].fillna("")

    # mapping (row_key, mes) → id_presupuesto  (solo meses con registro real)
    mapping: dict = {}
    for _, row in df.iterrows():
        key = tuple(str(row.get(c) or "") for c in cols_id)
        mapping[(key, int(row["mes"]))] = int(row["id_presupuesto"])

    # row_meta: datos constantes por fila para insertar nuevos meses
    meta_cols = ["id_carga", "seccion", "region", "anio",
                 "cve_prod", "estatus_excel", "precio"] + cols_id
    meta_cols = [c for c in meta_cols if c in df.columns]
    row_meta: dict = {}
    for _, row in df.iterrows():
        key = tuple(str(row.get(c) or "") for c in cols_id)
        if key not in row_meta:
            row_meta[key] = {c: row.get(c) for c in meta_cols}

    meta_map = df.groupby(cols_id, dropna=False)[
        [c for c in ["precio", "cve_prod", "estatus_excel"] if c in df.columns]
    ].first().reset_index()

    pivot = df.pivot_table(
        index=cols_id,
        columns="mes",
        values="valor",
        aggfunc="sum",
        fill_value=0.0,
    ).reset_index()
    pivot = pivot.rename(columns=_MESES)
    pivot = pivot.merge(meta_map, on=cols_id, how="left")

    meses_presentes = [_MESES[m] for m in range(1, 13) if _MESES[m] in pivot.columns]

    # indicador SAE por fila
    def _status(cve):
        return "🟢" if str(cve or "").strip() in sae_set else "🟠"

    def _label(cve):
        code = str(cve or "").strip()
        return code_to_label.get(code, code)  # fallback: raw code

    pivot["_status"] = pivot["cve_prod"].apply(_status) if "cve_prod" in pivot.columns else "🟠"
    pivot["_cve_prod_label"] = pivot["cve_prod"].apply(_label) if "cve_prod" in pivot.columns else ""

    col_order = cols_id + ["_status", "_cve_prod_label", "estatus_excel", "precio"] + meses_presentes
    pivot = pivot[[c for c in col_order if c in pivot.columns]]
    return pivot, mapping, row_meta


_ENCABEZADOS_EXPORT = {
    "company": "Company",
    "cliente_excel": "Cliente",
    "codigo_origen": "Código origen",
    "producto_excel": "Producto",
    "_cve_prod_label": "Cve prod / SAE",
    "_status": "En catálogo SAE",
    "estatus_excel": "Estatus",
    "precio": "Costo USD/unidad",
}


def _pivot_a_excel_bytes(hojas: list[tuple[str, str, pd.DataFrame]]) -> bytes:
    """hojas: lista de (nombre_hoja, seccion, dataframe ya armado por _construir_pivot)."""
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for nombre_hoja, seccion, df in hojas:
        ws = wb.create_sheet(title=nombre_hoja[:31])

        if df is None or df.empty:
            ws.append(["sin datos para esta sección"])
            continue

        df_export = df.copy()
        df_export["_status"] = df_export["_status"].map({"🟢": "Sí", "🟠": "No"}).fillna("No")

        cols_id = [c for c in _COLS_ID if c in df_export.columns]
        meses = [m for m in _MESES.values() if m in df_export.columns]
        col_order = cols_id + ["_cve_prod_label", "_status", "estatus_excel", "precio"] + meses
        df_export = df_export[[c for c in col_order if c in df_export.columns]]

        columnas = list(df_export.columns)
        ws.append([_ENCABEZADOS_EXPORT.get(c, c.upper()) for c in columnas])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        ws.freeze_panes = "A2"

        fmt_mes = "#,##0.0000" if seccion == "KG" else "#,##0.00"
        meses_idx = {m: columnas.index(m) + 1 for m in meses}
        precio_idx = columnas.index("precio") + 1 if "precio" in columnas else None

        for _, fila in df_export.iterrows():
            # openpyxl no acepta NaN/pd.NA como valor de celda ("Cannot
            # convert <NA> to Excel") — se limpian a None (celda en blanco)
            ws.append([None if pd.isna(v) else v for v in fila])

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for idx in meses_idx.values():
                row[idx - 1].number_format = fmt_mes
            if precio_idx:
                row[precio_idx - 1].number_format = "#,##0.0000"

        for idx, col_name in enumerate(columnas, start=1):
            largo = len(str(_ENCABEZADOS_EXPORT.get(col_name, col_name)))
            valores = df_export[col_name].astype(str).tolist()[:200]
            if valores:
                largo = max(largo, max(len(v) for v in valores))
            ws.column_dimensions[get_column_letter(idx)].width = min(largo + 2, 40)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


# ── plantilla de carga ("layout") ───────────────────────────────────────────
#
# Formato "standard" del parser (utils/presupuesto_compras_excel_parser.py):
#   - fila ancla de sección: debe contener TURNOVER+VOLUME+KG (→ KG) o
#     TURNOVER+USD (→ USD), en cualquier celda de la fila.
#   - fila ancla de región: debe contener MEXICO o CAM+CARIBE.
#   - fila de encabezado: columnas FIJAS por posición — col0 estatus,
#     col1 company, col2 cliente, col3 código origen, col4 cve_prod,
#     col5 producto, col6 precio, col7+ meses (detectados por nombre, ES/EN).
#   - los datos terminan en la primera fila con "producto" vacío.
#   - estatus válidos: BUDGETED, BUDGETEED, NOT IN BGT (o vacío).
# Se evita a propósito cualquier texto "CLAVE SAE" / "CODIGO UNIVERSAL" para
# no disparar por error uno de los parsers de vendedor (brewing/baking/juice).

_LAYOUT_HEADERS = [
    "ESTATUS", "COMPANY", "CLIENTE", "CODIGO ORIGEN", "CVE PROD",
    "PRODUCTO", "PRECIO USD/KG",
    "ENE", "FEB", "MAR", "ABR", "MAY", "JUN",
    "JUL", "AGO", "SEP", "OCT", "NOV", "DIC",
]

_LAYOUT_BLOQUES = [
    ("2026 BGT Calculated TURNOVER in VOLUME KG", "MEXICO", [
        ["Budgeted", "NZMX", "Cliente Ejemplo 1", "CON-0001", "B0001", "PRODUCTO EJEMPLO 1", 10.5,
         100, 100, 100, 120, 120, 120, 110, 110, 110, 100, 100, 100],
        ["Budgeted", "NZMX", "Cliente Ejemplo 2", "CON-0002", "B0002", "PRODUCTO EJEMPLO 2", 8.25,
         50, 50, 60, 60, 70, 70, 70, 60, 60, 50, 50, 50],
    ]),
    ("2026 BGT Calculated TURNOVER in USD", "MEXICO", [
        ["Budgeted", "NZMX", "Cliente Ejemplo 1", "CON-0001", "B0001", "PRODUCTO EJEMPLO 1", 10.5,
         1050, 1050, 1050, 1260, 1260, 1260, 1155, 1155, 1155, 1050, 1050, 1050],
    ]),
    ("2026 BGT Calculated TURNOVER in VOLUME KG", "CAM & CARIBE", [
        ["Not in BGT", "NZNA", "Cliente CAM Ejemplo", None, "B0003", "PRODUCTO EJEMPLO 3", 9.95,
         200, 200, 0, 200, 200, 0, 200, 200, 0, 200, 200, 0],
    ]),
    ("2026 BGT Calculated TURNOVER in USD", "CAM & CARIBE", [
        ["Not in BGT", "NZNA", "Cliente CAM Ejemplo", None, "B0003", "PRODUCTO EJEMPLO 3", 9.95,
         1990, 1990, 0, 1990, 1990, 0, 1990, 1990, 0, 1990, 1990, 0],
    ]),
]

_LAYOUT_INSTRUCCIONES = [
    "PLANTILLA DE CARGA — PRESUPUESTO DE COMPRAS",
    "",
    "Reglas para que la carga se procese correctamente:",
    "1. No borres ni renombres las filas que dicen \"TURNOVER in VOLUME KG\" / \"TURNOVER in USD\" "
    "ni las que dicen \"MEXICO\" / \"CAM & CARIBE\" — son las que el sistema usa para saber en qué "
    "sección/región va cada bloque de datos.",
    "2. No borres la fila de encabezados (ESTATUS, COMPANY, CLIENTE, ...) de cada bloque.",
    "3. Las columnas están en un orden fijo: ESTATUS, COMPANY, CLIENTE, CODIGO ORIGEN, CVE PROD, "
    "PRODUCTO, PRECIO USD/KG y luego los 12 meses — no insertes ni borres columnas en medio.",
    "4. ESTATUS solo acepta: Budgeted, Budgeteed o Not in BGT (o dejarlo vacío).",
    "5. CVE PROD es la clave del producto en SAE (código B-xxxx). Si coincide con el catálogo, "
    "en la app aparecerá marcado en verde 🟢; si no, en naranja 🟠 (puedes corregirlo después "
    "desde la app con el buscador de producto SAE).",
    "6. En los bloques \"TURNOVER in VOLUME KG\" las cifras de los meses son KILOGRAMOS.",
    "7. En los bloques \"TURNOVER in USD\" las cifras de los meses son DÓLARES.",
    "8. Una fila en blanco marca el final de cada bloque de datos — no dejes filas en blanco "
    "en medio de un bloque con datos.",
    "9. Borra las filas de ejemplo (PRODUCTO EJEMPLO 1/2/3) antes de cargar tu información real; "
    "se incluyen solo para mostrar el formato esperado.",
]


def _generar_layout_presupuesto_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "layout_carga"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    seccion_font = Font(bold=True)

    for anchor_seccion, anchor_region, filas_ejemplo in _LAYOUT_BLOQUES:
        ws.append([anchor_seccion])
        for cell in ws[ws.max_row]:
            cell.font = seccion_font
        ws.append([anchor_region])
        for cell in ws[ws.max_row]:
            cell.font = seccion_font

        ws.append(_LAYOUT_HEADERS)
        for cell in ws[ws.max_row]:
            cell.fill = header_fill
            cell.font = header_font

        for fila in filas_ejemplo:
            ws.append(fila)

        ws.append([])  # fila en blanco: cierra el bloque

    for idx, header in enumerate(_LAYOUT_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(len(header) + 2, 12)
    ws.column_dimensions["F"].width = 24
    ws.freeze_panes = "A1"

    ws_instr = wb.create_sheet("instrucciones")
    for linea in _LAYOUT_INSTRUCCIONES:
        ws_instr.append([linea])
    ws_instr["A1"].font = Font(bold=True, size=13)
    ws_instr.column_dimensions["A"].width = 110
    for row in ws_instr.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _guardar_pivot(
    orig: pd.DataFrame,
    edited: pd.DataFrame,
    mapping: dict,
    row_meta: dict,
    seccion: str,
    region: Optional[str],
    cols_id: list[str],
    usuario_id: int,
    label_to_code: dict,
    id_carga: int,
    anio: int,
    code_to_linea: dict,
) -> tuple[int, int]:
    """Calcula los cambios fila por fila y los aplica en un solo lote (una conexión)."""
    inserts: list[dict] = []
    updates: list[dict] = []
    cve_prod_updates: list[dict] = []
    identidad_updates: list[dict] = []
    errores = 0

    for i in range(len(orig)):
        es_nueva = bool(orig.iloc[i].get("_nueva"))

        if es_nueva:
            producto = str(edited.iloc[i].get("producto_excel") or "").strip()
            sin_llenar = not producto and all(
                not str(edited.iloc[i].get(c) or "").strip() for c in cols_id
            )
            if sin_llenar:
                # fila agregada pero nunca llenada: se ignora sin marcar error
                continue
            if not producto:
                errores += 1
                continue

            row_key = tuple(str(edited.iloc[i].get(c) or "").strip() for c in cols_id)
            meta = {
                "id_carga": id_carga,
                "seccion": seccion,
                "region": region,
                "anio": anio,
                **{c: edited.iloc[i].get(c) for c in cols_id},
            }
        else:
            row_key = tuple(str(orig.iloc[i].get(c) or "") for c in cols_id)
            meta = dict(row_meta.get(row_key, {}))
            meta_orig = meta  # identidad tal como está hoy en BD, antes de cualquier rename

            # cambios en la identidad de la fila (company/cliente/código/producto):
            # mientras la línea no esté congelada se puede editar cualquier campo,
            # incluida su identidad — se actualiza en BD por id_carga + identidad
            # anterior, igual que cve_prod_updates
            row_key_edit = tuple(str(edited.iloc[i].get(c) or "") for c in cols_id)
            producto_edit_id = str(edited.iloc[i].get("producto_excel") or "").strip()
            if row_key_edit != row_key and producto_edit_id:
                identidad_nueva = {
                    "company": str(edited.iloc[i].get("company") or "").strip() or None,
                    "cliente_excel": str(edited.iloc[i].get("cliente_excel") or "").strip() or None,
                    "codigo_origen": str(edited.iloc[i].get("codigo_origen") or "").strip() or None,
                    "producto_excel": producto_edit_id,
                }
                identidad_updates.append({
                    "id_carga": int(meta_orig.get("id_carga") or 0),
                    "producto_excel_orig": str(meta_orig.get("producto_excel") or ""),
                    "cliente_excel_orig": meta_orig.get("cliente_excel") or None,
                    "codigo_origen_orig": meta_orig.get("codigo_origen") or None,
                    "company_orig": meta_orig.get("company") or None,
                    **identidad_nueva,
                })
                # los inserts de meses nuevos para esta fila usan ya la
                # identidad renombrada, para no quedar inconsistentes con el
                # update anterior
                meta = {**meta, **identidad_nueva}

        # ── cve_prod / cve_linea (nuevo valor u origen — la línea siempre se
        # deriva del producto SAE elegido, nunca la captura el usuario) ───────
        cve_edit_lbl = str(edited.iloc[i].get("_cve_prod_label") or "")
        cve_edit_cod = label_to_code.get(cve_edit_lbl, cve_edit_lbl.strip() or None)
        cve_linea_edit = code_to_linea.get(cve_edit_cod, ("", ""))[0] if cve_edit_cod else None

        # ── estatus (editable) ────────────────────────────────────────────────
        estatus_orig = "" if es_nueva else str(orig.iloc[i].get("estatus_excel") or "")
        estatus_edit = str(edited.iloc[i].get("estatus_excel") or "").strip()
        estatus_cambio = es_nueva or estatus_orig != estatus_edit

        if es_nueva:
            meta["cve_prod"] = cve_edit_cod
            meta["cve_linea"] = cve_linea_edit
            meta["estatus_excel"] = estatus_edit or None
        else:
            cve_orig_lbl = str(orig.iloc[i].get("_cve_prod_label") or "")
            if cve_orig_lbl != cve_edit_lbl:
                cve_prod_updates.append({
                    "id_carga": int(meta_orig.get("id_carga") or 0),
                    "producto_excel": str(meta_orig.get("producto_excel") or ""),
                    "cliente_excel": meta_orig.get("cliente_excel") or None,
                    "codigo_origen": meta_orig.get("codigo_origen") or None,
                    "company": meta_orig.get("company") or None,
                    "cve_linea": cve_linea_edit,
                    "cve_prod": cve_edit_cod,
                })

        # ── cambios en precio / valores mensuales ───────────────────────────
        precio_orig = 0.0 if es_nueva else float(orig.iloc[i].get("precio") or 0)
        precio_edit = float(edited.iloc[i].get("precio") or 0)
        precio_cambio = es_nueva or abs(precio_edit - precio_orig) > 1e-6

        for mes_num in range(1, 13):
            mes_name = _MESES[mes_num]
            if mes_name not in orig.columns:
                continue

            val_orig = 0.0 if es_nueva else float(orig.iloc[i].get(mes_name) or 0)
            val_edit = float(edited.iloc[i].get(mes_name) or 0)
            val_cambio = abs(val_edit - val_orig) > 1e-4

            if es_nueva and abs(val_edit) < 1e-9:
                # fila nueva: no crea registros para meses que quedaron en cero
                continue

            if not es_nueva and not val_cambio and not precio_cambio and not estatus_cambio:
                continue

            precio_final = precio_edit if precio_cambio else precio_orig
            valor_final = val_edit if val_cambio else val_orig

            if seccion == "KG":
                cantidad_kg = valor_final
                importe = round(valor_final * precio_final, 2)
            else:
                cantidad_kg = 0.0
                importe = valor_final

            id_pv = None if es_nueva else mapping.get((row_key, mes_num))

            if id_pv:
                updates.append({
                    "id_presupuesto": id_pv,
                    "valor": valor_final,
                    "precio": precio_final,
                    "cantidad_kg": cantidad_kg,
                    "importe": importe,
                    "estatus_excel": estatus_edit if estatus_cambio else None,
                })
            else:
                inserts.append({
                    "id_carga": int(meta.get("id_carga") or 0),
                    "seccion": seccion,
                    "region": meta.get("region") or None,
                    "anio": int(meta.get("anio") or 0),
                    "mes": mes_num,
                    "company": meta.get("company") or None,
                    "cliente_excel": meta.get("cliente_excel") or None,
                    "codigo_origen": meta.get("codigo_origen") or None,
                    "producto_excel": str(meta.get("producto_excel") or ""),
                    "cve_prod": meta.get("cve_prod") or None,
                    "cve_linea": meta.get("cve_linea") or None,
                    "estatus_excel": meta.get("estatus_excel") or None,
                    "precio": precio_final,
                    "valor": valor_final,
                    "cantidad_kg": cantidad_kg,
                    "importe": importe,
                    "usuario_id": usuario_id,
                })

    if not inserts and not updates and not cve_prod_updates and not identidad_updates:
        return 0, errores

    try:
        resultado = guardar_presupuesto_compras_batch_ctrl(
            inserts=inserts,
            updates=updates,
            cve_prod_updates=cve_prod_updates,
            identidad_updates=identidad_updates,
        )
        cambios = (
            resultado.get("insertados", 0)
            + resultado.get("actualizados", 0)
            + resultado.get("cve_prod_actualizados", 0)
            + resultado.get("identidad_actualizados", 0)
        )
    except Exception:
        cambios = 0
        errores += len(inserts) + len(updates) + len(cve_prod_updates) + len(identidad_updates)

    return cambios, errores


# ── panel: selector de carga ──────────────────────────────────────────────────

def _selector_carga() -> Optional[int]:
    df = obtener_cargas_presupuesto_compras_ctrl(limit=50, usuario_id=_get_usuario_id())

    if df is None or df.empty:
        st.info("aún no hay presupuestos cargados")
        return None

    opciones = {
        f"{r['id_carga']} | {r['nombre_archivo']} | {r['anio']} | {r.get('version', '')} | {r.get('comentarios', '') or ''}": int(r["id_carga"])
        for r in df.to_dict(orient="records")
    }
    labels = list(opciones.keys())
    default = st.session_state.get("pc_id_carga")
    idx = next((i for i, l in enumerate(labels) if opciones[l] == default), 0)

    label = st.selectbox("presupuesto", options=labels, index=idx, key="pc_select_carga")
    id_carga = opciones[label]
    st.session_state["pc_id_carga"] = id_carga
    return id_carga


# ── panel: crear presupuesto manual (sin depender de un Excel) ────────────────

def _panel_crear_manual() -> None:
    with st.expander("➕ crear presupuesto manual (sin Excel)"):
        col1, col2, col3 = st.columns([1, 1, 2])
        with col1:
            anio = st.number_input(
                "año", min_value=2020, max_value=2100, value=2026, step=1, key="pc_manual_anio"
            )
        with col2:
            version = st.text_input("versión", value="manual", key="pc_manual_version")
        with col3:
            comentarios = st.text_input("comentarios", value="", key="pc_manual_comentarios")

        if st.button("crear presupuesto manual", key="pc_btn_manual"):
            usuario_id = _get_usuario_id()
            if usuario_id <= 0:
                st.error("no se encontró el usuario en sesión")
                return
            try:
                id_carga = registrar_carga_presupuesto_compras_ctrl(
                    nombre_archivo="Presupuesto manual",
                    hoja_origen="manual",
                    anio=int(anio),
                    version=version or None,
                    comentarios=comentarios or None,
                    usuario_id=usuario_id,
                )
                st.session_state["pc_id_carga"] = int(id_carga)
                st.success(f"presupuesto manual creado — id={id_carga}")
                st.rerun()
            except Exception as e:
                st.error(f"error al crear el presupuesto manual: {e}")


# ── panel: carga de Excel ─────────────────────────────────────────────────────

def _panel_carga() -> None:
    st.download_button(
        "⬇️ descargar layout de carga",
        data=_generar_layout_presupuesto_bytes(),
        file_name="layout_presupuesto_compras.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="pc_btn_layout",
        help="plantilla en blanco con el formato que espera la carga (incluye hoja de instrucciones)",
    )

    archivo = st.file_uploader(
        "sube el archivo de presupuesto compras",
        type=["xlsx", "xls"],
        key="pc_archivo",
    )
    if archivo is None:
        return

    hojas = _obtener_hojas(archivo)
    if not hojas:
        st.error("no fue posible leer las hojas del archivo")
        return

    col1, col2, col3 = st.columns([2, 1, 2])
    with col1:
        hoja = st.selectbox("hoja", options=hojas, key="pc_hoja")
    with col2:
        anio = st.number_input("año", min_value=2020, max_value=2100, value=2026, step=1, key="pc_anio")
    with col3:
        version = st.text_input("versión", value="v1", key="pc_version")

    comentarios = st.text_input("comentarios", value="", key="pc_comentarios")

    if st.button("cargar presupuesto", type="primary", use_container_width=True, key="pc_btn_cargar"):
        usuario_id = _get_usuario_id()
        if usuario_id <= 0:
            st.error("no se encontró el usuario en sesión")
            return
        try:
            archivo.seek(0)
            res = cargar_excel_directo_presupuesto_compras_ctrl(
                archivo=archivo,
                nombre_archivo=archivo.name,
                hoja=hoja,
                anio=int(anio),
                usuario_id=usuario_id,
                version=version or None,
                comentarios=comentarios or None,
                reemplazar=True,
            )
            st.session_state["pc_id_carga"] = int(res["id_carga"])
            st.success(
                f"cargado — id={res['id_carga']} | "
                f"tablas={res['tablas_detectadas']} | "
                f"registros={res['total_registros']}"
            )
            st.rerun()
        except Exception as e:
            st.error(f"error al cargar: {e}")


# ── formulario: agregar registro manual (en vez de fila vacía en la tabla) ────

def _form_agregar_registro(
    *,
    seccion: str,
    region: Optional[str],
    id_carga: int,
    work_key: str,
    cols_id: list[str],
    meses_presentes: list[str],
    sae_opciones: list[str],
    label_to_code: dict,
    code_to_label: dict,
    code_to_desc: dict,
    code_to_precio: dict,
    code_to_linea: dict,
    code_to_origen: dict,
    code_to_unidad: dict,
    sae_set: set,
    clientes_set: set,
    clientes_opciones: list[str],
) -> None:
    prefix = f"pc_man_{seccion}_{region}_{id_carga}"
    with st.expander("➕ agregar registro"):
        st.caption(
            "campos obligatorios: **nombre producto** y **al menos un mes con valor distinto de cero** "
            "— company, cliente y estatus son opcionales; costo, línea y código origen se "
            "obtienen de SAE al elegir el producto y no se pueden editar a mano"
        )
        col1, col2 = st.columns(2)
        with col1:
            company = st.text_input("company (opcional)", key=f"{prefix}_company")
        with col2:
            cli_sel_key = f"{prefix}_cliente_sel"
            cliente_key = f"{prefix}_cliente"

            def _autofill_cliente() -> None:
                label = st.session_state.get(cli_sel_key) or ""
                if label:
                    st.session_state[cliente_key] = label

            st.selectbox(
                "cliente SAE (obligatorio si estatus = Budgeted)",
                clientes_opciones, key=cli_sel_key, on_change=_autofill_cliente,
            )
            cliente = st.text_input(
                "cliente (libre si estatus no es Budgeted)", key=cliente_key,
            )

        sel_key = f"{prefix}_prod_sel"
        nom_key = f"{prefix}_prod_nom"

        def _autofill_producto() -> None:
            code = label_to_code.get(st.session_state.get(sel_key) or "", "")
            st.session_state[nom_key] = code_to_desc.get(code, "") if code else ""

        cve_label = st.selectbox(
            "producto SAE (opcional — autocompleta nombre, costo, línea y código origen)",
            sae_opciones, key=sel_key, on_change=_autofill_producto,
        )
        producto_excel = st.text_input("nombre producto *obligatorio*", key=nom_key)

        # precio, línea y código origen son solo lectura y dependen únicamente
        # del producto seleccionado: se recalculan en cada render (en vez de
        # guardarse en session_state vía on_change) para que nunca queden
        # desincronizados de lo que el selectbox muestra, sin importar qué
        # otro widget del formulario haya disparado el rerun (p. ej. el
        # cliente SAE)
        _code_sel = label_to_code.get(cve_label or "", "")
        precio = float(code_to_precio.get(_code_sel, 0.0)) if _code_sel else 0.0
        linea_desc = code_to_linea.get(_code_sel, ("", ""))[1] if _code_sel else ""
        codigo_origen = code_to_origen.get(_code_sel, "") if _code_sel else ""
        unidad_sel = (code_to_unidad.get(_code_sel, "") if _code_sel else "") or "unidad"

        col4, col5, col6, col7 = st.columns(4)
        with col4:
            estatus_excel = st.selectbox(
                "estatus (opcional)", ["", "Budgeted", "Not in BGT", "Prospecto"], key=f"{prefix}_estatus",
            )
        with col5:
            st.number_input(
                f"costo USD/{unidad_sel} (ult. costo SAE / tipo de cambio Banxico, no editable)",
                min_value=0.0, format="%.4f", value=precio, disabled=True,
                key=f"{prefix}_precio_{_code_sel or 'none'}",
            )
        with col6:
            st.text_input(
                "línea (de SAE, no editable)", value=linea_desc, disabled=True,
                key=f"{prefix}_linea_{_code_sel or 'none'}",
            )
        with col7:
            st.text_input(
                "código origen (de SAE, no editable)", value=codigo_origen, disabled=True,
                key=f"{prefix}_origen_{_code_sel or 'none'}",
            )

        # se arma en bloques de hasta 6 columnas por fila para mantener el
        # orden ene→dic; los 12 meses están disponibles sin restricción
        st.caption("*obligatorio* captura al menos un mes con valor distinto de cero")
        fmt = "%.4f" if seccion == "KG" else "%.2f"
        meses_form = meses_presentes
        valores: dict[str, float] = {}
        CHUNK = 6
        for inicio in range(0, len(meses_form), CHUNK):
            fila_meses = meses_form[inicio:inicio + CHUNK]
            cols = st.columns(len(fila_meses))
            for col, m in zip(cols, fila_meses):
                with col:
                    valores[m] = st.number_input(
                        m.upper(), value=0.0, format=fmt, key=f"{prefix}_val_{m}",
                    )

        if st.button("agregar", key=f"{prefix}_btn"):
            if not producto_excel.strip():
                st.error("ingresa el nombre del producto")
                return
            if not any(abs(v) > 1e-9 for v in valores.values()):
                st.error("captura al menos un mes con valor distinto de cero")
                return
            if estatus_excel == "Budgeted" and cliente.strip() not in clientes_set:
                st.error(
                    "con estatus \"Budgeted\" el cliente debe ser uno del catálogo SAE — "
                    "selecciónalo en \"cliente SAE\" (Not in BGT / Prospecto no lo requieren)"
                )
                return
            code = label_to_code.get(cve_label, cve_label.strip() or None)
            fila = {c: "" for c in cols_id}
            fila.update({
                "company": company.strip(),
                "cliente_excel": cliente.strip(),
                "codigo_origen": codigo_origen.strip(),
                "producto_excel": producto_excel.strip(),
                "_status": "🟢" if code in sae_set else "🟠",
                "_cve_prod_label": code_to_label.get(code, "") if code else "",
                "estatus_excel": estatus_excel,
                "precio": float(precio),
                "_nueva": True,
                "_estatus_linea": "captura",
                "_estatus_linea_badge": _ESTATUS_LINEA_BADGE["captura"],
            })
            for m in meses_presentes:
                fila[m] = valores.get(m, 0.0)

            # se agrega directo al estado de trabajo, sin tocar ver_key: no
            # hay nada nuevo que recargar de BD, solo mostrar la fila
            st.session_state[work_key] = pd.concat(
                [st.session_state[work_key], pd.DataFrame([fila])], ignore_index=True
            )
            st.success("registro agregado — revísalo en la tabla y guarda los cambios")
            st.rerun()


# ── panel: stock y órdenes de compra por llegar (SAE, solo lectura) ───────────

def _panel_stock_ordenes_sae() -> None:
    with st.container(border=True):
        st.markdown("**📦 stock y órdenes de compra por llegar (SAE)**")

        st.caption("stock de productos")
        df_stock = obtener_existencias_productos_pv_compras_ctrl()
        if df_stock is None or df_stock.empty:
            st.info("sin datos de existencias en SAE")
        else:
            filtro = st.text_input(
                "buscar producto (clave o nombre)", key="pc_stock_filtro",
            )
            df_mostrar = df_stock
            if filtro.strip():
                f = filtro.strip().lower()
                df_mostrar = df_stock[
                    df_stock["cve_art"].astype(str).str.lower().str.contains(f, na=False)
                    | df_stock["descr"].astype(str).str.lower().str.contains(f, na=False)
                ]
            st.caption(f"{len(df_mostrar):,} producto(s)")
            st.dataframe(
                df_mostrar.rename(columns={
                    "cve_art": "clave", "descr": "producto", "linea": "línea",
                    "uni_med": "unidad", "existencia": "existencia",
                    "costo_prom": "costo prom.", "ult_costo": "último costo",
                })[["clave", "producto", "línea", "unidad", "existencia", "costo prom.", "último costo"]],
                use_container_width=True, height=300, hide_index=True,
            )

        st.caption("órdenes de compra por llegar")
        df_oc = obtener_ordenes_compra_pendientes_pv_compras_ctrl()
        if df_oc is None or df_oc.empty:
            st.info("sin órdenes de compra pendientes de recibir en SAE")
        else:
            st.caption(f"{len(df_oc):,} línea(s) de orden de compra")
            st.dataframe(
                df_oc.rename(columns={
                    "cve_doc": "orden", "fecha_doc": "fecha", "fecha_rec": "fecha requerida",
                    "proveedor": "proveedor", "cve_art": "clave", "producto": "producto",
                    "linea": "línea", "cantidad": "cantidad", "unidad": "unidad", "precio": "precio",
                })[[
                    "orden", "fecha", "fecha requerida", "proveedor",
                    "clave", "producto", "línea", "cantidad", "unidad", "precio",
                ]],
                use_container_width=True, height=300, hide_index=True,
            )


# ── panel: tabla pivot ────────────────────────────────────────────────────────

def _panel_pivot(id_carga: int) -> None:
    df_all = obtener_presupuesto_compras_ctrl(id_carga=id_carga)
    if df_all is None:
        df_all = pd.DataFrame()

    anio_default = None
    if df_all.empty:
        # presupuesto sin registros aún (p. ej. creado de forma manual, sin Excel):
        # se arma una estructura vacía con las columnas base para permitir captura manual
        df_all = pd.DataFrame(columns=_COLS_ID + [
            "mes", "anio", "seccion", "region", "valor", "importe",
            "cantidad_kg", "precio", "cve_prod", "estatus_excel",
            "id_carga", "id_presupuesto",
        ])
        carga_meta = obtener_cargas_presupuesto_compras_ctrl(id_carga=id_carga, limit=1)
        if carga_meta is not None and not carga_meta.empty and "anio" in carga_meta.columns:
            anio_default = int(carga_meta.iloc[0]["anio"])

    for col in ("valor", "importe", "cantidad_kg", "precio"):
        if col in df_all.columns:
            df_all[col] = pd.to_numeric(df_all[col], errors="coerce").fillna(0.0)
        else:
            df_all[col] = 0.0

    if df_all["valor"].eq(0).all() and "importe" in df_all.columns:
        df_all["valor"] = df_all["importe"]

    df_all["mes"] = pd.to_numeric(df_all["mes"], errors="coerce").fillna(0).astype(int)

    # catálogo SAE (cacheado 1 hora)
    sae_set, code_to_label, label_to_code, code_to_desc, code_to_precio, code_to_linea, code_to_origen, code_to_unidad, sae_opciones = _catalogo_sae()
    clientes_set, clientes_opciones = _catalogo_clientes_sae()

    if "anio" in df_all.columns and not df_all.empty:
        anio_actual = int(df_all["anio"].iloc[0])
    else:
        anio_actual = int(anio_default or pd.Timestamp.today().year)

    # se muestran siempre los 12 meses (aunque no tengan datos aún) para poder
    # capturar cualquier mes al agregar o completar un registro
    meses_todos = list(_MESES.values())

    col_exp1, col_exp2 = st.columns([1, 3])
    with col_exp1:
        export_key = f"pc_export_bytes_{id_carga}"
        if st.button("📊 generar Excel", key=f"pc_export_gen_{id_carga}", use_container_width=True):
            hojas = []
            for label, seccion, region in _TABS_PIVOT:
                mask = pd.Series(True, index=df_all.index)
                if "seccion" in df_all.columns:
                    mask &= df_all["seccion"].astype(str) == seccion
                if region and "region" in df_all.columns:
                    mask &= df_all["region"].astype(str) == region
                df_sec_exp = df_all[mask].copy()
                pivot_exp, _, _ = _construir_pivot(df_sec_exp, sae_set, code_to_label)
                hojas.append((label, seccion, pivot_exp))
            st.session_state[export_key] = _pivot_a_excel_bytes(hojas)

        if st.session_state.get(export_key):
            st.download_button(
                "⬇️ descargar Excel",
                data=st.session_state[export_key],
                file_name=f"presupuesto_compras_{id_carga}_{anio_actual}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"pc_export_dl_{id_carga}",
            )

    sub_tabs = st.tabs([t[0] for t in _TABS_PIVOT])

    for tab_ui, (label, seccion, region) in zip(sub_tabs, _TABS_PIVOT):
        with tab_ui:
            # `orig_key` guarda la última copia confiable de BD (pivot + mapping +
            # row_meta), y solo se reconstruye cuando `ver_key` cambia (justo
            # después de agregar/guardar/eliminar). `work_key` es el estado que
            # se edita en vivo: se pasa como `data=` a AgGrid en cada render y se
            # vuelve a guardar con lo que el grid reporte, para que un rerun
            # disparado por la edición de OTRA celda nunca le reenvíe al grid un
            # estado viejo y revierta lo que el usuario acaba de escribir.
            ver_key = f"pc_ver_{seccion}_{region}_{id_carga}"
            orig_key = f"pc_orig_{seccion}_{region}_{id_carga}"
            work_key = f"pc_work_{seccion}_{region}_{id_carga}"
            st.session_state.setdefault(ver_key, 0)

            necesita_reload = (
                orig_key not in st.session_state
                or st.session_state[orig_key]["ver"] != st.session_state[ver_key]
            )

            if necesita_reload:
                mask = pd.Series(True, index=df_all.index)
                if "seccion" in df_all.columns:
                    mask &= df_all["seccion"].astype(str) == seccion
                if region and "region" in df_all.columns:
                    mask &= df_all["region"].astype(str) == region
                df_sec = df_all[mask].copy()

                cols_id_reload = [c for c in _COLS_ID if c in df_all.columns]
                pivot_db, mapping_db, row_meta_db = _construir_pivot(df_sec, sae_set, code_to_label)

                for m in meses_todos:
                    if m not in pivot_db.columns:
                        pivot_db[m] = 0.0

                col_order = (
                    cols_id_reload + ["_status", "_cve_prod_label", "estatus_excel", "precio"] + meses_todos
                )
                pivot_db = pivot_db[[c for c in col_order if c in pivot_db.columns]]
                pivot_db["_nueva"] = False

                # estatus de autorización por línea (no por mes): si no hay
                # fila en presupuesto_compras_lineas para esa combinación
                # company/cliente/código/producto, es "captura" implícito
                lineas_df = obtener_presupuesto_compras_lineas_ctrl(id_carga)
                estatus_por_linea: dict[tuple, str] = {}
                if lineas_df is not None and not lineas_df.empty:
                    for r in lineas_df.to_dict("records"):
                        clave = tuple(
                            str(r.get(c) or "") for c in
                            ("company", "cliente_excel", "codigo_origen", "producto_excel")
                        )
                        estatus_por_linea[clave] = str(r.get("estatus") or "captura")

                def _estatus_linea_de(row):
                    clave = tuple(str(row.get(c) or "") for c in cols_id_reload)
                    return estatus_por_linea.get(clave, "captura")

                pivot_db["_estatus_linea"] = (
                    pivot_db.apply(_estatus_linea_de, axis=1) if not pivot_db.empty else "captura"
                )
                pivot_db["_estatus_linea_badge"] = pivot_db["_estatus_linea"].map(_ESTATUS_LINEA_BADGE).fillna("🔵 captura")

                # la columna de autorización va al inicio de la tabla
                pivot_db = pivot_db[
                    ["_estatus_linea_badge"] + [c for c in pivot_db.columns if c != "_estatus_linea_badge"]
                ]

                st.session_state[orig_key] = {
                    "ver": st.session_state[ver_key],
                    "pivot": pivot_db,
                    "mapping": mapping_db,
                    "row_meta": row_meta_db,
                    "cols_id": cols_id_reload,
                }
                st.session_state[work_key] = pivot_db.copy()

            orig_data = st.session_state[orig_key]
            pivot = orig_data["pivot"]
            mapping = orig_data["mapping"]
            row_meta = orig_data["row_meta"]
            cols_id = orig_data["cols_id"]
            meses_presentes = meses_todos

            decimales = 2 if seccion == "USD" else 4

            _form_agregar_registro(
                seccion=seccion,
                region=region,
                id_carga=id_carga,
                work_key=work_key,
                cols_id=cols_id,
                meses_presentes=meses_presentes,
                sae_opciones=sae_opciones,
                label_to_code=label_to_code,
                code_to_label=code_to_label,
                code_to_desc=code_to_desc,
                code_to_precio=code_to_precio,
                code_to_linea=code_to_linea,
                code_to_origen=code_to_origen,
                code_to_unidad=code_to_unidad,
                sae_set=sae_set,
                clientes_set=clientes_set,
                clientes_opciones=clientes_opciones,
            )

            work_df = st.session_state[work_key]
            if work_df.empty:
                st.info(f"sin datos para {label} — usa \"➕ agregar registro\" para capturar uno")
                continue

            # única restricción de edición: una línea deja de ser editable en
            # cuanto se solicita autorización ("enviada") y mientras esté
            # "autorizada" — "captura" y "rechazada" permiten editar
            # cualquier campo sin restricción
            _NOT_FROZEN = (
                "!(params.data && (params.data._estatus_linea === 'enviada'"
                " || params.data._estatus_linea === 'autorizada'))"
            )
            editable_no_congelada = JsCode(f"function(params) {{ return {_NOT_FROZEN}; }}")

            gb = GridOptionsBuilder.from_dataframe(work_df)
            gb.configure_default_column(editable=False, resizable=True, width=100)
            gb.configure_selection("multiple", use_checkbox=True, header_checkbox=True)

            # cliente, producto y código origen quedan fijos una vez
            # capturado el registro (cliente y producto se definen al
            # agregarlo, código origen viene de SAE) — solo "company"
            # permanece editable en la tabla
            _CAMPOS_ID_BLOQUEADOS = {"producto_excel", "cliente_excel", "codigo_origen"}
            for c in cols_id:
                gb.configure_column(
                    c, headerName=c.replace("_excel", "").replace("_", " "),
                    editable=False if c in _CAMPOS_ID_BLOQUEADOS else editable_no_congelada,
                    width=130,
                )

            gb.configure_column("_nueva", hide=True)
            gb.configure_column("_estatus_linea", hide=True)
            gb.configure_column(
                "_estatus_linea_badge", headerName="autorización", editable=False, width=120,
                headerTooltip="🔵 captura  |  🟡 enviada (esperando autorización)  |  🟢 autorizada  |  🔴 rechazada",
            )
            gb.configure_column(
                "_status", headerName="SAE", editable=False, width=70,
                headerTooltip="🟢 producto en catálogo SAE  |  🟠 no encontrado en SAE",
            )
            gb.configure_column(
                "_cve_prod_label",
                headerName="cve prod",
                # producto SAE: fijo una vez capturado, igual que costo y línea —
                # se elige al agregar el registro, no se reasigna después
                editable=False,
                width=200,
            )
            gb.configure_column(
                "estatus_excel", headerName="status", editable=editable_no_congelada, width=110,
                cellEditor="agSelectCellEditor",
                cellEditorParams={"values": ["", "Budgeted", "Not in BGT", "Prospecto"]},
            )
            gb.configure_column(
                "precio",
                headerName="costo USD/unidad",
                # costo: ult_costo SAE / tipo de cambio Banxico al agregar el registro, no editable en la tabla
                editable=False,
                width=110,
                type=["numericColumn"],
                valueFormatter=_value_formatter_js(4),
            )
            for m in meses_presentes:
                gb.configure_column(
                    m,
                    headerName=m.upper(),
                    editable=editable_no_congelada,
                    width=90,
                    type=["numericColumn"],
                    cellEditor="agNumberCellEditor",
                    cellStyle=_CELL_STYLE_VALORES,
                    valueFormatter=_value_formatter_js(decimales),
                )

            # totales al final — solo lectura, calculados en vivo a partir de
            # los meses ya capturados en el estado de trabajo (no se guardan
            # como columnas propias del work_key, solo se muestran)
            work_df_mostrar = work_df.copy()
            suma_meses = (
                work_df_mostrar[meses_presentes].apply(pd.to_numeric, errors="coerce").fillna(0.0).sum(axis=1)
                if meses_presentes else pd.Series(0.0, index=work_df_mostrar.index)
            )
            precio_num = pd.to_numeric(work_df_mostrar.get("precio"), errors="coerce").fillna(0.0)
            if seccion == "KG":
                work_df_mostrar["total_kg_anio"] = suma_meses
                work_df_mostrar["total_usd_anio"] = suma_meses * precio_num
            else:
                work_df_mostrar["total_kg_anio"] = 0.0
                work_df_mostrar["total_usd_anio"] = suma_meses

            gb.configure_column(
                "total_kg_anio", headerName="Total Kg", editable=False, width=120,
                type=["numericColumn"], valueFormatter=_value_formatter_js(4),
            )
            gb.configure_column(
                "total_usd_anio", headerName="Total USD", editable=False, width=120,
                type=["numericColumn"], valueFormatter=_value_formatter_js(2),
            )

            st.caption(
                "🟢 en SAE  |  🟠 no en SAE  |  🟩 valor positivo  |  🟥 valor negativo "
                " — cliente, producto, cve prod, código origen y costo quedan fijos desde que se agrega"
                " el registro  |  selecciona filas con el checkbox para eliminarlas o para solicitar su"
                " autorización  |  🔵 captura 🟡 enviada 🟢 autorizada 🔴 rechazada — mientras esté en"
                " captura o rechazada puedes editar company, estatus y los meses; una vez enviada o"
                " autorizada ya no se puede editar nada"
            )

            grid_response = AgGrid(
                work_df_mostrar,
                gridOptions=gb.build(),
                update_on=[("cellValueChanged", 600), "selectionChanged"],
                data_return_mode=DataReturnMode.AS_INPUT,
                fit_columns_on_grid_load=False,
                allow_unsafe_jscode=True,
                height=min(56 + len(work_df) * 35, 680),
                key=f"pc_pivot_{seccion}_{region}_{id_carga}_{st.session_state[ver_key]}",
            )
            edited = pd.DataFrame(grid_response.get("data", []))
            edited = edited.drop(columns=["total_kg_anio", "total_usd_anio"], errors="ignore")

            # se guarda de inmediato lo que el grid reportó como estado de
            # trabajo, para que el próximo rerun (disparado por la edición de
            # OTRA celda) siga desde aquí en vez de reconstruir desde BD y
            # revertir lo que el usuario acaba de escribir
            if not edited.empty:
                st.session_state[work_key] = edited

            # al elegir cve_prod en una fila nueva, se llena "producto" con el
            # nombre del producto del catálogo SAE
            if not edited.empty and "_cve_prod_label" in edited.columns and "producto_excel" in edited.columns:
                hubo_cambio = False
                work_actual = st.session_state[work_key].copy()
                for i in range(len(work_actual)):
                    if not bool(work_actual.iloc[i].get("_nueva")):
                        continue
                    lbl = str(work_actual.iloc[i].get("_cve_prod_label") or "").strip()
                    if not lbl:
                        continue
                    code = label_to_code.get(lbl, lbl)
                    nombre = code_to_desc.get(code, "")
                    if nombre and str(work_actual.iloc[i].get("producto_excel") or "") != nombre:
                        work_actual.at[i, "producto_excel"] = nombre
                        hubo_cambio = True
                    if "_status" in work_actual.columns and code in sae_set:
                        work_actual.at[i, "_status"] = "🟢"

                if hubo_cambio:
                    # se mutó el estado de trabajo directamente, sin tocar
                    # ver_key: no hace falta recargar de BD, solo redibujar
                    st.session_state[work_key] = work_actual
                    edited = work_actual
                    st.rerun()

            seleccionadas = grid_response.get("selected_rows")
            if seleccionadas is None:
                seleccionadas = []
            elif isinstance(seleccionadas, pd.DataFrame):
                seleccionadas = seleccionadas.to_dict("records")

            col_save, col_del, col_aut = st.columns(3)
            with col_save:
                if st.button(
                    "💾 guardar cambios",
                    type="primary",
                    use_container_width=True,
                    key=f"pc_save_{seccion}_{region}_{id_carga}",
                ):
                    # `_guardar_pivot` compara `orig`/`edited` fila por fila
                    # posicionalmente; `pivot` (línea base de BD) no incluye las
                    # filas nuevas que solo viven en `work_key`, así que se
                    # rellena con placeholders "_nueva" para igualar el largo
                    # antes de diffear.
                    work_actual_save = st.session_state[work_key]
                    n_extra = len(work_actual_save) - len(pivot)
                    if n_extra > 0:
                        relleno = pd.DataFrame(
                            [{**{c: "" for c in cols_id}, "_nueva": True}] * n_extra
                        )
                        orig_para_diff = pd.concat([pivot, relleno], ignore_index=True)
                    else:
                        orig_para_diff = pivot

                    cambios, errores = _guardar_pivot(
                        orig_para_diff, work_actual_save, mapping, row_meta,
                        seccion, region, cols_id, _get_usuario_id(), label_to_code,
                        id_carga, anio_actual, code_to_linea,
                    )
                    if cambios:
                        st.success(f"guardados {cambios} registros")
                        # se recarga desde BD: descarta filas nuevas vacías y
                        # refleja lo recién guardado como la nueva línea base
                        st.session_state[ver_key] += 1
                        st.rerun()
                    if errores:
                        st.error(f"{errores} filas con error")
                    if not cambios and not errores:
                        st.info("sin cambios detectados")
            with col_del:
                if st.button(
                    "🗑️ eliminar seleccionados",
                    use_container_width=True,
                    disabled=not seleccionadas,
                    key=f"pc_del_{seccion}_{region}_{id_carga}",
                ):
                    # una línea enviada o autorizada ya no se puede eliminar
                    congeladas = [
                        f for f in seleccionadas
                        if str(f.get("_estatus_linea") or "captura") in ("enviada", "autorizada")
                    ]
                    seleccionadas = [
                        f for f in seleccionadas
                        if str(f.get("_estatus_linea") or "captura") not in ("enviada", "autorizada")
                    ]
                    if congeladas:
                        st.warning(
                            f"{len(congeladas)} línea(s) enviada(s)/autorizada(s) no se eliminaron "
                            "(ya no se pueden editar ni eliminar)."
                        )

                    registros_borrados = 0
                    filas_bd = 0
                    claves_borradas = {
                        tuple(str(f.get(c) or "").strip() for c in cols_id)
                        for f in seleccionadas
                    }
                    nuevas_borradas = [
                        f for f in seleccionadas if f.get("_nueva")
                    ]
                    for fila in seleccionadas:
                        if fila.get("_nueva"):
                            continue
                        try:
                            n = eliminar_registro_presupuesto_compras_ctrl(
                                id_carga=id_carga,
                                seccion=seccion,
                                region=region,
                                producto_excel=str(fila.get("producto_excel") or ""),
                                cliente_excel=fila.get("cliente_excel") or None,
                                codigo_origen=fila.get("codigo_origen") or None,
                                company=fila.get("company") or None,
                            )
                            filas_bd += n
                            if n:
                                registros_borrados += 1
                        except Exception:
                            pass

                    # se quitan de inmediato del estado de trabajo (filas nuevas
                    # y/o ya guardadas); si hubo borrado real en BD, además se
                    # fuerza una recarga desde BD para que quede consistente
                    work_actual = st.session_state[work_key]
                    mask_fuera = work_actual.apply(
                        lambda r: tuple(str(r.get(c) or "").strip() for c in cols_id) in claves_borradas,
                        axis=1,
                    )
                    st.session_state[work_key] = work_actual[~mask_fuera].reset_index(drop=True)

                    if registros_borrados:
                        st.session_state[ver_key] += 1

                    if registros_borrados or nuevas_borradas:
                        st.success(
                            f"{registros_borrados} registro(s) eliminados ({filas_bd} filas de detalle)"
                            + (f", {len(nuevas_borradas)} fila(s) nueva(s) descartadas" if nuevas_borradas else "")
                        )
                        st.rerun()
                    else:
                        st.info("no se eliminó nada")
            with col_aut:
                elegibles = [
                    f for f in seleccionadas
                    if str(f.get("_estatus_linea") or "captura") in ("captura", "rechazada")
                    and str(f.get("producto_excel") or "").strip()
                ]
                if st.button(
                    "📤 solicitar autorización",
                    use_container_width=True,
                    disabled=not elegibles,
                    key=f"pc_aut_{seccion}_{region}_{id_carga}",
                ):
                    usuario = st.session_state.get("usuario") or {}
                    usuario_id = int(usuario.get("id") or 0)
                    usuario_nombre = str(usuario.get("nombre") or usuario.get("username") or "").strip()
                    usuario_email = str(usuario.get("email") or "").strip()

                    tipo_aut = _tipo_autorizacion_linea()
                    estatus_destino = "autorizada" if tipo_aut == "sin_autorizacion" else "enviada"

                    for fila in elegibles:
                        _cambiar_estatus_linea_compras(
                            id_carga=id_carga,
                            company=fila.get("company") or None,
                            cliente_excel=fila.get("cliente_excel") or None,
                            codigo_origen=fila.get("codigo_origen") or None,
                            producto_excel=str(fila.get("producto_excel") or "").strip(),
                            estatus_nuevo=estatus_destino,
                            usuario_id=usuario_id,
                            usuario_nombre=usuario_nombre,
                            usuario_email=usuario_email,
                        )

                    if estatus_destino == "enviada":
                        nombre_rol = _rol_autorizador_linea()
                        token = st.session_state.get("microsoft_token")
                        ok_mail, msg_mail = _enviar_notificacion_autorizador_compras(
                            lineas=elegibles, id_carga=id_carga, anio=anio_actual,
                            nombre_rol=nombre_rol, token=token, remitente=usuario_email,
                        )
                        st.success(f"{len(elegibles)} línea(s) enviada(s) a autorización ({nombre_rol})")
                        if not ok_mail:
                            st.warning(f"no se pudo enviar el correo de notificación: {msg_mail}")
                    else:
                        st.success(f"{len(elegibles)} línea(s) autorizada(s) automáticamente (tu rol no requiere autorización)")

                    st.session_state[ver_key] += 1
                    st.rerun()

    _panel_stock_ordenes_sae()


# ── panel: gestión de cargas ──────────────────────────────────────────────────

def _panel_gestionar_cargas() -> None:
    df = obtener_cargas_presupuesto_compras_ctrl(limit=200, usuario_id=_get_usuario_id())

    if df is None or df.empty:
        st.info("no hay cargas registradas")
        return

    cols_mostrar = [c for c in
                    ["id_carga", "nombre_archivo", "anio", "version", "estatus", "comentarios", "created_at"]
                    if c in df.columns]
    st.dataframe(df[cols_mostrar], use_container_width=True, hide_index=True)

    st.divider()
    st.markdown("#### eliminar carga")
    st.caption("esto borra la carga, sus registros de presupuesto y el staging asociado de forma permanente")

    opciones = {
        f"{r['id_carga']} | {r['nombre_archivo']} | {r['anio']} | {r.get('version', '')}": int(r["id_carga"])
        for r in df.to_dict(orient="records")
    }
    label = st.selectbox("selecciona la carga a eliminar", options=list(opciones.keys()), key="pc_gc_select")
    id_sel = opciones[label]

    confirmar = st.checkbox(f"confirmo que quiero eliminar la carga {id_sel}", key="pc_gc_confirmar")

    if st.button("🗑️ eliminar carga", type="primary", disabled=not confirmar, key="pc_gc_btn_eliminar"):
        try:
            eliminar_carga_completa_presupuesto_compras_ctrl(id_sel)
            st.success(f"carga {id_sel} eliminada correctamente")
            if st.session_state.get("pc_id_carga") == id_sel:
                del st.session_state["pc_id_carga"]
            st.rerun()
        except Exception as e:
            st.error(f"error al eliminar: {e}")


