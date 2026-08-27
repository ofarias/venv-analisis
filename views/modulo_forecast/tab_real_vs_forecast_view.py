from __future__ import annotations

from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from controllers.forecast_controller import (
    obtener_forecast_detalle_ctrl,
    obtener_presupuesto_resumen_por_anio_ctrl,
    obtener_presupuesto_compras_resumen_por_anio_ctrl,
    obtener_presupuesto_finanzas_resumen_por_anio_ctrl,
    _catalogo_productos_sae,
    _compras_historicas_sae,
    _ventas_historicas_sae,
)


_MESES = {1:"ene",2:"feb",3:"mar",4:"abr",5:"may",6:"jun",
          7:"jul",8:"ago",9:"sep",10:"oct",11:"nov",12:"dic"}

_TABS_SEC = [
    ("KG México",        "KG",  "MEXICO"),
]

# series disponibles para comparar; "mes_suffix" son las columnas que ya arma
# _construir_comparativo ({mes}_real / {mes}_real_compras / {mes}_fc / {mes}_pv / {mes}_pc / {mes}_pf / {mes}_real_filt)
_SERIES_META = {
    "real":          {"label": "Real (SAE)",              "total_col": "total_real",          "mes_suffix": "_real"},
    "real_filtrado": {"label": "Real (SAE Filtrado)",      "total_col": "total_real_filtrado", "mes_suffix": "_real_filt"},
    "real_compras":  {"label": "Real Compras (SAE)",       "total_col": "total_real_compras",  "mes_suffix": "_real_compras"},
    "fc":            {"label": "Forecast",                 "total_col": "total_fc",             "mes_suffix": "_fc"},
    "pv":            {"label": "Presupuesto Ventas",        "total_col": "total_presupuesto",   "mes_suffix": "_pv"},
    "pc":            {"label": "Presupuesto Compras",       "total_col": "total_pc",             "mes_suffix": "_pc"},
    "pf":            {"label": "Presupuesto Finanzas",      "total_col": "total_pf",             "mes_suffix": "_pf"},
}
_SERIES_ORDEN = ["real", "real_filtrado", "real_compras", "fc", "pv", "pc", "pf"]


def _col_sae(seccion: str) -> str:
    return "cantidad" if seccion == "KG" else "importe"


def _col_pv(seccion: str) -> str:
    return "total_kg" if seccion == "KG" else "total_importe"


# presupuesto ventas y compras comparten los mismos nombres de columna
# generados por _resumir_presupuesto (total_kg / total_importe)
_col_pc = _col_pv


def _construir_comparativo(
    df_sae: pd.DataFrame,
    df_sae_compras: pd.DataFrame,
    df_fc: pd.DataFrame,
    df_pv: pd.DataFrame,
    df_pc: pd.DataFrame,
    df_pf: pd.DataFrame,
    seccion: str,
    meses: list[int],
) -> pd.DataFrame:
    """
    Retorna pivot wide:
      cve_prod | producto | total_real | total_real_filtrado | total_real_compras |
      total_fc | total_presupuesto | total_pc | total_pf | cumplimiento_%
      + ene_real | ene_real_compras | ene_fc | ene_pv | ene_pc | ene_pf | ene_Δ% | feb_real | ...

    total_real_filtrado / {mes}_real_filt tienen el mismo valor que total_real
    / {mes}_real (el real de SAE) fila por fila — lo que cambia con "Real (SAE
    Filtrado)" es el conjunto de filas mostradas, filtrado en el llamador
    (mostrar_tab_real_vs_forecast) a los códigos presentes en las demás series
    elegidas para comparar (Forecast/Presupuesto Ventas/Compras/Finanzas), en
    vez de incluir también los códigos que solo existen en SAE.

    total_real_compras / {mes}_real_compras vienen de compras ya recibidas en
    SAE (compc01/par_compc01) — es la referencia correcta al comparar contra
    Presupuesto Compras (real vs real, no ventas vs compras); ver
    _referencia_para en el llamador.
    """
    col_sae = _col_sae(seccion)
    col_pv = _col_pv(seccion)
    col_pc = _col_pc(seccion)

    # ── agrupar ventas SAE por cve_art × mes ──────────────────────────────────
    if df_sae is not None and not df_sae.empty and "cve_art" in df_sae.columns:
        df_sae = df_sae.copy()
        df_sae["mes"] = pd.to_numeric(df_sae["mes"], errors="coerce").astype(int)
        df_sae[col_sae] = pd.to_numeric(df_sae[col_sae], errors="coerce").fillna(0.0)
        sae_grp = df_sae.groupby(["cve_art", "mes"], as_index=False).agg(
            real=(col_sae, "sum"),
            producto=("producto", "first"),
        )
        sae_grp = sae_grp[sae_grp["mes"].isin(meses)]
    else:
        sae_grp = pd.DataFrame(columns=["cve_art", "mes", "real", "producto"])

    # ── agrupar compras SAE por cve_art × mes ─────────────────────────────────
    if df_sae_compras is not None and not df_sae_compras.empty and "cve_art" in df_sae_compras.columns:
        df_sae_compras = df_sae_compras.copy()
        df_sae_compras["mes"] = pd.to_numeric(df_sae_compras["mes"], errors="coerce").astype(int)
        df_sae_compras[col_sae] = pd.to_numeric(df_sae_compras[col_sae], errors="coerce").fillna(0.0)
        sae_compras_grp = df_sae_compras.groupby(["cve_art", "mes"], as_index=False).agg(
            real_compras=(col_sae, "sum"),
        )
        sae_compras_grp = sae_compras_grp[sae_compras_grp["mes"].isin(meses)]
    else:
        sae_compras_grp = pd.DataFrame(columns=["cve_art", "mes", "real_compras"])

    # ── agrupar forecast por cve_prod × mes ───────────────────────────────────
    if df_fc is not None and not df_fc.empty:
        df_fc = df_fc.copy()
        df_fc["mes"] = pd.to_numeric(df_fc["mes"], errors="coerce").astype(int)
        df_fc["forecast"] = pd.to_numeric(df_fc["forecast"], errors="coerce").fillna(0.0)
        fc_grp = df_fc[df_fc["mes"].isin(meses)].groupby(["cve_prod", "mes"], as_index=False).agg(
            fc=("forecast", "sum"),
            producto_fc=("producto_excel", "first"),
        )
    else:
        fc_grp = pd.DataFrame(columns=["cve_prod", "mes", "fc", "producto_fc"])

    # ── agrupar presupuesto de ventas por cve_prod × mes ──────────────────────
    if df_pv is not None and not df_pv.empty and "cve_prod" in df_pv.columns:
        df_pv = df_pv.copy()
        df_pv["mes"] = pd.to_numeric(df_pv["mes"], errors="coerce").astype(int)
        df_pv[col_pv] = pd.to_numeric(df_pv[col_pv], errors="coerce").fillna(0.0)
        pv_grp = df_pv[df_pv["mes"].isin(meses)].groupby(["cve_prod", "mes"], as_index=False).agg(
            pv=(col_pv, "sum"),
            producto_pv=("producto_excel", "first"),
        )
    else:
        pv_grp = pd.DataFrame(columns=["cve_prod", "mes", "pv", "producto_pv"])

    # ── agrupar presupuesto de compras por cve_prod × mes ─────────────────────
    if df_pc is not None and not df_pc.empty and "cve_prod" in df_pc.columns:
        df_pc = df_pc.copy()
        df_pc["mes"] = pd.to_numeric(df_pc["mes"], errors="coerce").astype(int)
        df_pc[col_pc] = pd.to_numeric(df_pc[col_pc], errors="coerce").fillna(0.0)
        pc_grp = df_pc[df_pc["mes"].isin(meses)].groupby(["cve_prod", "mes"], as_index=False).agg(
            pc=(col_pc, "sum"),
            producto_pc=("producto_excel", "first"),
        )
    else:
        pc_grp = pd.DataFrame(columns=["cve_prod", "mes", "pc", "producto_pc"])

    # ── agrupar presupuesto de finanzas por cve_prod (clave sku) × mes ────────
    if df_pf is not None and not df_pf.empty and "cve_prod" in df_pf.columns:
        df_pf = df_pf.copy()
        df_pf["mes"] = pd.to_numeric(df_pf["mes"], errors="coerce").astype(int)
        df_pf[col_pv] = pd.to_numeric(df_pf[col_pv], errors="coerce").fillna(0.0)
        pf_grp = df_pf[df_pf["mes"].isin(meses)].groupby(["cve_prod", "mes"], as_index=False).agg(
            pf=(col_pv, "sum"),
            producto_pf=("producto_excel", "first"),
        )
    else:
        pf_grp = pd.DataFrame(columns=["cve_prod", "mes", "pf", "producto_pf"])

    # ── unión por cve_prod = cve_art ──────────────────────────────────────────
    sae_grp = sae_grp.rename(columns={"cve_art": "cve_prod"})
    sae_compras_grp = sae_compras_grp.rename(columns={"cve_art": "cve_prod"})
    merged = pd.merge(
        fc_grp, sae_grp,
        on=["cve_prod", "mes"],
        how="outer",
    )
    merged = pd.merge(
        merged, sae_compras_grp,
        on=["cve_prod", "mes"],
        how="outer",
    )
    merged = pd.merge(
        merged, pv_grp,
        on=["cve_prod", "mes"],
        how="outer",
    )
    merged = pd.merge(
        merged, pc_grp,
        on=["cve_prod", "mes"],
        how="outer",
    )
    merged = pd.merge(
        merged, pf_grp,
        on=["cve_prod", "mes"],
        how="outer",
    )
    merged = merged.fillna({"real": 0.0, "real_compras": 0.0, "fc": 0.0, "pv": 0.0, "pc": 0.0, "pf": 0.0})

    # nombre de producto: preferir el de SAE, fallback forecast, presupuesto, pres. finanzas
    merged["producto"] = (
        merged["producto"].fillna(merged["producto_fc"]).fillna(merged["producto_pv"])
        .fillna(merged["producto_pc"]).fillna(merged["producto_pf"]).fillna(merged["cve_prod"])
    )
    merged = merged.drop(columns=["producto_fc", "producto_pv", "producto_pc", "producto_pf"], errors="ignore")
    merged["cve_prod"] = merged["cve_prod"].fillna("").astype(str)

    if merged.empty:
        return pd.DataFrame()

    # ── pivot wide ────────────────────────────────────────────────────────────
    filas: list[dict] = []
    for cve_prod in sorted(merged["cve_prod"].unique()):
        sub = merged[merged["cve_prod"] == cve_prod]
        producto = sub["producto"].iloc[0] if not sub.empty else cve_prod
        fila: dict = {"cve_prod": cve_prod, "producto": producto}

        total_real = 0.0
        total_real_filtrado = 0.0
        total_real_compras = 0.0
        total_fc   = 0.0
        total_pv   = 0.0
        total_pc   = 0.0
        total_pf   = 0.0
        for mes in meses:
            row = sub[sub["mes"] == mes]
            real = float(row["real"].sum()) if not row.empty else 0.0
            real_compras = float(row["real_compras"].sum()) if not row.empty else 0.0
            fc   = float(row["fc"].sum())   if not row.empty else 0.0
            pv   = float(row["pv"].sum())   if not row.empty else 0.0
            pc   = float(row["pc"].sum())   if not row.empty else 0.0
            pf   = float(row["pf"].sum())   if not row.empty else 0.0
            real_filt = real
            mn = _MESES[mes]
            fila[f"{mn}_real"] = real
            fila[f"{mn}_real_filt"] = real_filt
            fila[f"{mn}_real_compras"] = real_compras
            fila[f"{mn}_fc"]   = fc
            fila[f"{mn}_pv"]   = pv
            fila[f"{mn}_pc"]   = pc
            fila[f"{mn}_pf"]   = pf
            fila[f"{mn}_Δ%"]   = round((real - fc) / fc * 100, 1) if fc != 0 else None
            fila[f"{mn}_Δ%_pv"] = round((real - pv) / pv * 100, 1) if pv != 0 else None
            fila[f"{mn}_Δ%_pc"] = round((real_compras - pc) / pc * 100, 1) if pc != 0 else None
            fila[f"{mn}_Δ%_pf"] = round((real - pf) / pf * 100, 1) if pf != 0 else None
            total_real += real
            total_real_filtrado += real_filt
            total_real_compras += real_compras
            total_fc   += fc
            total_pv   += pv
            total_pc   += pc
            total_pf   += pf

        fila["total_real"] = round(total_real, 2)
        fila["total_real_filtrado"] = round(total_real_filtrado, 2)
        fila["total_real_compras"] = round(total_real_compras, 2)
        fila["total_fc"]   = round(total_fc, 2)
        fila["total_presupuesto"] = round(total_pv, 2)
        fila["total_pc"]   = round(total_pc, 2)
        fila["total_pf"]   = round(total_pf, 2)
        fila["cumplimiento_%"] = round(total_real / total_fc * 100, 1) if total_fc != 0 else None
        fila["cumplimiento_pv_%"] = round(total_real / total_pv * 100, 1) if total_pv != 0 else None
        fila["cumplimiento_pc_%"] = round(total_real_compras / total_pc * 100, 1) if total_pc != 0 else None
        fila["cumplimiento_pf_%"] = round(total_real / total_pf * 100, 1) if total_pf != 0 else None
        filas.append(fila)

    return pd.DataFrame(filas)


def _elegir_comparacion(series_sel: list[str]) -> tuple[str, list[str]]:
    """De las series elegidas, decide cuál es la referencia (baseline) para los
    Δ%/cumplimiento: "real" si está entre las elegidas (mantiene el
    comportamiento histórico de esta pantalla), si no la primera elegida según
    _SERIES_ORDEN."""
    baseline = "real" if "real" in series_sel else next(k for k in _SERIES_ORDEN if k in series_sel)
    otros = [k for k in series_sel if k != baseline]
    return baseline, otros


def _referencia_para(otro: str, baseline: str) -> str:
    """Serie que realmente actúa como referencia al comparar contra `otro`:
    normalmente `baseline`, salvo que la referencia elegida sea "real"
    (ventas facturadas SAE) y lo comparado sea "Presupuesto Compras" — ahí la
    referencia real debe ser compras ya recibidas en SAE ("real_compras"), no
    ventas, porque comparar ventas contra un presupuesto de compras no tiene
    el mismo significado que comparar real vs plan."""
    if baseline == "real" and otro == "pc":
        return "real_compras"
    return baseline


def _agregar_comparaciones(
    df_comp: pd.DataFrame, baseline: str, otros: list[str], meses: list[int]
) -> pd.DataFrame:
    """Agrega columnas total_delta_{otro}_%, total_cumpl_{otro}_% y
    {mes}_delta_{otro}_% para cada serie no-baseline, relativas a la
    referencia efectiva de cada `otro` (ver _referencia_para).
    Si baseline es "real", el numerador es la referencia (igual que el
    cálculo histórico real-vs-plan); si no, el numerador es la serie "otro" y
    el denominador es la referencia."""
    df = df_comp.copy()

    for otro in otros:
        referencia = _referencia_para(otro, baseline)
        ref_total_col = _SERIES_META[referencia]["total_col"]
        otro_total_col = _SERIES_META[otro]["total_col"]
        num, den = (df[ref_total_col], df[otro_total_col]) if baseline == "real" \
            else (df[otro_total_col], df[ref_total_col])
        # NaN (no pd.NA): pd.NA hace que la columna pase a dtype "object" y
        # rompe tanto la comparación "val >= 0" del estilo de color como la
        # escritura a Excel con openpyxl (ninguno de los dos sabe qué hacer
        # con pd.NA); NaN es un float normal y ambos lo manejan bien.
        den_safe = den.replace(0, float("nan"))
        df[f"total_delta_{otro}_%"] = ((num - den) / den_safe * 100).round(1)
        df[f"total_cumpl_{otro}_%"] = (num / den_safe * 100).round(1)

        for mes in meses:
            mn = _MESES[mes]
            ref_mes_col = f"{mn}{_SERIES_META[referencia]['mes_suffix']}"
            otro_mes_col = f"{mn}{_SERIES_META[otro]['mes_suffix']}"
            if ref_mes_col not in df.columns or otro_mes_col not in df.columns:
                continue
            num_m, den_m = (df[ref_mes_col], df[otro_mes_col]) if baseline == "real" \
                else (df[otro_mes_col], df[ref_mes_col])
            den_m_safe = den_m.replace(0, float("nan"))
            df[f"{mn}_delta_{otro}_%"] = ((num_m - den_m) / den_m_safe * 100).round(1)

    return df


def _style_delta(df: pd.DataFrame) -> pd.io.formats.style.Styler:
    """Colorea columnas Δ%/cumplimiento generadas por _agregar_comparaciones:
    verde si ≥0, rojo si <0 (mismo criterio que usaba la versión anterior,
    fija a 4 columnas, ahora aplicado a las que existan según la selección)."""
    def _color(val):
        # pd.isna() cubre None/NaN/NaT/pd.NA (este último lo produce
        # _agregar_comparaciones cuando el denominador es 0); "val >= 0" con
        # pd.NA lanza TypeError porque su bool() es ambiguo, así que hay que
        # descartarlo ANTES de comparar.
        if pd.isna(val):
            return ""
        return "background-color: #d4edda; color: #155724" if val >= 0 else "background-color: #f8d7da; color: #721c24"

    delta_cols = [c for c in df.columns if ("_delta_" in c or "_cumpl_" in c) and c.endswith("_%")]
    styler = df.style
    for col in delta_cols:
        styler = styler.applymap(_color, subset=[col])
    return styler


def _comparativo_a_excel_bytes(nombre_hoja: str, df: pd.DataFrame, encabezados: dict[str, str]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja[:31]

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    columnas = list(df.columns)
    ws.append([encabezados.get(c, c) for c in columnas])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws.freeze_panes = "A2"

    for _, fila in df.iterrows():
        # openpyxl no acepta NaN/pd.NA como valor de celda ("Cannot convert
        # <NA> to Excel") — se limpian a None (celda en blanco) antes de escribir
        ws.append([None if pd.isna(v) else v for v in fila])

    for idx, col_name in enumerate(columnas, start=1):
        largo = len(str(encabezados.get(col_name, col_name)))
        valores = df[col_name].astype(str).tolist()[:200]
        if valores:
            largo = max(largo, max(len(v) for v in valores))
        ws.column_dimensions[get_column_letter(idx)].width = min(largo + 2, 40)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def mostrar_tab_real_vs_forecast(
    id_version: int,
    anio: int,
    meses: list[int],
    usuario_datos_id: int,
) -> None:
    if not meses:
        st.warning("selecciona meses en el tab de construcción")
        return

    st.caption(
        f"Comparativo de Ventas Reales SAE, Compras Reales SAE, Forecast, Presupuesto de Ventas, "
        f"Presupuesto de Compras y Presupuesto de Finanzas — año {anio}. Elige abajo qué series comparar. "
        "Verde = la serie de referencia ≥ la otra serie comparada. Rojo = lo contrario. "
        "Presupuesto de Ventas y Presupuesto de Compras solo consideran líneas ya **autorizadas**. "
        "Real (SAE Filtrado) solo muestra el real de los productos cuyo código existe tanto en Presupuesto "
        "de Ventas como en Presupuesto de Compras. "
        "Al comparar contra Presupuesto de Compras, la referencia real es \"Real Compras (SAE)\" "
        "(compras ya recibidas), no las ventas facturadas. "
        "⚠️ Presupuesto Finanzas no distingue región (MEXICO / CAM & Caribe): se muestra el mismo total en ambas."
    )

    # ── catálogo SAE para armar opciones de línea/producto ─────────────────────
    df_cat = _catalogo_productos_sae()
    lineas_disponibles = (
        sorted(df_cat["linea"].dropna().astype(str).str.strip().unique())
        if df_cat is not None and not df_cat.empty and "linea" in df_cat.columns
        else []
    )

    # ── fila de filtros: año / línea / producto ────────────────────────────────
    col_anio, col_linea, col_prod = st.columns(3)

    with col_anio:
        anio_sel = st.number_input(
            "año a comparar", min_value=2020, max_value=2030,
            value=anio, step=1, key="rvf_anio",
        )

    with col_linea:
        linea_sel = st.multiselect(
            "línea", options=lineas_disponibles, default=[], key="rvf_linea",
        )

    with col_prod:
        df_cat_prod = df_cat
        if df_cat_prod is not None and not df_cat_prod.empty and "status" in df_cat_prod.columns:
            df_cat_prod = df_cat_prod[df_cat_prod["status"].astype(str).str.strip().str.upper() == "A"]
        if df_cat_prod is not None and not df_cat_prod.empty and linea_sel and "linea" in df_cat_prod.columns:
            df_cat_prod = df_cat_prod[df_cat_prod["linea"].astype(str).str.strip().isin(linea_sel)]

        productos_opciones: dict[str, str] = {}
        if df_cat_prod is not None and not df_cat_prod.empty:
            for _, fila_cat in df_cat_prod.iterrows():
                cve = str(fila_cat["cve_art"]).strip()
                desc = str(fila_cat.get("descr") or "").strip()
                productos_opciones[f"{cve} - {desc}" if desc else cve] = cve

        producto_labels_sel = st.multiselect(
            "producto", options=sorted(productos_opciones.keys()), default=[], key="rvf_producto",
        )
    productos_sel = {productos_opciones[lbl] for lbl in producto_labels_sel}

    # carga datos SAE (cacheados)
    with st.spinner("cargando ventas SAE…"):
        df_sae = _ventas_historicas_sae(int(anio_sel))
    with st.spinner("cargando compras SAE…"):
        df_compras_sae = _compras_historicas_sae(int(anio_sel))

    # filtrar solo el año seleccionado
    if df_sae is not None and not df_sae.empty and "anio" in df_sae.columns:
        df_sae["anio"] = pd.to_numeric(df_sae["anio"], errors="coerce")
        df_sae = df_sae[df_sae["anio"] == int(anio_sel)]
    if df_compras_sae is not None and not df_compras_sae.empty and "anio" in df_compras_sae.columns:
        df_compras_sae["anio"] = pd.to_numeric(df_compras_sae["anio"], errors="coerce")
        df_compras_sae = df_compras_sae[df_compras_sae["anio"] == int(anio_sel)]

    # ── filtro por línea (INV01.LIN_PROD ligado con clin01/LINEAS.ID) ──────────
    productos_linea: set[str] | None = None
    if linea_sel:
        if df_sae is not None and not df_sae.empty and "linea" in df_sae.columns:
            df_sae = df_sae[df_sae["linea"].astype(str).str.strip().isin(linea_sel)]
        if df_compras_sae is not None and not df_compras_sae.empty and "linea" in df_compras_sae.columns:
            df_compras_sae = df_compras_sae[df_compras_sae["linea"].astype(str).str.strip().isin(linea_sel)]
        if df_cat is not None and not df_cat.empty:
            productos_linea = set(
                df_cat.loc[
                    df_cat["linea"].astype(str).str.strip().isin(linea_sel), "cve_art"
                ]
                .astype(str)
                .str.strip()
            )

    # ── filtro por tipo de producto (clave inicia con MP o PT) ─────────────────
    tipo_prod_sel = st.radio(
        "tipo de producto",
        options=["Todos", "Materia Prima (MP)", "Producto Terminado (PT)"],
        horizontal=True, key="rvf_tipo_prod",
    )
    prefijo_tipo = {"Materia Prima (MP)": "B", "Producto Terminado (PT)": "PT"}.get(tipo_prod_sel)

    if prefijo_tipo:
        if df_sae is not None and not df_sae.empty and "cve_art" in df_sae.columns:
            df_sae = df_sae[
                df_sae["cve_art"].astype(str).str.strip().str.upper().str.startswith(prefijo_tipo)
            ]
        if df_compras_sae is not None and not df_compras_sae.empty and "cve_art" in df_compras_sae.columns:
            df_compras_sae = df_compras_sae[
                df_compras_sae["cve_art"].astype(str).str.strip().str.upper().str.startswith(prefijo_tipo)
            ]

    # ── filtro por producto ─────────────────────────────────────────────────────
    if productos_sel:
        if df_sae is not None and not df_sae.empty and "cve_art" in df_sae.columns:
            df_sae = df_sae[df_sae["cve_art"].astype(str).str.strip().isin(productos_sel)]
        if df_compras_sae is not None and not df_compras_sae.empty and "cve_art" in df_compras_sae.columns:
            df_compras_sae = df_compras_sae[df_compras_sae["cve_art"].astype(str).str.strip().isin(productos_sel)]
        productos_linea = productos_sel if productos_linea is None else (productos_linea & productos_sel)

    # ── presupuesto de finanzas (no distingue región: se calcula una sola vez) ──
    with st.spinner("cargando presupuesto de finanzas…"):
        df_pf_all = obtener_presupuesto_finanzas_resumen_por_anio_ctrl(int(anio_sel))
    if productos_linea is not None and df_pf_all is not None and not df_pf_all.empty:
        df_pf_all = df_pf_all[
            df_pf_all["cve_prod"].astype(str).str.strip().isin(productos_linea)
        ]
    if prefijo_tipo and df_pf_all is not None and not df_pf_all.empty:
        df_pf_all = df_pf_all[
            df_pf_all["cve_prod"].astype(str).str.strip().str.upper().str.startswith(prefijo_tipo)
        ]

    # ── qué se va a comparar ────────────────────────────────────────────────────
    labels_a_key = {v["label"]: k for k, v in _SERIES_META.items()}
    sel_labels = st.multiselect(
        "comparar",
        options=[_SERIES_META[k]["label"] for k in _SERIES_ORDEN],
        default=["Real (SAE)", "Forecast"],
        key="rvf_series_sel",
        help=(
            "elige 2 o más series (ej. Forecast vs Real, Presupuesto Ventas vs Presupuesto Finanzas, "
            "Forecast vs Real vs Presupuesto Ventas, …). Si incluyes \"Real (SAE)\" se usa como referencia "
            "para las columnas Δ%/cumplimiento; si no, se usa la primera serie elegida."
        ),
    )
    series_sel = [labels_a_key[l] for l in sel_labels]
    if len(series_sel) < 2:
        st.warning("selecciona al menos 2 series para comparar (ej. Forecast vs Real)")
        return
    baseline, otros = _elegir_comparacion(series_sel)

    sub_tabs = st.tabs([t[0] for t in _TABS_SEC])

    for tab_ui, (label, seccion, region) in zip(sub_tabs, _TABS_SEC):
        with tab_ui:
            df_fc = obtener_forecast_detalle_ctrl(
                id_version=id_version, seccion=seccion, region=region, tipo="venta"
            )
            if productos_linea is not None and df_fc is not None and not df_fc.empty:
                df_fc = df_fc[
                    df_fc["cve_prod"].astype(str).str.strip().isin(productos_linea)
                ]
            if prefijo_tipo and df_fc is not None and not df_fc.empty:
                df_fc = df_fc[
                    df_fc["cve_prod"].astype(str).str.strip().str.upper().str.startswith(prefijo_tipo)
                ]

            # solo líneas autorizadas — un forecast no debe alimentarse (ni
            # compararse) contra presupuesto todavía en captura/enviado/rechazado
            df_pv = obtener_presupuesto_resumen_por_anio_ctrl(
                int(anio_sel), seccion, region, usuario_id=usuario_datos_id, solo_autorizados=True,
            )
            if productos_linea is not None and df_pv is not None and not df_pv.empty:
                df_pv = df_pv[
                    df_pv["cve_prod"].astype(str).str.strip().isin(productos_linea)
                ]
            if prefijo_tipo and df_pv is not None and not df_pv.empty:
                df_pv = df_pv[
                    df_pv["cve_prod"].astype(str).str.strip().str.upper().str.startswith(prefijo_tipo)
                ]

            df_pc = obtener_presupuesto_compras_resumen_por_anio_ctrl(
                int(anio_sel), seccion, region, usuario_id=usuario_datos_id, solo_autorizados=True,
            )
            if productos_linea is not None and df_pc is not None and not df_pc.empty:
                df_pc = df_pc[
                    df_pc["cve_prod"].astype(str).str.strip().isin(productos_linea)
                ]
            if prefijo_tipo and df_pc is not None and not df_pc.empty:
                df_pc = df_pc[
                    df_pc["cve_prod"].astype(str).str.strip().str.upper().str.startswith(prefijo_tipo)
                ]

            if (
                (df_fc is None or df_fc.empty)
                and (df_sae is None or df_sae.empty)
                and (df_compras_sae is None or df_compras_sae.empty)
                and (df_pv is None or df_pv.empty)
                and (df_pc is None or df_pc.empty)
                and (df_pf_all is None or df_pf_all.empty)
            ):
                st.info(f"sin datos para {label}")
                continue

            df_comp = _construir_comparativo(
                df_sae, df_compras_sae, df_fc, df_pv, df_pc, df_pf_all, seccion, meses
            )

            if df_comp is None or df_comp.empty:
                st.info(f"sin productos para comparar en {label}")
                continue

            # "Real (SAE Filtrado)": el universo de filas no es "todo SAE" sino
            # la unión de códigos de las demás series elegidas para comparar
            # (ej. si comparás contra Forecast={C,D,E} y SAE tiene {A,B,C},
            # se muestran C, D y E — A y B se descartan por no estar en
            # Forecast — y el real solo se llena para C, que sí viene de SAE).
            if "real_filtrado" in series_sel:
                fuente_por_serie = {"fc": df_fc, "pv": df_pv, "pc": df_pc, "pf": df_pf_all}
                otras_series = [k for k in series_sel if k not in ("real", "real_filtrado")]
                if otras_series:
                    codigos_otras = set()
                    for k in otras_series:
                        df_fuente = fuente_por_serie.get(k)
                        if df_fuente is not None and not df_fuente.empty and "cve_prod" in df_fuente.columns:
                            codigos_otras |= set(df_fuente["cve_prod"].astype(str).str.strip())
                    df_comp = df_comp[df_comp["cve_prod"].astype(str).str.strip().isin(codigos_otras)]
                    if df_comp.empty:
                        st.info(f"sin productos coincidentes con las series elegidas en {label}")
                        continue

            df_comp = _agregar_comparaciones(df_comp, baseline, otros, meses)
            unidad = "kg" if seccion == "KG" else "USD"

            # ── KPIs del total ─────────────────────────────────────────────
            # incluye "real_compras" aunque no esté en series_sel: se necesita
            # como referencia interna al comparar contra Presupuesto Compras
            # (ver _referencia_para)
            claves_totales = set(series_sel) | {"real_compras"}
            totales_sum = {
                k: df_comp[_SERIES_META[k]["total_col"]].sum()
                for k in claves_totales if _SERIES_META[k]["total_col"] in df_comp.columns
            }

            cols_tot = st.columns(len(series_sel))
            for col, k in zip(cols_tot, series_sel):
                col.metric(f"{_SERIES_META[k]['label']} {unidad}", f"{totales_sum[k]:,.0f}")

            if otros:
                cols_cumpl = st.columns(len(otros))
                cols_delta = st.columns(len(otros))
                for col_c, col_d, otro in zip(cols_cumpl, cols_delta, otros):
                    referencia = _referencia_para(otro, baseline)
                    referencia_label = _SERIES_META[referencia]["label"]
                    otro_label = _SERIES_META[otro]["label"]
                    num, den = (totales_sum[referencia], totales_sum[otro]) if baseline == "real" \
                        else (totales_sum[otro], totales_sum[referencia])
                    cumpl = round(num / den * 100, 1) if den else 0.0
                    col_c.metric(
                        f"Cumplimiento {otro_label} (vs {referencia_label})", f"{cumpl:.1f}%",
                        delta=f"{cumpl - 100:.1f}pp vs plan", delta_color="normal",
                    )
                    col_d.metric(
                        f"Diferencia {otro_label} (vs {referencia_label})", f"{num - den:+,.0f}",
                        delta_color="normal",
                    )

            st.divider()

            # ── tabla comparativa ──────────────────────────────────────────
            fmt_valor = "localized"  # sin decimales, con separador de miles

            col_cfg: dict = {
                "cve_prod": st.column_config.TextColumn("cve prod", disabled=True, width="small"),
                "producto": st.column_config.TextColumn("producto", disabled=True),
            }
            encabezados: dict[str, str] = {"cve_prod": "Cve prod", "producto": "Producto"}
            cols_orden = ["cve_prod", "producto"]

            for k in series_sel:
                total_col = _SERIES_META[k]["total_col"]
                titulo = f"total {_SERIES_META[k]['label'].lower()} {unidad}"
                col_cfg[total_col] = st.column_config.NumberColumn(titulo, format=fmt_valor, disabled=True)
                encabezados[total_col] = titulo
                cols_orden.append(total_col)

            for otro in otros:
                referencia_label = _SERIES_META[_referencia_para(otro, baseline)]["label"]
                ccol = f"total_cumpl_{otro}_%"
                titulo = f"cumplimiento % {_SERIES_META[otro]['label']} (vs {referencia_label})"
                col_cfg[ccol] = st.column_config.NumberColumn(titulo, format="%.1f", disabled=True)
                encabezados[ccol] = titulo
                cols_orden.append(ccol)

            for mes in meses:
                mn = _MESES[mes]
                for k in series_sel:
                    c = f"{mn}{_SERIES_META[k]['mes_suffix']}"
                    titulo = f"{mn.upper()} {_SERIES_META[k]['label'].lower()}"
                    col_cfg[c] = st.column_config.NumberColumn(titulo, format=fmt_valor, disabled=True)
                    encabezados[c] = titulo
                    cols_orden.append(c)
                for otro in otros:
                    referencia_label = _SERIES_META[_referencia_para(otro, baseline)]["label"]
                    c = f"{mn}_delta_{otro}_%"
                    titulo = f"{mn.upper()} Δ% {_SERIES_META[otro]['label']} (vs {referencia_label})"
                    col_cfg[c] = st.column_config.NumberColumn(titulo, format="%.1f", disabled=True)
                    encabezados[c] = titulo
                    cols_orden.append(c)

            df_show = df_comp[[c for c in cols_orden if c in df_comp.columns]].copy()

            # columnas de valores (kg/USD): redondear a entero para que "localized" no muestre decimales
            cols_valor = {_SERIES_META[k]["total_col"] for k in series_sel} | {
                f"{_MESES[m]}{_SERIES_META[k]['mes_suffix']}" for m in meses for k in series_sel
            }
            for c in cols_valor & set(df_show.columns):
                df_show[c] = pd.to_numeric(df_show[c], errors="coerce").round(0)

            styled = _style_delta(df_show)
            st.dataframe(
                styled,
                use_container_width=True,
                hide_index=True,
                column_config=col_cfg,
                height=min(56 + len(df_show) * 35, 680),
            )

            # ── gráfica de las series elegidas, por mes ──────────────────────
            with st.expander("📊 gráfica por mes"):
                datos_grafica: list[dict] = []
                for mes in meses:
                    mn = _MESES[mes]
                    fila_g = {"mes": mn.upper()}
                    for k in series_sel:
                        c = f"{mn}{_SERIES_META[k]['mes_suffix']}"
                        fila_g[_SERIES_META[k]["label"]] = df_comp[c].sum() if c in df_comp.columns else 0
                    datos_grafica.append(fila_g)
                df_g = pd.DataFrame(datos_grafica).set_index("mes")
                cols_chart = [_SERIES_META[k]["label"] for k in series_sel]
                st.bar_chart(df_g[cols_chart], use_container_width=True, height=300)

            # ── descarga ──────────────────────────────────────────────────
            df_export = df_show.rename(columns=encabezados)
            csv = df_export.to_csv(index=False).encode("utf-8")
            col_dl1, col_dl2 = st.columns(2)
            with col_dl1:
                st.download_button(
                    f"⬇️ CSV {label}",
                    data=csv,
                    file_name=f"real_vs_forecast_{seccion}_{region}_{anio_sel}_v{id_version}.csv",
                    mime="text/csv",
                    key=f"rvf_dl_{seccion}_{region}",
                    use_container_width=True,
                )
            with col_dl2:
                xlsx = _comparativo_a_excel_bytes(label, df_show, encabezados)
                st.download_button(
                    f"⬇️ Excel {label}",
                    data=xlsx,
                    file_name=f"real_vs_forecast_{seccion}_{region}_{anio_sel}_v{id_version}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"rvf_dl_xlsx_{seccion}_{region}",
                    use_container_width=True,
                )
